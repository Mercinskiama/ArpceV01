# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import os, codecs

from django.contrib.contenttypes.models import ContentType
from ErpBackOffice.utils.auth import auth
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpRequest, HttpResponseRedirect, JsonResponse
from django.template.response import SimpleTemplateResponse, TemplateResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User, Group
from django.db import transaction
from django.template import loader
from django.views import generic
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.utils import timezone
from django.core import serializers
from random import randint
from django.core.mail import send_mail
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from ErpBackOffice.dao.dao_place import dao_place
from ErpBackOffice.dao.dao_utilisateur import dao_utilisateur
from ErpBackOffice.dao.dao_devise import dao_devise
from ErpBackOffice.dao.dao_module import dao_module
from ErpBackOffice.dao.dao_sous_module import dao_sous_module
#from ErpBackOffice.dao.dao_action_utilisateur import dao_action_utilisateur
from ErpBackOffice.utils.identite import identite
from ErpBackOffice.utils.tools import ErpModule
from ErpBackOffice.models import Model_Personne,Model_LigneRegle
from ErpBackOffice.models import Model_ModuleOverModel
from ErpBackOffice.models import Model_Module
from ErpBackOffice.dao.dao_moduleovermodel import dao_moduleovermodel
from ErpBackOffice.dao.dao_wkf_workflow import dao_wkf_workflow
from ErpBackOffice.dao.dao_wkf_etape import dao_wkf_etape
from ErpBackOffice.dao.dao_wkf_condition import dao_wkf_condition
from ErpBackOffice.dao.dao_wkf_transition import dao_wkf_transition
from ErpBackOffice.dao.dao_personne import dao_personne
from ErpBackOffice.dao.dao_constante import dao_constante
from ErpBackOffice.dao.dao_sous_module import dao_sous_module
from ErpBackOffice.dao.dao_groupe_permission import dao_groupe_permission
from ErpBackOffice.dao.dao_groupe_menu import dao_groupe_menu
from ErpBackOffice.dao.dao_model import dao_model
from ErpBackOffice.dao.dao_regle import dao_regle
from ModuleConfiguration.dao.dao_permission import dao_permission
from ModuleConfiguration.dao.dao_actionutilisateur import dao_actionutilisateur
from ModuleConversation.dao.dao_notification import dao_notification
from ErpBackOffice.dao.dao_organisation import dao_organisation
from ModuleConfiguration.dao.dao_sousmodule import dao_sousmodule
from ModuleConfiguration.dao.dao_groupemenu import dao_groupemenu
import pandas as pd
from rest_framework.decorators import api_view
from ErpBackOffice.utils.print import weasy_print
from ErpBackOffice.utils.wkf_task import wkf_task
from ErpBackOffice.dao.dao_query_builder import dao_query_builder
from ErpBackOffice.utils.separateur import makeFloat, makeStringFromFloatExcel, makeInt, makeIntId, makeString, checkDateTimeFormat, checkDateFormat

from ModuleConfiguration.dao.dao_utilisateur_cof import *
from ErpBackOffice.utils.generateur import *
from ModuleConfiguration.models import *
#Pagination
from ErpBackOffice.utils.pagination import pagination
from ErpBackOffice.utils.utils import utils
from subprocess import call,run,Popen
from django.core.management import call_command
import os.path
import datetime, inspect, logging
import json, traceback, array, time, random, string, unidecode

monLog = logging.getLogger('logger')
var_module_id = 5 #ID du Module en cours
module = "ModuleConfiguration"
vars_module = {"name" : "MODULE_CONFIGURATION", "value" : 99 }



# Create your views here.
# TABLEAU DE BORD CONTROLLER
def get_dashboard(request):
    try:
        modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetDashboardAuthentification(5, request)

        if response != None:
            return response
        
        #utilisateurs
        users = dao_utilisateur.toListUtilisateur()
        last_users = users[:5]
        total_users = len(users)
        total_actif_users = dao_utilisateur.toListUtilisateursActifs().count()
        total_inactif_users = dao_utilisateur.toListUtilisateursInActifs().count()
        #utilisateurs connectés
        users_connected = sorted(dao_utilisateur.toListUtilisateur()[:5], key=lambda t: t.is_connected, reverse=True) #Filtrage des resultats sur une clé property
        all_users_connected = dao_utilisateur_cof.get_connected_users()
        number_users_connected = len(all_users_connected)
        number_users = dao_utilisateur_cof.toGetNombre().count()
        #Liste de derniers connectés
        nbre_employes_login = [2,1,1,0,1,5,1,0,4,4,1,8]
            
        #Rôles
        roles = dao_groupe_permission.toListGroupePermissions()
        last_roles = roles[:5]
        total_roles = len(roles)
        
        #Permissions
        permissions = dao_permission.toListPermission()
        last_permissions = permissions[:5]
        total_permissions = len(permissions)
        
        #Actions
        actions = dao_actionutilisateur.toListActionutilisateur() 
        last_actions = actions[:5]
        total_actions = len(actions)

        #Règles
        regles = dao_regle.toListRegles()
        last_regles = regles[:5]
        total_regles = len(regles)  
        
        #Modules
        applications = dao_module.toListModules()
        last_applications = applications[:5]
        total_applications = len(applications)
        
        #Sous-modules
        menus = dao_sousmodule.toListSousmodule()
        last_menus = menus[:5]
        total_menus = len(menus)
        
        menus_actifs = dao_sousmodule.toListSousmoduleActif()
        last_menus_actifs = menus_actifs[:5]
        total_menus_actifs = len(menus_actifs)
        
        #Groupe-menu
        groupemenus = dao_groupemenu.toListGroupemenu()
        last_groupemenus = groupemenus[:5]
        total_groupemenus = len(groupemenus)

        #WorkFlow
        workflows = dao_wkf_workflow.toListWorkflows()
        last_workflows = workflows[:5]
        total_workflows = len(workflows)
        
        wkf_etapes = dao_wkf_etape.toListEtapeWorkflows()
        last_wkf_etapes = wkf_etapes[:5]
        total_wkf_etapes = len(wkf_etapes)
        
        wkf_transitions = dao_wkf_transition.toListTransitions()
        last_wkf_transitions = wkf_transitions[:5]
        total_wkf_transitions = len(wkf_transitions)
        
        #notifications
        notifications = dao_notification.toListNotification()[:5]

        context = {
            'title' : 'Tableau de bord',
            "utilisateur" : utilisateur,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            "modules" : modules,
        	"module" : vars_module,
            "sous_modules":sous_modules,
            "module" : vars_module,
            'menu' : 0,
            "users_connected":users_connected,
            'all_users':all_users_connected,
            'toGetNombre':number_users,
            'number_users_connected':number_users_connected,
            'users' : users,
            'last_users' : last_users,
            'total_users' : total_users,
            "total_inactif_users" : total_inactif_users,
            "total_actif_users" : total_actif_users,
            "notifications": notifications,
            "nbre_employes_login": nbre_employes_login,
            'roles' : roles,
            'last_roles' : last_roles,
            'total_roles' : total_roles,
            'permissions' : permissions,
            'last_permissions' : last_permissions,
            'total_permissions' : total_permissions,
            'actions' : actions,
            'last_actions' : last_actions,
            'total_actions' : total_actions,
            'regles' : regles,
            'last_regles' : last_regles,
            'total_regles' : total_regles,
            'applications' : applications,
            'last_applications' : last_applications,
            'total_applications' : total_applications,
            'menus' : menus,
            'last_menus' : last_menus,
            'total_menus' : total_menus,
            'menus_actifs' : menus_actifs,
            'last_menus_actifs' : last_menus_actifs,
            'total_menus_actifs' : total_menus_actifs,
            'groupemenus' : groupemenus,
            'last_groupemenus' : last_groupemenus,
            'total_groupemenus' : total_groupemenus,
            'workflows' : workflows,
            'last_workflows' : last_workflows,
            'total_workflows' : total_workflows,
            'wkf_etapes' : wkf_etapes,
            'last_wkf_etapes' : last_wkf_etapes,
            'total_wkf_etapes' : total_wkf_etapes,
            'wkf_transitions' : wkf_transitions,
            'last_wkf_transitions' : last_wkf_transitions,
            'total_wkf_transitions' : total_wkf_transitions,
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/index.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        return auth.toReturnFailed(request, e, traceback.format_exc(), reverse("backoffice_index"))


# UTILISATEURS CONTROLLERS
def get_lister_utilisateurs(request):
    # modules, utilisateur,response = auth.toGetAuth(request)
    permission_number = 546
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    # model = dao_utilisateur.toListUtilisateur()
    #*******Filtre sur les règles **********#
    model = dao_model.toListModel(dao_utilisateur.toListUtilisateur(), permission_number, groupe_permissions, identite.utilisateur(request))
	#******* End Regle *******************#
    context = {
		'title' : 'Liste des utilisateurs',
        'model' : model,
        'sous_modules': sous_modules,
        "utilisateur" : utilisateur,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        "modules" : modules,
        "module" : ErpModule.MODULE_CONFIGURATION,
        'menu' : 1,
        'degrade': 'module_configuration'
	}
    template = loader.get_template("ErpProject/ModuleConfiguration/utilisateur/list.html")
    return HttpResponse(template.render(context, request))

def get_creer_utilisateur(request):
    permission_number = 547
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    context = {
        'title' : 'Nouvel utilisateur',
        "utilisateur" : utilisateur,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        "modules" :modules,
        "sous_modules": sous_modules,
        "module" : ErpModule.MODULE_CONFIGURATION,
        'menu' : 1
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/utilisateur/add.html")
    return HttpResponse(template.render(context, request))

@transaction.atomic
def post_creer_utilisateur(request):
    sid = transaction.savepoint()
    try:
        auteur =  identite.utilisateur(request)
        nom_complet = request.POST["nom_complet"]
        est_particulier = est_particulier = True
        email = request.POST["email"]
        phone = request.POST["phone"]
        image = ""

        personne = dao_personne.toCreateEmploye("","","",nom_complet, image, email , phone,"", 3, True)
        #utilisateur = dao_utilisateur.toSaveUtilisateur(personne)
        utilisateur = dao_personne.toSaveEmploye(auteur, personne)
        if utilisateur != None :
            messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
            transaction.savepoint_commit(sid)
            return HttpResponseRedirect(reverse('module_configuration_details_utilisateur', args=(utilisateur.id,)))
        else :
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue lors de la création de l'utilisateur")
            return HttpResponseRedirect(reverse('module_configuration_add_utilisateur'))
    except Exception as e:
        print("ERREUR !")
        print(e)
        transaction.savepoint_rollback(sid)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_add_utilisateur'))

def get_modifier_utilisateur(request, ref):
    permission_number = 548
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        id = int(ref)
        utilisateur = dao_utilisateur.toGetUtilisateur(id)
        context = {
            'title' : 'Modifier %s' % utilisateur.nom_complet,
            'model' : utilisateur,
            "utilisateur" : utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'sous_modules': sous_modules,
            "modules" : modules,
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 1
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/utilisateur/update.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_utilisateurs'))

def post_modifier_utilisateur(request):
    id = int(request.POST["ref"])
    try:
        auteur =  identite.utilisateur(request)
        nom_complet = request.POST["nom_complet"]
        est_particulier = True
        email = request.POST["email"]
        phone = request.POST["phone"]
        adresse = request.POST["adresse"]
        commune_quartier_id = int(request.POST["commune_quartier_id"])
        image= ""

        personne = dao_personne.toGetEmploye(id)
        personne.nom_complet = nom_complet
        personne.email = email
        personne.user.username = email
        personne.phone = phone
        personne.adresse = adresse
        personne.commune_quartier_id = commune_quartier_id
        #personne = dao_personne.toCreateEmploye("","","",nom_complet, image, email , phone,"", 3, 0, True,"","2019-12-31","","2019-12-31",None,None,None,True,None,None)
        #is_done = dao_utilisateur.toUpdateUtilisateur(id, personne)
        #is_done = dao_personne.toUpdateEmploye(id,personne)
        personne.save()
        personne.user.save()

        #if is_done == True :
        if personne != None :
            return HttpResponseRedirect(reverse('module_configuration_details_utilisateur', args=(id,)))
        else :
            messages.add_message(request, messages.ERROR, "Une erreur est survenue lors de la mise à jour des informations de l'utilisateur")
            return HttpResponseRedirect(reverse('module_configuration_update_utilisateur', args=(id,)))
    except Exception as e:
        #print("ERREUR !")
        #print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_update_utilisateur', args=(id,)))

def get_details_utilisateur(request, ref):
    permission_number = 546
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    
    if response != None:
        return response
    try:
        ref = int(ref)
        model = dao_utilisateur.toGetUtilisateur(ref)

        context = {
            'title' : model.nom_complet,
            'model' : model,
            'roles' : dao_groupe_permission.toListGroupePermissionDeLaPersonne(model.id),
            'utilisateur' : utilisateur,
            'actions':[],
            'sous_modules': sous_modules,
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'modules' : modules,
            'roles_modules' : [],
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 1
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/utilisateur/item.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_utilisateurs'))

# ROLES CONTROLLERS
def get_lister_roles(request):
    permission_number = 550
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response
    try:

        model = dao_groupe_permission.toListGroupePermissions()

        #print("model", model)
        try:
            view = str(request.GET.get("view","list"))
        except Exception as e:
            view = "list"

        #Pagination
        model = pagination.toGet(request, model)


        context = {
            'title' : 'Liste des roles',
            'model' : model,
            'view' : view,
            "utilisateur" : utilisateur,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            "modules" :modules,
            'sous_modules': sous_modules,
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 2
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/role/list.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_dashboard"))


def get_details_role(request, ref):
    permission_number = 550
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response
    try:
        ref = int(ref)
        model = dao_groupe_permission.toGetGroupePermission(ref)

        #print("Groupe %s" % model)
        modules = dao_module.toListModulesByPermission(model)
        #print("Modules %s" % modules)

        sous_modules_list = []
        actions = []

        for item in modules:
            sous_modules_list.extend(dao_sous_module.toListSousModulesByGroupePermission(item, model))

        for item in sous_modules_list:
            #actions.extend(dao_droit.toListDroitAutroses(item.id,model.designation))
            actions.extend(model.permissions.filter(sous_module_id = item.id))


        #print("Sous Module %s" % sous_modules_list)

        utilisateurs = dao_utilisateur.toListUtilisateursDuRole(model.id)
        role = dao_groupe_permission.toGetGroupePermission(ref)
        used_role = Model_GroupePermissionUtilisateur.objects.filter(groupe_permission = role.id)

        print('utilisateur role:', utilisateurs)
        context = {
            'title' : model.designation,
            'model' : model,
            'modules_du_role' : modules,
            'sous_modules_list': sous_modules_list,
            'sous_modules': sous_modules,
            'actions_utilisateur': actions,
            'modules' : dao_module.toListModulesInstalles(),
            "module" : ErpModule.MODULE_CONFIGURATION,
            'utilisateurs': used_role,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'utilisateur': utilisateur,
            'menu' : 2
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/role/item.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_list_roles"))


def get_creer_role(request):
    permission_number = 551
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)


    if response != None:
        return response
    try:
        sous_modules_list = []
        actions_utilisateur = []

        for item in modules:
            sous_modules_list.extend(dao_sous_module.toListSousModulesOf(item.id))

        for item in sous_modules_list:
            #actions_utilisateur.extend(dao_droit.toListDroitNonAutroses(item.id,'role_nouveau'))
            actions_utilisateur.extend(dao_permission.toListPermissionsOfSousModule(item.id))


        context = {
            'title' : 'Nouveau role',
            'utilisateur' : utilisateur,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'modules' : modules,
            'sous_modules_list' : sous_modules_list,
            'sous_modules': sous_modules,
            'actions_utilisateur' : actions_utilisateur,
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 2
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/role/add.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_list_roles"))

@transaction.atomic
def post_creer_role(request):
    sid = transaction.savepoint()
    try:
        #print(request.POST)
        erreur_survenue = False
        nom_role = request.POST["nom_role"]
        auteur = identite.utilisateur(request)
        role = dao_groupe_permission.toCreateGroupePermission(nom_role)
        role = dao_groupe_permission.toSaveGroupePermission(auteur, role)
        #print("what is the role", role)

        list_module_id = request.POST.getlist('module_id', None)
        list_sous_module_id = request.POST.getlist('sous_module_id', None)
        # list_permissions_id = request.POST.getlist('action_id', None)
        list_permissions_id = request.POST.getlist('my_multi_select2[]', None)

        for i in range(0, len(list_permissions_id)):
            permission_id = int(list_permissions_id[i])
            is_done = dao_groupe_permission.toAddPermission(role, permission_id)
            if is_done == False:
                erreur_survenue = True
                break

        '''if erreur_survenue == True:
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue au moment de la création du role")
            return HttpResponseRedirect(reverse("module_configuration_add_role"))

        for i in range(0, len(list_sous_module_id)):
            sous_module_id = int(list_sous_module_id[i])
            is_done = dao_role.toAddSousModuleInRole(auteur, sous_module_id, role.id)

            if is_done == False:
                erreur_survenue = True
                break

        if erreur_survenue == True:
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue au moment de la création du role")
            return HttpResponseRedirect(reverse("module_configuration_add_role"))

        for i in range(0, len(list_action_id)):
            droit_id = int(list_action_id[i])
            is_done = dao_droit.toAddDroit(droit_id, role.id)

            if is_done == False:
                erreur_survenue = True
                break'''

        if erreur_survenue == True:
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue au moment de la création du role")
            return HttpResponseRedirect(reverse("module_configuration_add_role"))
        else:
            messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
            transaction.savepoint_commit(sid)
            return HttpResponseRedirect(reverse("module_configuration_details_role", args=(role.id,)))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        transaction.savepoint_rollback(sid)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_add_role"))


def get_modifier_role(request, ref):
    permission_number = 552
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        ref = int(ref)
        model = dao_groupe_permission.toGetGroupePermission(ref)
        modules = dao_module.toListModulesByPermission(model)
        sous_modules_list = []
        actions = []
        request.session['nom_role'] = model.designation

        for item in modules:
            sous_modules_list.extend(dao_sous_module.toListSousModulesOf(item.id))

        for item in sous_modules_list:
            #actions_utilisateur.extend(dao_droit.toListDroitNonAutroses(item.id,'role_nouveau'))
            actions.extend(dao_permission.toListPermissionsOfSousModule(item.id))

        context = {
            'title' : 'Role %s' % model.designation,
            'model' : model,
            'modules_du_role' : modules,
            'modules' : dao_module.toListModulesInstalles(),
            "module" : ErpModule.MODULE_CONFIGURATION,
            'sous_modules_list': sous_modules_list,
            'sous_modules': sous_modules,
            'actions_utilisateur': actions,
            'utilisateur' : utilisateur,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'menu' : 2
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/role/update.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_list_roles"))


@transaction.atomic
def post_modifier_role(request):
    ref= int(request.POST["ref"])
    sid = transaction.savepoint()

    try:
        erreur_survenue = False
        nom_role = request.POST["nom_role"]
        groupe_permission = dao_groupe_permission.toCreateGroupePermission(nom_role)
        is_done = dao_groupe_permission.toUpdateGroupePermission(ref, groupe_permission)

        if is_done ==  False:
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue au moment de la mise à jour du role")
            return HttpResponseRedirect(reverse("module_configuration_update_role", args=(ref,)))
        else:
            transaction.savepoint_commit(sid)
            return HttpResponseRedirect(reverse("module_configuration_details_role", args=(ref,)))

    except Exception as e:
        #print("ERREUR POST MODIFIER ROLE")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_update_role", args=(ref,)))

def get_ajouter_droits(request, ref):
    permission_number = 552
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        ref = int(ref)
        model = dao_groupe_permission.toGetGroupePermission(ref)

        sous_modules_list = []
        actions_non_autorisees = []
        modules_installes = dao_module.toListModulesInstalles()

        #print("MODULES %s" % modules_installes)

        for item in modules_installes:
            sous_modules_list.extend(dao_sous_module.toListSousModulesOf(item.id))


        for item in sous_modules_list:
            #actions_non_autorisees.extend(dao_droit.toListDroitNonAutroses(item.id,model.nom_role))
            actions_non_autorisees.extend(dao_permission.toListPermissionsNonAutorizeOfSousModule(model.id, item.id))
            #print("SOUS MOD %s" % item.id)
            #print("print %s" % model.permissions.exclude(sous_module_id = item.id))

        #print("ACTION")
        #print(actions_non_autorisees)

        context = {
            'title' : 'Ajout des droits au role %s' % model.designation,
            'model' : model,
            'modules_installes' : modules_installes,
            'sous_modules_list': sous_modules_list,
            'sous_modules': sous_modules,
            'actions_non_autorisees': actions_non_autorisees,
            'modules' : modules,
            "module" : ErpModule.MODULE_CONFIGURATION,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'utilisateur': utilisateur,
            'menu' : 2
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/droit/add.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_list_roles"))

@transaction.atomic
def post_ajouter_droits(request):
    ref= int(request.POST["ref"])
    sid = transaction.savepoint()

    try:
        role = dao_groupe_permission.toGetGroupePermission(ref)
        auteur = identite.utilisateur(request)
        erreur_survenue = False
        list_module_id = request.POST.getlist('module_id', None)
        list_sous_module_id = request.POST.getlist('sous_module_id', None)
        # list_permissions_id = request.POST.getlist('action_id', None)
        list_permissions_id = request.POST.getlist('my_multi_select2[]', None)

        for i in range(0, len(list_permissions_id)):
            permission_id = int(list_permissions_id[i])

            is_done = dao_groupe_permission.toAddPermission(role, permission_id)
            if is_done == False:
                erreur_survenue = True
                break

        if erreur_survenue == True:
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue au moment de l'ajout des droits au role")
            return HttpResponseRedirect(reverse("module_configuration_add_rights", args=(ref,)))
        else:
            transaction.savepoint_commit(sid)
            messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
            return HttpResponseRedirect(reverse("module_configuration_details_role", args=(ref,)))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        transaction.savepoint_rollback(sid)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_add_rights", args=(ref,)))

def get_retirer_droits(request, ref):
    permission_number = 552
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        ref = int(ref)
        model = dao_groupe_permission.toGetGroupePermission(ref)
        modules_du_role = dao_module.toListModulesByPermission(model)
        sous_modules_du_role = []
        actions_du_role = []

        for item in modules:
            sous_modules_du_role.extend(dao_sous_module.toListSousModulesByGroupePermission(item, model))

        for item in sous_modules_du_role:
            #actions.extend(dao_droit.toListDroitAutroses(item.id,model.designation))
            actions_du_role.extend(model.permissions.filter(sous_module_id = item.id))

        context = {

            'title' : 'Retirer des droits au role %s' % model.designation,
            'model' : model,
            'modules_installes' : modules_du_role,
            'sous_modules_list': sous_modules_du_role,
            'sous_modules': sous_modules,
            'actions_non_autorisees': actions_du_role,
            'modules' : modules,
            "module" : ErpModule.MODULE_CONFIGURATION,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'utilisateur': utilisateur,
            'menu' : 2,
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/droit/remove.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_list_roles"))

@transaction.atomic
def post_retirer_droits(request):
    ref= int(request.POST["ref"])
    sid = transaction.savepoint()

    try:
        role = dao_groupe_permission.toGetGroupePermission(ref)
        auteur = identite.utilisateur(request)
        erreur_survenue = False
        list_module_id = request.POST.getlist('module_id', None)
        list_sous_module_id = request.POST.getlist('sous_module_id', None)
        # list_permissions_id = request.POST.getlist('action_id', None)
        list_permissions_id = request.POST.getlist('my_multi_select2[]', None)

        for i in range(0, len(list_permissions_id)):
            permission_id = int(list_permissions_id[i])

            is_done = dao_groupe_permission.toRemovePermission(role, permission_id)
            if is_done == False:
                erreur_survenue = True
                break

        if erreur_survenue == True:
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue au moment de la mise à jour du role")
            return HttpResponseRedirect(reverse("module_configuration_remove_rights", args=(ref,)))
        else:
            transaction.savepoint_commit(sid)
            return HttpResponseRedirect(reverse("module_configuration_details_role", args=(role.id,)))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        transaction.savepoint_rollback(sid)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_remove_rights", args=(ref,)))

def get_attribuer_role(request, ref):
    permission_number = 548
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    if response != None:
        return response

    try:
        ref = int(ref)
        utilisateur = dao_utilisateur.toGetUtilisateur(ref)

        roles = dao_groupe_permission.toListGroupePermissions()

        #print("ROLES %s" % roles)

        context = {
            'title' : "Attribuer un nouveau rôle",
            'model' : utilisateur,
            'roles' : roles ,
            "utilisateur" : utilisateur,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            "modules" : modules,
            'sous_modules': sous_modules,
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 1
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/role/attribute.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_utilisateurs'))

def post_attribuer_role(request):
    ref = int(request.POST["ref"])
    try:
        utilisateur = dao_utilisateur.toGetUtilisateur(ref)
        role_id = int(request.POST["role_id"])
        auteur = identite.utilisateur(request)
        is_done = dao_groupe_permission.toAttributeGroupePermission(auteur, utilisateur.id, role_id)

        if is_done == True:
            return HttpResponseRedirect(reverse('module_configuration_details_utilisateur', args=(ref,)))
        else:
            messages.add_message(request, messages.ERROR, "Une erreur est survenue au moment de l'attribution du role à l'utilisateur")
            return HttpResponseRedirect(reverse("module_configuration_attribuer_role", args=(ref,)))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_attribuer_role", args=(ref,)))

def get_retirer_role(request, ref_utilisateur, ref_role):
    try:
        ref_utilisateur = int(ref_utilisateur)
        utilisateur = dao_utilisateur.toGetUtilisateur(ref_utilisateur)

        ref_role = int(ref_role)
        is_done = dao_groupe_permission.toRetireGroupePermission(utilisateur.id, ref_role)

        if is_done == True:
            return HttpResponseRedirect(reverse('module_configuration_details_utilisateur', args=(ref_utilisateur,)))
        else:
            messages.add_message(request, messages.ERROR, "Une erreur est survenue au moment de l'attribution du role à l'utilisateur")
            return HttpResponseRedirect(reverse("module_configuration_details_utilisateur", args=(ref_utilisateur,)))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_utilisateurs'))

# PLACE CONTROLLERS
def get_lister_devises(request):
    pass

def get_json_list_places_filles(request):
    try:
        data = []
        parent_id = int(request.GET["ref"])
        places = dao_place.toListPlacesFilles(parent_id)
        for place in places:
            item = {
                "id" : place.id,
                "designation" : place.designation,
                "code_telephone" : place.code_telephone,
                "code_pays" : place.code_pays,
                "place_type" : place.place_type
            }
            data.append(item)
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse([], safe=False)

#CONFIGURATION CONTROLLER
def get_configuration(request):
    permission_number = 570
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        #ref = int(ref)
        #condition_reglement = dao_condition_reglement.toGetConditionReglement(ref)
        organisation = dao_organisation.toGetMainOrganisation()
        context = {
            'title' : 'Paramètres généraux' ,
            'organisation' : organisation,
            "utilisateur" : utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'sous_modules': sous_modules,
            "modules" : modules,
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 11
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/configuration/index.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_configuration'))

def post_modifier_configuration(request):
    id = int(request.POST["ref"])
    try:
        auteur = identite.utilisateur(request)
        designation = request.POST["designation"]

        categorie = None
        is_done = False

        if is_done == True : return HttpResponseRedirect(reverse("module_configuration_configuration", args=(id,)))
        else : return HttpResponseRedirect(reverse('module_configuration_configuration', args=(id,)))
    except Exception as e:
        #print("ERREUR !")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_configuration', args=(id,)))

def get_json_employee(request):

        employes = Model_Personne.objects.filter(type = "UTILISATEUR").filter(est_employe = True)

        data = []
        for employe in employes:
            employe_json = {}
            employe_json['id']= employe.id
            employe_json['nom_complet'] = employe.nom_complet
            data.append(employe_json)
        return JsonResponse(data, safe=False)


# MODULES CONTROLLERS
def get_lister_modules(request):
    try:
        permission_number = 554
        modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

        if response != None:
            return response
        
        #*******Filtre sur les règles **********#
        model = dao_model.toListModel(dao_module.toListModules(), permission_number, groupe_permissions, identite.utilisateur(request))
        #******* End Regle *******************#
        try:
            view = str(request.GET.get("view","list"))
        except Exception as e:
            view = "list"
        model = pagination.toGet(request, model)
    
        context = {
            'title' : 'Liste des modules',
            'model' : model,
            'view' : view,
            "utilisateur" : utilisateur,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            "modules" : modules,
            "module" : vars_module,
            'sous_modules': sous_modules,           
            'menu' : 3
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/module/list.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_dashboard'))

def get_creer_module(request):
    try:
        permission_number = 555
        modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

        if response != None:
            return response
        context = {
            'title' : 'Nouveau module',
            'utilisateur' : utilisateur,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'modules' : modules,
            "module" : vars_module,
            'sous_modules': sous_modules,            
            'menu' : 3
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/module/add.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_creer_module(request):
    sid = transaction.savepoint()
    try:
        nom_module = request.POST["nom_module"]
        description = request.POST["description"]
        url_vers=request.POST["url_vers"]
        numero_ordre = request.POST["numero_ordre"]
        icon_module = request.POST["icon_module"]
        couleur = request.POST["couleur"]
        model_dashboard = int(request.POST["model_dashboard"])
        est_installe = False
        if "est_installe" in request.POST : est_installe = True
        auteur = identite.utilisateur(request)

        module_object = dao_module.toCreateModule(nom_module, description, url_vers, numero_ordre, icon_module, couleur, est_installe)
        module_object = dao_module.toSaveModule(module_object)
        print("Module creee")

        #On génère les différents fichiers du module
        parametrage_module(module_object, request, model_dashboard)

        transaction.savepoint_commit(sid)             
        messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
        return HttpResponseRedirect(reverse('module_configuration_details_module', args=(module_object.id,)))
    except Exception as e:
        transaction.savepoint_rollback(sid)  
        return auth.toReturnFailed(request, e, traceback.format_exc())


def get_details_module(request, ref):
    permission_number = 554
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response
    try:
        ref = int(ref)
        model = dao_module.toGetModule(ref)
        modeles = ContentType.objects.filter(app_label = model.nom_application)
        groupesmenus = dao_groupe_menu.toListGroupeOfModule(ref)
        menus = dao_sous_module.toListSousModulesOfModule(ref)

        context = {
            'title' : 'Détails du Module ' + model.nom_module,
            'model' : model,
            'modeles': modeles,
            'modules' : modules,
            "module" : vars_module,
            'groupesmenus': groupesmenus,
            'menus' : menus,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'sous_modules': sous_modules, 
            'utilisateur': utilisateur,
            'menu' : 3
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/module/item.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_list_modules"))

def get_modifier_module(request, ref):
    permission_number = 556
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        id = int(ref)
        module = dao_module.toGetModule(id)
        context = {
            'title' : 'Modifier %s' % module.nom_module,
            'model' : module,
            "utilisateur" : utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'sous_modules': sous_modules,
            "modules" : modules,
            "module" : vars_module,           
            'menu' : 3
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/module/update.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR POUL")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_modules'))

def post_modifier_module(request):
    id = int(request.POST["ref"])
    try:
        nom_module = request.POST["nom_module"]
        description = request.POST["description"]
        url_vers=request.POST["url_vers"]
        numero_ordre = request.POST["numero_ordre"]
        icon_module = request.POST["icon_module"]
        couleur = request.POST["couleur"]
        est_installe = False
        if "est_installe" in request.POST : est_installe = True        

        module = dao_module.toCreateModule(nom_module, description, url_vers, numero_ordre, icon_module,couleur, est_installe)
        is_done = dao_module.toUpdateModule(id,module)

        if is_done == True :
            return HttpResponseRedirect(reverse('module_configuration_details_module', args=(id,)))
        else :
            messages.add_message(request, messages.ERROR, "Une erreur est survenue lors de la mise à jour des informations du module")
            return HttpResponseRedirect(reverse('module_configuration_update_module', args=(id,)))
    except Exception as e:
        #print("ERREUR !")
        #print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_update_module', args=(id,)))

def parametrage_module(module, request, model_dashboard = 1):
    try:
        print("Debut parametrage_module")
        nomModule = module.nom_application
        nameModuleUp = 'MODULE_{0}'.format(unidecode.unidecode(module.nom_module.upper().replace(" ","_")))
        nom_pattern = 'module_{0}'.format(unidecode.unidecode(module.nom_module.lower().replace(" ","_")))
        auteur = identite.utilisateur(request)        

        #Creation du Module via invite de commande
        call(["python","manage.py","startapp",nomModule])
        print("cmd startapp lancer")


        #Manipulation de views.py dans  le sous-dossier de l'application
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\views.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_views_py_dossier_application = '# -*- coding: utf-8 -*-\nfrom __future__ import unicode_literals\nimport traceback\nfrom django.shortcuts import render, redirect\nfrom django.http import HttpResponse, HttpRequest, HttpResponseRedirect, JsonResponse\nfrom django.template.response import SimpleTemplateResponse, TemplateResponse\nfrom django.contrib.auth import authenticate, login, logout\nfrom django.contrib.auth.models import User, Group\nfrom django.template import loader\nfrom django.views import generic\nfrom django.views.generic.edit import CreateView, UpdateView, DeleteView\nfrom django.urls import reverse_lazy, reverse\nfrom django.contrib import messages\nfrom django.utils import timezone\nfrom django.core import serializers\nfrom random import randint\nfrom django.core.mail import send_mail\nfrom django.conf import settings\nfrom django.core.files.storage import FileSystemStorage\nfrom django.core.files.base import ContentFile\nfrom django.core.files.storage import default_storage\nfrom ErpBackOffice.utils.identite import identite\nfrom ErpBackOffice.utils.tools import ErpModule\nfrom ErpBackOffice.utils.utils import utils\nimport os, calendar\nimport json\nimport pandas as pd\nfrom openpyxl import load_workbook, Workbook, styles\nfrom copy import copy\nfrom io import BytesIO\nfrom rest_framework.decorators import api_view\nimport base64, uuid\nfrom locale import atof, setlocale, LC_NUMERIC\nimport numpy as np\nfrom dateutil.relativedelta import relativedelta\nfrom ErpBackOffice.utils.separateur import makeFloat, checkDateTimeFormat, checkDateFormat, makeStringFromFloatExcel, makeInt, makeIntId, makeString\nfrom django.db import transaction\nfrom ErpBackOffice.dao.dao_organisation import dao_organisation\nfrom ErpBackOffice.dao.dao_utilisateur import dao_utilisateur\nfrom ErpBackOffice.dao.dao_wkf_workflow import dao_wkf_workflow\nfrom ErpBackOffice.dao.dao_wkf_etape import dao_wkf_etape\nfrom ErpBackOffice.dao.dao_wkf_historique import dao_wkf_historique\nfrom ErpBackOffice.models import *\nfrom ErpBackOffice.dao.dao_model import dao_model\nfrom ErpBackOffice.dao.dao_place import dao_place\nfrom ErpBackOffice.dao.dao_module import dao_module\nfrom ErpBackOffice.dao.dao_devise import dao_devise\nfrom ModuleConversation.dao.dao_notification import dao_notification\nfrom ModuleConversation.dao.dao_temp_notification import dao_temp_notification\nfrom ErpBackOffice.utils.pagination import pagination\nfrom ErpBackOffice.utils.auth import auth\nfrom ErpBackOffice.utils.wkf_task import wkf_task\nfrom ErpBackOffice.utils.endpoint import endpoint\nfrom ErpBackOffice.utils.print import weasy_print\nfrom ErpBackOffice.utils.utils import utils\nfrom ModuleConfiguration.dao.dao_query import dao_query\nfrom ErpBackOffice.dao.dao_query_builder import dao_query_builder\n\n\n#LOGGING\nimport logging, inspect, unidecode\nfrom {2}.models import *\nfrom ErpBackOffice.models import *\nfrom ModuleConfiguration.models import *\nmonLog = logging.getLogger("logger")\nmodule= "{2}"\nvar_module_id = {3}\nvars_module = {{"name" : "{1}", "value" : {4} }}\n\n\ndef get_index(request):\n\ttry:\n\t\tmodules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetDashboardAuthentification(var_module_id, request)\n\t\tif response != None:\n\t\t\treturn response\n\n\t\t#NOTIFCATION\n\t\tmodule_name = vars_module["name"]\n\t\ttemp_notif_list = dao_temp_notification.toGetListTempNotificationUnread(identite.utilisateur(request).id, module_name)\n\t\ttemp_notif_count = temp_notif_list.count()\n\t\tmessage_no_open = dao_temp_notification.toCountTempNotificationUnread(identite.utilisateur(request).id, module_name)\n\n\t\tcontext = {{\n\t\t\t"title" : "Accueil",\n\t\t\t"utilisateur" : utilisateur,\n\t\t\t"organisation": dao_organisation.toGetMainOrganisation(),\n\t\t\t"temp_notif_count": temp_notif_count,\n\t\t\t"temp_notif_list": temp_notif_list,\n\t\t\t"msg_no_open": message_no_open,\n\t\t\t"sous_modules":sous_modules,\n\t\t\t"modules" : modules,\n\t\t\t"module" : vars_module\n\t\t}}\n\n\t\ttemplate = loader.get_template("ErpProject/{2}/index.html")\n\t\treturn HttpResponse(template.render(context, request))\n\texcept Exception as e:\n\t\treturn auth.toReturnFailed(request, e, traceback.format_exc(), reverse(\'backoffice_index\'))'.format(module.description,nameModuleUp,nomModule,module.id,module.numero_ordre)
        texte_a_ajouter_views_py_dossier_application = texte_a_ajouter_views_py_dossier_application + '\n\n\ndef get_dashboard(request):\n\ttry:\n\t\tmodules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetDashboardAuthentification(var_module_id, request)\n\t\tif response != None:\n\t\t\treturn response\n\n\t\t#NOTIFCATION\n\t\tmodule_name = vars_module["name"]\n\t\ttemp_notif_list = dao_temp_notification.toGetListTempNotificationUnread(identite.utilisateur(request).id, module_name)\n\t\ttemp_notif_count = temp_notif_list.count()\n\n\t\tcontext = {{\n\t\t\t"title" : "Tableau de Bord",\n\t\t\t"utilisateur" : utilisateur,\n\t\t\t"organisation": dao_organisation.toGetMainOrganisation(),\n\t\t\t"temp_notif_count": temp_notif_count,\n\t\t\t"temp_notif_list": temp_notif_list,\n\t\t\t"sous_modules":sous_modules,\n\t\t\t"modules" : modules,\n\t\t\t"module" : vars_module\n\t\t}}\n\n\t\ttemplate = loader.get_template("ErpProject/{2}/dashboard.html")\n\t\treturn HttpResponse(template.render(context, request))\n\texcept Exception as e:\n\t\treturn auth.toReturnFailed(request, e, traceback.format_exc(), reverse(\'backoffice_index\'))'.format(module.description,nameModuleUp,nomModule,module.id,module.numero_ordre)
        fichier.write(texte_a_ajouter_views_py_dossier_application)
        fichier.close()
        print("views.py dans module cree")

        #Manipulation pour la creation fichier urls.py dans le sous-dossier de l'Application
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\urls.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_urls_py_dossier_application = '''from django.conf.urls import include, url\nfrom . import views\nurlpatterns = [
    url(r'^$', views.get_index, name='{0}_index'),
    url(r'^tableau', views.get_dashboard, name='{0}_tableau_de_bord'),
]
    '''.format(nom_pattern)
        fichier.write(texte_a_ajouter_urls_py_dossier_application)
        fichier.close()
        print("urls.py dans module cree")

        #Création fichier index.html (Page d'acceuil du Module)
        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}".format(nomModule)
        os.mkdir(utils.format_path(path))
        path = path + "\\index.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')    
        texte_a_ajouter_index_html_template = genIndexTemplate(nomModule,nom_pattern)
        fichier.write(texte_a_ajouter_index_html_template)
        fichier.close()
        print("index.html dans template cree")
        
        #Création fichier dashboard.html (Tableau de Bord du Module)
        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}".format(nomModule)
        path = path + "\\dashboard.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')    
        texte_a_ajouter_dashboard_html_template1 = genDashBoardTemplate1(nomModule)
        texte_a_ajouter_dashboard_html_template2 = genDashBoardTemplate2(nomModule)
        texte_a_ajouter_dashboard_html_template3 = genDashBoardTemplate3(nomModule)
        if model_dashboard == 1:
            fichier.write(texte_a_ajouter_dashboard_html_template1)
        elif model_dashboard == 2:
            fichier.write(texte_a_ajouter_dashboard_html_template2)
        elif model_dashboard == 3:
            fichier.write(texte_a_ajouter_dashboard_html_template3)
        fichier.close()
        print("dashboard.html dans template cree")

        #Creation du dossier Shared  et de son fichier Layout.html
        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}\\shared".format(nomModule)
        os.mkdir(utils.format_path(path))
        path = path + "\\layout.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_layout_html_dossier_template = genLayoutHtmlOfModule()
        fichier.write(texte_a_ajouter_layout_html_dossier_template)
        fichier.close()
        print("layout.html dans template shared cree")        

        #Ajout Style du Module
        path = os.path.abspath(os.path.curdir)
        path = path + "\\static\\ErpProject\\css\\custom.css"
        fichier = codecs.open(utils.format_path(path), "a", encoding='utf-8')
        if module.couleur == "Vert":
            texte_a_ajouter_custom_css = genCssVertTemplate(module.nom_module)
        elif module.couleur == "Bleu foncé":
            texte_a_ajouter_custom_css = genCssBleuTemplate(module.nom_module)
        elif module.couleur == "Bleu ciel":
            texte_a_ajouter_custom_css = genCssBleuCielTemplate(module.nom_module)
        elif module.couleur == "Pourpre":
            texte_a_ajouter_custom_css = genCssPourpreTemplate(module.nom_module)
        elif module.couleur == "Magenta":
            texte_a_ajouter_custom_css = genCssMagentaTemplate(module.nom_module)
        else: texte_a_ajouter_custom_css = genCssOrangeTemplate(module.nom_module)
        fichier.write(texte_a_ajouter_custom_css)
        fichier.close()
        print("custom.css couleur du module ajoute")

        #Création et Initialisation du fichier Model py
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\models.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_models_py_dossier_application = "# -*- coding: utf-8 -*-\nfrom __future__ import unicode_literals\nfrom django.db import models\nfrom ErpBackOffice.models import Model_Personne\n\n# Create your models here.\n"
        fichier.write(texte_a_ajouter_models_py_dossier_application)
        fichier.close()
        print("models.py dans module ajoute")

        #Création et Initialisation du fichier Admin py
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\admin.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_admin_py_dossier_application = "from django.contrib import admin\nfrom . import models\n"
        fichier.write(texte_a_ajouter_admin_py_dossier_application)
        fichier.close()
        print("admin.py dans module ajoute")

        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\dao\\".format(nomModule)
        os.mkdir(utils.format_path(path))
        print("dossier dao dans module ajoute")    


        #On crée les groupes-menus par défaut du nouveau module ("Nom du module", "Business Inteligence" et "Configuration")
        designation = module.nom_module.capitalize()
        groupemenu = dao_groupemenu.toCreateGroupemenu(designation,"achats.svg", "Groupe Menu {}".format(designation), module.id, 1)
        groupemenu = dao_groupemenu.toSaveGroupemenu(auteur, groupemenu) 
        print("Groupe Menu 1 cree")

        designation = "Rapports"
        groupemenu = dao_groupemenu.toCreateGroupemenu(designation,"file.svg", "Groupe Menu {}".format(designation), module.id, 8)
        groupemenu = dao_groupemenu.toSaveGroupemenu(auteur, groupemenu) 
        print("Groupe Menu 2 cree")

        designation = "Analyses"
        groupemenu = dao_groupemenu.toCreateGroupemenu(designation,"line-chart-for-business.svg", "Groupe Menu {}".format(designation), module.id, 9)
        groupemenu = dao_groupemenu.toSaveGroupemenu(auteur, groupemenu) 
        print("Groupe Menu 3 cree")
        
        designation = "Configurations"
        groupemenu = dao_groupemenu.toCreateGroupemenu(designation,"setting.svg", "Groupe Menu {}".format(designation), module.id, 10)
        groupemenu = dao_groupemenu.toSaveGroupemenu(auteur, groupemenu) 
        print("Groupe Menu 4 cree")

        #On crée le menu pour le tableau de bord du module
        nom_sous_module = "Tableau de Bord"
        url_vers = '{}_tableau_de_bord'.format(nom_pattern)
        sous_module = dao_sous_module.toCreateSousModule(module.id, nom_sous_module, nom_sous_module,  numero_ordre = 0, url_vers = url_vers, est_dashboard = True, icon_menu = "dashboard.svg")
        sous_module = dao_sous_module.toSaveSousModule(sous_module)
        print("Sous Menu 1 cree")
        
        permission = dao_permission.toCreatePermission(sous_module.id, "ADMIN_DASHBOARD_{}".format(unidecode.unidecode(module.nom_module.upper().replace(" ","_"))), dao_permission.toGetLatestNumeroOrdre() + 1)
        permission = dao_permission.toSavePermission(auteur, permission)
        print("FIN CREATION PERMISSION ADMIN DASHBOARD")
        
        permission = dao_permission.toCreatePermission(sous_module.id, "USER_DASHBOARD_{}".format(unidecode.unidecode(module.nom_module.upper().replace(" ","_"))), dao_permission.toGetLatestNumeroOrdre() + 1)
        permission = dao_permission.toSavePermission(auteur, permission)
        print("FIN CREATION PERMISSION USER DASHBOARD")


        #Parametrage dans Application Settings à propos de ce module comme new Installed Apps
        path = os.path.abspath(os.path.curdir)
        path = path + "\\ErpProject\\settings.py"
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_a_ajouter_settings_py_dossier_projet="\nINSTALLED_APPS.append('{0}')".format(nomModule)
        fichier.write(texte_a_ajouter_settings_py_dossier_projet)
        fichier.close()
        print("module ajoute dans INSTALLED_APPS sur settings.py") 
        
        #Manipulation pour ajouter liens urls dans le repertoire du projet principal
        path = os.path.abspath(os.path.curdir)
        path = path + "\\ErpProject\\urls.py"
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_a_ajouter_urls_py_dossier_projet="\nurlpatterns.append(url(r'^{0}', include('{1}.urls')))".format(module.url_vers,nomModule)
        fichier.write(texte_a_ajouter_urls_py_dossier_projet)
        fichier.close()
        print("urls.py dans projet cree")
        print("Fin parametrage_module")
    except Exception as e:
        print("ERREUR PARAMETRAGE FILES OF THIS MODULE!")
        print(e)
        pass


#MODELE DANS MODULE
def get_creer_modele(request, ref):
    permission_number = 555
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        id = int(ref)

        module = dao_module.toGetModule(id)
        groupe = dao_sous_module.toGetSousModuleDuModuleGroupeOf(ref,"parent")
        context = {
            'title' : 'Nouveau modele',
            'utilisateur' : utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'modules' : modules,
            "module" : vars_module,
            'sous_modules': sous_modules,
            'groupe':groupe,
            'model': module,
            
            'menu' : 4
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/module/modele/add.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        print("ERREUR MODEL ON GET CREER MODELE ")
        print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_details_module', args=(id,)))


def post_creer_modele(request,ref):
    try:
        id = int(ref)
        module = dao_module.toGetModule(id)
        nomModule = module.nom_application
        nom_pattern = 'module_{0}'.format(unidecode.unidecode(module.nom_module.lower().replace(" ","_")))
        db_table = str(request.POST['db_table'])
        verbose_name = str(request.POST['verbose_name'])
        verbose_name_plural = str(request.POST['verbose_name_plural'])
        nom_modele = request.POST["nom_modele"]

        list_nom_champ = request.POST.getlist("nom_champ",None)
        list_nom_verbose = request.POST.getlist("nom_verbose",None)
        list_type_champ = request.POST.getlist("type_champ",None)
        list_taille = request.POST.getlist("taille",None)
        list_default = request.POST.getlist("default",None)
        list_nullable = request.POST.getlist("nullable",None)
        list_foreign = request.POST.getlist("foreign",None)
        list_ondelete = request.POST.getlist("ondelete",None)
        list_choix = request.POST.getlist("choix",None)
        list_nom_choix = request.POST.getlist("nom_choix",None)
        list_valeur_choix = request.POST.getlist("valeur_choix",None)

        #Standardisation denomination modele
        nom_modele = nom_modele.replace(" ","_").capitalize()
        nomModele = 'Model_{0}'.format(nom_modele)

        #Création Modele dans le fichier models.py de ErpBackOffice
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\models.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_a_ajouter = ""
        
        
        #On ajoute dabord les listes de choix
        for i in range(0,len(list_nom_champ)):
            choix = 0
            valeur_select_choix = False
            try:
                choix = int(list_choix[i])
                if choix == 1:
                    valeur_select_choix = True
                elif choix == 2:
                    valeur_select_choix = False
            except Exception as e: pass
            
            nom_choix = ""
            try:
                nom_choix = list_nom_choix[i]
            except Exception as e: pass
            
            valeur_choix = ""
            try:
                valeur_choix = list_valeur_choix[i]
            except Exception as e: pass
            
            if valeur_select_choix == True:
                texte_a_ajouter = texte_a_ajouter + "\n{0}  =  ({1})\n".format(nom_choix, valeur_choix)
            
        texte_a_ajouter = texte_a_ajouter + "\nclass {0}(models.Model):".format(nomModele)
        fichier.write(texte_a_ajouter)
        fichier.close()

        #les champs du Modele dans le mm fichier
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_=""

        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e: pass
            
            nom_verbose = nom_champ.replace("_", " ").capitalize()
            try:
                nom_verbose = list_nom_verbose[i]
            except Exception as e: pass

            type_data = ""
            try:
                type_data = str(list_type_champ[i])
            except Exception as e: pass
            
            taille = 100
            try:
                taille = int(list_taille[i])
                if taille == 0: taille = 100
            except Exception as e: pass
            
            valeur_default = ""
            try:
                valeur_default = list_default[i]
            except Exception as e: pass
            
            nullable = 0
            valeur_null = False
            try:
                nullable = int(list_nullable[i])
                if nullable == 1:
                    valeur_null = True
                elif nullable == 2:
                    valeur_null = False
            except Exception as e: pass

            try:
                foreign = list_foreign[i]
            except Exception as e: pass
            
            try:
                ondelete = list_ondelete[i]
            except Exception as e: pass
            
            choix = 0
            valeur_select_choix = False
            try:
                choix = int(list_choix[i])
                if choix == 1:
                    valeur_select_choix = True
                elif choix == 2:
                    valeur_select_choix = False
            except Exception as e: pass

            nom_choix = ""            
            try:
                nom_choix = list_nom_choix[i]
            except Exception as e: pass


            texte_models = "\n\t{0}    =    models.{1}(".format(nom_champ,type_data)
            if type_data in ("ForeignKey", "OneToOneField",  "ManyToManyField") :
                related_name = "{0}s_{1}".format(nom_modele.lower(), nom_champ)
                texte_models = texte_models + '"{0}", related_name = "{1}", '.format(foreign, related_name)
                if type_data == "ForeignKey":
                    texte_models = texte_models + "on_delete=models.{0}, ".format(ondelete)
                    list_nom_champ[i] = '{0}_id'.format(nom_champ)
            elif type_data in ("CharField", "IntegerField", "FloatField", "EmailField") :
                if valeur_select_choix : texte_models = texte_models + "choices = {0}, ".format(nom_choix)
                if valeur_default != "" : texte_models = texte_models + "default = {}, ".format(valeur_default)
                if type_data == "CharField" : texte_models = texte_models + "max_length = {}, ".format(taille)
            elif type_data == "BooleanField":
                if valeur_default != "" : valeur_default = "False"
                texte_models = texte_models + "default = {}, ".format(valeur_default)
                
            if valeur_null : texte_models = texte_models + "null = True, blank = True, "
            texte_models = texte_models + 'verbose_name = "{}" )'.format(nom_verbose)

            texte_ = texte_ + texte_models
            #fichier.write(texte_models)

        texte_ = texte_ + "\n\tstatut    =    models.ForeignKey('ErpBackOffice.Model_Wkf_Etape', on_delete=models.SET_NULL, blank=True, null=True)\n\tetat    =    models.CharField(max_length=50, blank=True, null=True)\n\tcreation_date    =    models.DateTimeField(auto_now_add = True)\n\tupdate_date    =    models.DateTimeField(auto_now = True)\n\tauteur    =    models.ForeignKey(Model_Personne, on_delete = models.SET_NULL, related_name = 'auteur_{0}s', null = True, blank = True)".format(nom_modele.lower())
        texte_ = texte_ + "\n\n\tdef __str__(self):\n\t\treturn self.{0}\n".format(list_nom_champ[0])
        texte_ = texte_ + '\n\n\tclass Meta:\n\t\tverbose_name = "{0}"\n\t\tverbose_name_plural = "{1}"'.format(verbose_name, verbose_name_plural)
        if db_table != "": texte_ = texte_ + '\n\t\tdb_table = "{0}"\n\n'.format(db_table)

        #On ajoute dabord les listes de choix
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e: pass
            
            choix = 0
            valeur_select_choix = False
            try:
                choix = int(list_choix[i])
                if choix == 1:
                    valeur_select_choix = True
                elif choix == 2:
                    valeur_select_choix = False
            except Exception as e: pass
            
            nom_choix = ""
            try:
                nom_choix = list_nom_choix[i]
            except Exception as e: pass
            
            valeur_choix = ""
            try:
                valeur_choix = list_valeur_choix[i]
            except Exception as e: pass
            
            if valeur_select_choix == True:
                texte_ = texte_ + "\n\t@property\n\tdef value_{0}(self):\n\t\tif self.{0}: return dict({1})[int(self.{0})]\n".format(nom_champ, nom_choix)
                texte_ = texte_ + "\n\t@property\n\tdef list_{0}(self):\n\t\tlist = []\n\t\tfor key, value in {1}:\n\t\t\titem = {{'id' : key,'designation' : value}}\n\t\t\tlist.append(item)\n\t\treturn list\n".format(nom_champ, nom_choix)
        fichier.write(texte_)
        fichier.close()
        print("Fin MAJ Model")

        #Enregistrement dans Admin py afin que le model soit visible dans le panel Admin
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\admin.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_a_ajouter_admin_py_dossier_application = "\nadmin.site.register(models.{0})".format(nomModele)
        fichier.write(texte_a_ajouter_admin_py_dossier_application)
        fichier.close()
        print("admin.py cree")

        #Enregistrement dans la table prévue pour
        mod = dao_moduleovermodel.toCreateModuleOverModel(nomModele,module.id)
        mod = dao_moduleovermodel.toSaveModuleOverModel(mod)
        #Popen('python manage.py migrate')
        print("Object model cree")

        #Fin des Opérations, on retourne Http
        messages.add_message(request, messages.SUCCESS, "Opération Réussie. Veuillez éxecuter 'python manage.py makemigrations' suivi de 'python manage.py migrate'")
        return HttpResponseRedirect(reverse('module_configuration_details_module', args=(id,)))
    except Exception as e:
        print("ERREUR !")
        print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_add_modele', args=(id,)))
    
def get_exporter_modele(request, ref):
	permission_number = 554
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

	if response != None:
		return response
	try:
		ref = int(ref)
		model = dao_module.toGetModule(ref)
		modeles = ContentType.objects.filter(app_label = model.nom_application)
		groupesmenus = dao_groupe_menu.toListGroupeOfModule(ref)
		groupe_menu1 = groupesmenus.filter(designation = model.nom_module).first()
		groupe_menu2 = groupesmenus.filter(designation = "Configurations").first()
		menus = dao_sous_module.toListSousModulesOfModule(ref)

		#On génère le fichier excel file_name =  "Liste_modeles.xlsx"
		from openpyxl import load_workbook, Workbook, styles
		from io import BytesIO

		wb = Workbook()
		ws = wb.active

		titleStyle = styles.NamedStyle(name = 'title_style')
		titleStyle.font = styles.Font(name = 'Calibri', size = 11, color = 'FF000000', bold=True)
		titleStyle.alignment = styles.Alignment(horizontal='left', vertical='top',wrap_text=True)

		pageStyle = styles.NamedStyle(name = 'page_style')
		pageStyle.font = styles.Font(name = 'Calibri', size = 11, color = 'FF000000')
		pageStyle.alignment = styles.Alignment(vertical='bottom')
		#borderStyle = styles.Side(style = 'dashDot', color = 'FF00FF')
		#pageStyle.border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)

		#On crée les entêtes du tableau
		ws['A1'] = 'module_id'
		ws['A1'].style = titleStyle
		ws['B1'] = 'modele_id'
		ws['B1'].style = titleStyle
		ws['C1'] = 'modele_name'
		ws['C1'].style = titleStyle
		ws['D1'] = 'numero_ordre'
		ws['D1'].style = titleStyle
		ws['E1'] = 'groupe_menu_id'
		ws['E1'].style = titleStyle
		ws['F1'] = 'groupe_menu_name'
		ws['F1'].style = titleStyle
		ws['G1'] = 'related_models'
		ws['G1'].style = titleStyle
		ws['H1'] = 'generate_reporting'
		ws['H1'].style = titleStyle
		ws['I1'] = 'generate_bi'
		ws['I1'].style = titleStyle
		ws['J1'] = 'est_actif'
		ws['J1'].style = titleStyle

		row = 2

		for item in modeles:
			cell1 = ws.cell(row=row,column=1)
			cell1.value = model.id
			cell1.style = pageStyle
			cell1.data_type = 'n'
   
			cell2 = ws.cell(row=row,column=2)
			cell2.value = item.id
			cell2.style = pageStyle
			cell2.data_type = 'n'
   
			cell3 = ws.cell(row=row,column=3)
			cell3.value = item.model
			cell3.style = pageStyle
   
			cell4 = ws.cell(row=row,column=4)
			num_order = row - 1
			cell4.value = makeInt(num_order)
			cell4.style = pageStyle
			cell4.data_type = 'n'

			if row == 2:				
				groupe_menu_id = groupe_menu1.id if groupe_menu1 else 0
				groupe_menu_name = groupe_menu1.designation if groupe_menu1 else ""
			elif row == 3:
				groupe_menu_id = groupe_menu1.id if groupe_menu1 else 0
				groupe_menu_name = groupe_menu1.designation if groupe_menu1 else ""
			elif row == 4:
				groupe_menu_id = groupe_menu1.id if groupe_menu1 else 0
				groupe_menu_name = groupe_menu1.designation if groupe_menu1 else ""
			elif row == 5:
				groupe_menu_id = groupe_menu2.id if groupe_menu2 else 0
				groupe_menu_name = groupe_menu2.designation if groupe_menu2 else ""
			else:
				groupe_menu_id = groupe_menu2.id if groupe_menu2 else 0
				groupe_menu_name = groupe_menu2.designation if groupe_menu2 else ""
       
			cell5 = ws.cell(row=row,column=5)
			cell5.value = groupe_menu_id
			cell5.style = pageStyle
			cell5.data_type = 'n'
   
			cell6 = ws.cell(row=row,column=6)
			cell6.value = groupe_menu_name
			cell6.style = pageStyle
   
			cell7 = ws.cell(row=row,column=7)
			cell7.value = ''
			cell7.style = pageStyle
   
			cell8 = ws.cell(row=row,column=8)
			cell8.value = "Non"
			cell8.style = pageStyle
   
			cell9 = ws.cell(row=row,column=9)
			cell9.value = "Non"
			cell9.style = pageStyle
   
			cell10 = ws.cell(row=row,column=10)
			cell10.value = "Oui"
			cell10.style = pageStyle

			row = row+1

		for col in ws.columns:
			max_length = 0
			column = col[0].column_letter # Get the column name
			for cell in col:
				try: # Necessary to avoid error on empty cells
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = (max_length + 2) * 1.2
			ws.column_dimensions[column].width = adjusted_width


		buffer = BytesIO()
		wb.save(buffer)
		excell_file = buffer.getvalue()
		buffer.close()
		response = HttpResponse(excell_file, content_type='application/xlsx')
		response['Content-Disposition'] = 'inline;filename=Liste_modeles.xlsx'
		return response
	except Exception as e:
		#print("ERREUR")
		#print(e)
		messages.error(request,e)
		return HttpResponseRedirect(reverse("module_configuration_list_modules"))

def get_creer_framework(request):
    permission_number = 555
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        context = {
            'title' : 'Generer un squelette',
            'utilisateur' : utilisateur,
            'modules' : modules,
            "module" : vars_module,
            'sous_modules': sous_modules,
            
            'menu' : 3
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/framework/generate.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR MODEL ON GET CREER FRAMEWORK")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_modules'))


def post_creer_framework(request):
    try:
        #Création du Module dans la table module
        name_module = request.POST["name_module"]
        numero_ordre = int(request.POST["numero_ordre"])
        module = Model_Module()
        #print("youv")
        module.nom_module = name_module
        #print("youv2")
        module.numero_ordre = numero_ordre
        #HERE7
        module.url_vers = "{0}/".format(name_module.lower())
        module.save()

        module = dao_module.toGetModuleOf(numero_ordre)
        #print("youva")
        #print(module.nom_module)
        framework_setting_module(module)

        nomModule = module.nom_application


        list_fonction = request.POST.getlist("fonction",None)
        nomModele= ""
        #Création Modele dans le fichier models.py
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\models.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')

        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass

            #Standardisation denomination modele
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModele = 'Model_{0}'.format(nom_modele)

            #Ecritrure dans le fichier models.py
            texte_a_ajouter_models_py_dossier_application = "\nclass {0}(models.Model):\n\tpass\n\n".format(nomModele)
            fichier.write(texte_a_ajouter_models_py_dossier_application)


        fichier.close()


        #Migrations et execution
        #print("move on")
        #Popen('python manage.py makemigrations '+nomModule+ ' | python manage.py migrate',shell=True)
        #run('python manage.py makemigrations '+nomModule+ '; ./ manage.py migrate')
        #print('passed')

        #Enregistrement dans Admin py afin que le model soit visible dans le panel Admin
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\admin.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModele = 'Model_{0}'.format(nom_modele)
            texte_a_ajouter_admin_py_dossier_application = "\nadmin.site.register(models.{0})".format(nomModele)
            fichier.write(texte_a_ajouter_admin_py_dossier_application)
        fichier.close()
        #print('un')

        #Just au cas où
        #Popen('python manage.py migrate')

        #Creation dossier dao
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\dao".format(nomModule)
        os.mkdir(utils.format_path(path))

        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            nomdao = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModele = 'Model_{0}'.format(nom_modele)
            nomdao="dao_{0}".format(nom_modele.lower())
            path = os.path.abspath(os.path.curdir)
            path = path + "\\{0}\\dao\\{1}.py".format(nomModule,nomdao)
            fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
            texte_a_ajouter_dao_dossier_application = "from __future__ import unicode_literals\nfrom {0}.models import {1}\n\nclass {2}(object):".format(nomModule,nomModele,nomdao)
            fichier.write(texte_a_ajouter_dao_dossier_application)
            fichier.close()
            fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
            texte_about_methode_dao = "\n\tid = 0"
            #fichier.write(texte_add_hors)

            texte_about_methode_dao = texte_about_methode_dao + "\n\n\t@staticmethod\n\tdef toList():\n\t\treturn {1}.objects.all()\n".format(nom_modele,nomModele)
            #fichier.write(texte_about_methode_dao)
            texte_about_methode_dao = texte_about_methode_dao + "\n\t@staticmethod\n\tdef toGet(id):\n\t\ttry:\n\t\t\treturn Model_{2}.objects.get(pk = id)\n\t\texcept Exception as e:\n\t\t\treturn None\n\n\t@staticmethod\n\tdef toDelete(id):\n\t\ttry:\n\t\t\t{0} = Model_{2}.objects.get(pk = id)\n\t\t\t{0}.delete()\n\t\t\treturn True\n\t\texcept Exception as e:\n\t\t\treturn False".format(nom_modele.lower(),nom_modele.upper(),nom_modele.capitalize())
            fichier.write(texte_about_methode_dao)
            fichier.close()
        #print('deux')


        #Creation des fonctions CRUD Dans views.py
        nameModuleUp = 'MODULE_{0}'.format(unidecode.unidecode(module.nom_module.upper().replace(" ","_")))

        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\views.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')

        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            nomdao = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModele = 'Model_{0}'.format(nom_modele)
            nomdao="dao_{0}".format(nom_modele.lower())
            nom_pattern = 'module_{0}'.format(unidecode.unidecode(module.nom_module.lower().replace(" ","_")))

            texte_addenda_views_py_appli = "\n\nfrom {4}.dao.{1} import {1}\n\ndef get_lister_{0}(request):\n\tis_connect = identite.est_connecte(request)\n\tif is_connect == False: return HttpResponseRedirect(reverse('backoffice_connexion'))\n\n\tmodel = {1}.toList{2}()\n\tcontext ={{'title' : 'Liste des {0}','model' : model,'utilisateur': utilisateur,'modules' : dao_module.toListModulesInstalles(),'module' : ErpModule.{3},'menu' : 2}}\n\ttemplate = loader.get_template('ErpProject/{4}/{0}/list.html')\n\treturn HttpResponse(template.render(context, request))".format(nom_modele.lower(),nomdao,nom_modele.capitalize(),nameModuleUp,nomModule)
            #fichier.write(texte_addenda_views_py_appli)
            texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\n\n\ndef get_creer_{0}(request):\n\tis_connect = identite.est_connecte(request)\n\tif is_connect == False: return HttpResponseRedirect(reverse('backoffice_connexion'))\n\tcontext = {{'title' : 'Nouvelle {0}','utilisateur': utilisateur,'modules' : dao_module.toListModulesInstalles(),'module' : ErpModule.{1},'menu' : 2}}\n\ttemplate = loader.get_template('ErpProject/{2}/{0}/add.html')\n\treturn HttpResponse(template.render(context, request))".format(nom_modele.lower(),nameModuleUp,nomModule)
            #fichier.write(another_text_addenda)
            texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\n\n\ndef post_creer_{0}(request):\n\ttry:\n\t\treturn HttpResponseRedirect(reverse('{1}_list_{0}'))\n\texcept Exception as e:\n\t\t#print('Erreur lors de l\\'enregistrement')\n\t\t#print(e)\n\t\treturn HttpResponseRedirect(reverse('{1}_add_{0}'))".format(nom_modele.lower(),nom_pattern)
            #fichier.write(another_one)
            texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\n\n\ndef get_details_{0}(request, ref):\n\tis_connect = identite.est_connecte(request)\n\tif is_connect == False: return HttpResponseRedirect(reverse('backoffice_connexion'))\n\ttry:\n\t\tref = int(ref)\n\t\ttemplate = loader.get_template('ErpProject/{1}/{0}/item.html')\n\t\tcontext = {{'title' : 'Details d\\'un(e) {0}','utilisateur': utilisateur,'modules' : dao_module.toListModulesInstalles(),'module' : ErpModule.{2},'menu' : 4}}\n\t\treturn HttpResponse(template.render(context, request))\n\texcept Exception as e:\n\t\t#print('ERREUR s')\n\t\t#print(e)\n\t\treturn HttpResponseRedirect(reverse('{3}_list_{0}'))".format(nom_modele.lower(),nom_modele.capitalize(),nameModuleUp,nom_pattern)
            fichier.write(texte_addenda_views_py_appli)
        fichier.close()
        #print('trois')

        #Manipulation pour la creation fichier urls.py dans le sous-dossier de l'Application
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\urls.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            nomdao = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModele = 'Model_{0}'.format(nom_modele)
            nomdao = "dao_{0}".format(nom_modele.lower())
            nom_pattern = 'module_{0}'.format(unidecode.unidecode(module.nom_module.lower().replace(" ","_")))

            texte_a_ajouter_urls_py_dossier_ap="\nurlpatterns.append(url(r'^{0}/list', views.get_lister_{0}, name = '{1}_list_{0}'))".format(nom_modele.lower(),nom_pattern)
            #fichier.write(texte_a_ajouter_urls_py_dossier_ap)
            texte_a_ajouter_urls_py_dossier_ap= texte_a_ajouter_urls_py_dossier_ap + "\nurlpatterns.append(url(r'^{0}/add', views.get_creer_{0}, name = '{1}_add_{0}'))".format(nom_modele.lower(),nom_pattern)
            #fichier.write(texte_a_ajouter_urls_py_dossier_ap)
            texte_a_ajouter_urls_py_dossier_ap= texte_a_ajouter_urls_py_dossier_ap + "\nurlpatterns.append(url(r'^{0}/post_add', views.post_creer_{0}, name = '{1}_post_add_{0}'))".format(nom_modele.lower(),nom_pattern)
            #fichier.write(texte_a_ajouter_urls_py_dossier_ap)
            texte_a_ajouter_urls_py_dossier_ap= texte_a_ajouter_urls_py_dossier_ap + "\nurlpatterns.append(url(r'^{0}/item/(?P<ref>[0-9]+)/$', views.get_details_{0}, name = '{1}_detail_{0}'))".format(nom_modele.lower(),nom_pattern)
            fichier.write(texte_a_ajouter_urls_py_dossier_ap)
        fichier.close()
        #print('quatre')

        #Creation du dossier Shared  et de son fichier Layout.html
        #Popen('python manage.py makemigrations '+nomModule+ ' | python manage.py migrate',shell=True)

        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}\\shared".format(nomModule)
        os.mkdir(utils.format_path(path))
        path = path + "\\layout.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_layout_html_dossier_template = '''{{% extends "ErpProject/ErpBackOffice/shared/layout.html" %}} {{% block content %}} {{% load static %}}
<!-- Suite Menu Nav -->

<!-- Menu lateral -->
<div class="sidebar" role="navigation" style="">

    <div class="contenair-profil" style="background-color: transparent;">
        <img src="/static/ErpProject/image/icone_profile.png" class="profil">
        <label class="nom-admin">{{{{ utilisateur.nom_complet }}}}</label>
        <P class="fonction">Administrateur</P>
        <div class="divider" style="background-color: transparent;"></div>
    </div>
    <div class="sidebar-nav navbar-collapse" style="background-color: transparent;">
        <ul class=" nav" id="side-menu" style="background-color: transparent;">
            <li class="li-menu">
                <a href="#" class="a-menu"><i class="fa fa-dashboard fa-fw"></i> Accueil <span class="fa fa arrow"></span></a>
                <!-- nav-second-niveau -->
                <ul class="nav nav-second-level">
                    <li class="{{% if menu == 1 %}}{{ 'active' }}{{% endif %}} chargement-au-click">
                        <a href="{{% url 'module_{0}_index' %}}" class="a-menu">Home</a>
                    </li>
                </ul>
                <!-- /.nav-second-niveau -->
            </li>'''.format(module.nom_module.lower().replace(" ","_"))

        fichier.write(texte_a_ajouter_layout_html_dossier_template)
        fichier.close()
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            nomdao = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModele = 'Model_{0}'.format(nom_modele)
            nomdao = "dao_{0}".format(nom_modele.lower())
            nom_pattern = 'module_{0}'.format(unidecode.unidecode(module.nom_module.lower().replace(" ","_")))

            texte2 = '''<li class="li-menu"><a href="{{% url '{0}_list_{1}' %}}" class="a-menu"><i class="fa fa-dashboard fa-fw"></i> {2} <span class="fa fa arrow"></span></a></li>'''.format(nom_pattern,nom_modele.lower(),nom_fonction.capitalize())
            fichier.write(texte2)

        #print('cinq')
        texte3 ='''</ul>
    </div>

    <!-- /.sidebar-collapse -->
</div>
<!-- /.Menu lateral -->
</nav>
<!-- /.Menu Navbar -->

<!-- Corps de la page (A definir dans chaque fonction du module) -->
<div id="page-wrapper" style="background-color:white;">
    {% block page %}{% endblock %}
</div>
<!-- Fin Corps de la page -->
{% endblock %}'''
        fichier.write(texte3)
        fichier.close()
        #print('cinq m')

        #Creation des fichiers add list item
        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            nomdao = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModele = 'Model_{0}'.format(nom_modele)
            nomdao="dao_{0}".format(nom_modele.lower())
            nom_pattern = 'module_{0}'.format(unidecode.unidecode(module.nom_module.lower().replace(" ","_")))

            try:
                path = os.path.abspath(os.path.curdir)
                path = path + "\\templates\\ErpProject\\{0}\\{1}".format(nomModule,nom_modele.lower())
                os.mkdir(utils.format_path(path))
            except Exception as e:
                pass
            path = path + "\\list.html"
            fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
            #print('in')

            texte = '''{{% extends "ErpProject/{0}/shared/layout.html" %}}
{{% block page %}}
   <div class="row">
    <ul class="breadcrumb">
        <li><a><span class="mif-home"></span></a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_index' %}}">Accueil</a></li>
        <li>{{{{ title }}}}</li>
    </ul>
</div>
    <div class="row">
        <button onclick="javascript:window.location.assign('{{% url '{2}_add_{3}' %}}')" class="button small-button rounded primary chargement-au-click">Creer</button>
    </div>
    <hr style="background-color: #f7f5f5;">
    <table id="example" class="display nowrap border bordered striped" cellspacing="0" style="width:100%">
            <thead>
                <tr>
                    <th style="width: 20px; background-color:#2e416a; white"></th>
            </thead>
            <tbody>

                    <tr>
                        <td>
                            <label class="small-check">
                                <input type="checkbox">
                                <span class="check"></span>
                            </label>
                        </td>
                        <td>
                        </td>
                    </tr>

            </tbody>
        </table>
    </div>
{{% endblock %}}'''.format(nomModule,module.nom_module.lower().replace(" ","_"),nom_pattern,nom_modele.lower())
            fichier.write(texte)
            fichier.close()
            path = os.path.abspath(os.path.curdir)
            path = path + "\\templates\\ErpProject\\{0}\\{1}".format(nomModule,nom_modele.lower())
            path = path + "\\add.html"
            fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
            #print('in1')
            texte10 = '''{{% extends "ErpProject/{0}/shared/layout.html" %}}
{{% block page %}}
<div class="row">
    <ul class="breadcrumb">
        <li><a><span class="mif-home"></span></a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_index' %}}">Accueil</a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_list_{3}' %}}">Liste</a></li>
        <li>{{{{ title }}}}</li>
    </ul>
</div>
    <div class="row">
        <button onclick="javascript:document.getElementById('submit').click()" class="button small-button rounded primary">Valider</button>
        <button onclick="javascript:window.location.assign('{{% url '{2}_list_{3}' %}}')" class="button small-button rounded chargement-au-click" style="margin-left: 5px">Annuler</button>
    </div>
    <hr style="background-color: #f7f5f5;">
    <div class="row" style="margin-top: 10px">
        <form action="{{% url '{2}_post_add_{3}' %}}" method="POST"
            data-role="validator"
            data-show-required-state="false"
            data-hint-mode="line"
            data-hint-background="bg-red"
            data-hint-color="fg-white"
            data-hide-error="5000"
            novalidate="novalidate"
            data-on-error-input="notifyOnErrorInput"
            data-show-error-hint="false">
            {{% csrf_token %}}
            <input id="submit" type="submit" style="display: none">
            <input id="nombreLigne" type="hidden" value="0" />
            <div class="row">

            </div>

        </form>
    </div>


{{% endblock %}}'''.format(nomModule,module.nom_module.lower().replace(" ","_"),nom_pattern,nom_modele.lower())
            fichier.write(texte10)
            fichier.close()
            path = os.path.abspath(os.path.curdir)
            path = path + "\\templates\\ErpProject\\{0}\\{1}".format(nomModule,nom_modele.lower())
            path = path + "\\item.html"
            fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
            #print('six')
            texte11='''{{% extends "ErpProject/{0}/shared/layout.html" %}}

{{% block page %}}
    {{% load static %}}
   <div class="row">
    <ul class="breadcrumb">
        <li><a><span class="mif-home"></span></a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_index' %}}">Accueil</a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_list_{3}' %}}">Liste</a></li>
        <li>{{{{ title }}}}</li>
    </ul>
</div>
    <div class="row padding30 bg-white" style="padding-top:15px">
        <div class="row">
            <div class="row cells2">
                <div class="cell" style="padding:0 5px">
                    <div class="row padding20 no-padding-top no-padding-bottom" style="border-right: 1px solid #f7f5f5;">

                    </div>
                </div>
             </div>
        </div>
</div>
{{% endblock %}}'''.format(nomModule,module.nom_module.lower().replace(" ","_"),nom_pattern,nom_modele.lower())
            fichier.write(texte11)
            fichier.close()
            #print('septs')

        #print("above all")
        #Popen('python manage.py makemigrations '+nomModule+ ' | python manage.py migrate',shell=True)

        #Fin des Opérations, on retourne Http
        messages.add_message(request, messages.SUCCESS, "Opération Réussie. Veuillez éxecuter 'python manage.py makemigrations' suivi de 'python manage.py migrate'")
        return HttpResponseRedirect(reverse('module_configuration_generate_framework'))

    except Exception as e:
        #print("ERREUR !")
        #print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_generate_framework'))

def framework_setting_module(module):
    try:
        #print('before')
        nomModule = module.nom_application
        #print('after')

        #Creation du Module via invite de commande
        #Popen('python manage.py startapp '+nomModule, shell = True)
        call(["python","manage.py","startapp",nomModule])
        #Manipulation pour ajouter liens urls dans le repertoire du projet principal
        path = os.path.abspath(os.path.curdir)
        path = path + "\\ErpProject\\urls.py"
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_a_ajouter_urls_py_dossier_projet="\nurlpatterns.append(url(r'^{0}', include('{1}.urls')))".format(module.url_vers,nomModule)
        fichier.write(texte_a_ajouter_urls_py_dossier_projet)
        fichier.close()

        #Manipulation de views.py dans  le sous-dossier de l'application er creation de index.html
        nameModuleUp = 'MODULE_{0}'.format(unidecode.unidecode(module.nom_module.upper().replace(" ","_")))
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\views.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_views_py_dossier_application = '# -*- coding: utf-8 -*-\nfrom __future__ import unicode_literals\nfrom django.shortcuts import render, redirect\nfrom django.http import HttpResponse, HttpRequest, HttpResponseRedirect, JsonResponse\nfrom django.contrib.auth import authenticate, login, logout\nfrom django.contrib.auth.models import User, Group\nfrom django.template import loader\nfrom django.views import generic\nfrom django.views.generic.edit import CreateView, UpdateView, DeleteView\nfrom django.urls import reverse_lazy, reverse\nfrom django.contrib import messages\nfrom django.utils import timezone\nfrom django.core import serializers\nfrom random import randint\nfrom django.core.mail import send_mail\nfrom django.conf import settings\nfrom django.core.files.storage import FileSystemStorage\nfrom django.core.files.base import ContentFile\nfrom django.core.files.storage import default_storage\nfrom ErpBackOffice.utils.identite import identite\nfrom ErpBackOffice.utils.tools import ErpModule\nimport datetime\nimport json\nfrom django.db import transaction\nfrom ErpBackOffice.dao.dao_place import dao_place\nfrom ErpBackOffice.dao.dao_compte import dao_compte\nfrom ErpBackOffice.dao.dao_module import dao_module\nfrom ErpBackOffice.dao.dao_devise import dao_devise\n\ndef get_index(request):\n\tcontext = {{"title" : "{0}","utilisateur" : identite.utilisateur(request),"modules" : dao_module.toListModulesInstalles(),"module" : ErpModule.{1},"menu" : 1}}\n\ttemplate = loader.get_template("ErpProject/{2}/index.html")\n\treturn HttpResponse(template.render(context, request))'.format(module.description,nameModuleUp,nomModule)
        texte_a_ajouter_views_py_dossier_application = texte_a_ajouter_views_py_dossier_application + '\n\ndef get_dashboard(request):\n\tcontext = {{"title" : "{0}","utilisateur" : identite.utilisateur(request),"modules" : dao_module.toListModulesInstalles(),"module" : ErpModule.{1},"menu" : 1}}\n\ttemplate = loader.get_template("ErpProject/{2}/dashboard.html")\n\treturn HttpResponse(template.render(context, request))'.format(module.description,nameModuleUp,nomModule)
        
        fichier.write(texte_a_ajouter_views_py_dossier_application)
        fichier.close()


        #Manipulation pour la creation fichier urls.py dans le sous-dossier de l'Application
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\urls.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_urls_py_dossier_application = '''from django.conf.urls import include, url\nfrom . import views\nurlpatterns = [
        url(r'^$', views.get_index, name='module_{0}_index'),
        ]
    '''.format(module.nom_module.lower().replace(" ","_"))

        fichier.write(texte_a_ajouter_urls_py_dossier_application)
        fichier.close()

        #Création fichier index.html
        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}".format(nomModule)
        os.mkdir(utils.format_path(path))
        path = path + "\\index.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        fichier.close()



        #Manipulation fichier index.html
        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}\\index.html".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_index_html_dossier_template = '''{{% extends "ErpProject/{0}/shared/layout.html" %}}
    {{% block page %}}
    {{% load static %}}

    <div class="container">
        <h3 class="pb-3 mb-4 font-italic border-bottom">
            {1}
        </h3>

        <div class="row">
            <div class="col-md-4">
                <h4>Page Generated avec Succes</h4>
            </div>
        </div>
    </div>
    {{% endblock %}}
    '''.format(nomModule,"Module GenerE via le Framework. A vous de le configurer")
        fichier.write(texte_a_ajouter_index_html_dossier_template)
        fichier.close()


        #Parametrage dans Application Settings à propos de ce module comme new Installed Apps
        path = os.path.abspath(os.path.curdir)
        path = path + "\\ErpProject\\settings.py"
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_a_ajouter_settings_py_dossier_projet="\nINSTALLED_APPS.append('{0}')".format(nomModule)
        fichier.write(texte_a_ajouter_settings_py_dossier_projet)
        fichier.close()

        #AJOUT dans ErpModule.utils
        path = os.path.abspath(os.path.curdir)
        path = path + "\\ErpBackOffice\\utils\\tools.py"
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_a_ajouter_tools_py_dossier_backoffice="\n    {0}={1}".format(nameModuleUp,module.numero_ordre)
        fichier.write(texte_a_ajouter_tools_py_dossier_backoffice)
        fichier.close()

        #Création et Initialisation du fichier Model py
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\models.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_models_py_dossier_application = "# -*- coding: utf-8 -*-\nfrom __future__ import unicode_literals\nfrom django.db import models\n\n# Create your models here.\n"
        fichier.write(texte_a_ajouter_models_py_dossier_application)
        fichier.close()

        #Création et Initialisation du fichier Admin py
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\admin.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_admin_py_dossier_application = "from django.contrib import admin\nfrom . import models\n"
        fichier.write(texte_a_ajouter_admin_py_dossier_application)
        fichier.close()
    except Exception as e:
        #print("ERREUR GENERATION FILES OF THIS MODULE!")
        #print(e)
        pass






#TEST
def get_creer_test(request):
    permission_number = 559
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response
    try:
        context = {
            'title' : 'Generer un test unitaire',
            'utilisateur' : utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'modules' : modules,
            "module" : vars_module,
            'sous_modules': sous_modules,
            
            'menu' : 5
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/test/generate.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR MODEL ON GET CREER TEST")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_modules'))


def post_creer_test(request):
    try:
        #Création du Module dans la table module

        nomModule = request.POST["nom_module"]
        name_modele = request.POST["name_modele"]

        list_fonction = request.POST.getlist("name_modele",None)

        try:
            path = os.path.abspath(os.path.curdir)
            path = path + "\\{0}\\tests\\unit".format(nomModule)
            #Creation du dossier tests et unit si ceux ci n'existent pas
            os.makedirs(path)
        except Exception as e:
            pass


        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass

            #Standardisation denomination modele
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModele = 'Model_{0}'.format(nom_modele)

            nomModeleEnUn = ""
            nommodelesplit = nom_modele.split("_")
            i = 0
            while i < len(nommodelesplit):
                nomModeleEnUn = nomModeleEnUn + nommodelesplit[i].capitalize()
                i = i + 1

            path = os.path.abspath(os.path.curdir)
            path = path + "\\{0}\\tests\\unit".format(nomModule)
            path = path + "\\test_dao_{0}.py".format(nomModeleEnUn.lower())
            fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')

            texte_contenu ="# -*- coding: utf-8 -*-\nfrom __future__ import unicode_literals\nfrom django.test import TestCase\nfrom pprint import pprint\nfrom {0}.dao.dao_{1} import dao_{1}\nfrom ErpBackOffice.models import {2}\nfrom ErpBackOffice.models import Model_Personne\n\nclass Test_Dao{3} (TestCase):\n".format(nomModule,nom_modele.lower(),nomModele,nomModeleEnUn)
            texte_contenu = texte_contenu + "\n\t@classmethod\n\tdef setUpTestData(cls):\n\t\t#Creation d'un Auteur\n\t\tModel_Personne.objects.create(nom_complet = 'Serena')\n\t\t#Enregistrement dans la BD Test de deux objets {0} pour besoin de test\n\t\t#A CONFIGURER LES PARAMETRES ENTRE PARENTHESE SELON L'EXEMPLE  !!!!!!\n\t\t{0}.objects.create()\n\t\t{0}.objects.create()\n\n\tdef setUp(self):\n\t\t#Affectation de l'auteur dans une variable\n\t\tself.auteur = Model_Personne.objects.get(pk=1)\n".format(nomModele)
            texte_contenu = texte_contenu + "\n\tdef test_CreateSave{0}(self):\n\t\tpprint ('test_Create{0}')\n\t\t#A Configurer\n\t\tobjet = dao_{1}.toCreate{0}()\n\t\tself.assertIsInstance(dao_{1}.toSave{0}(self.auteur,objet),{2})\n\t\tp#print('SUCCES')\n".format(nomModeleEnUn,nom_modele,nomModele)
            texte_contenu = texte_contenu + "\n\tdef test_Update{0}(self):\n\t\tpprint ('test_Update{0}')\n\t\tobjet = dao_{1}.toCreate{0}()\n\t\tself.assertIsInstance(dao_{1}.toUpdate{0}(1,objet),{2})\n\t\tp#print({2}.objects.get(pk=1))\n\t\tp#print('SUCCES')\n".format(nomModeleEnUn,nom_modele,nomModele)
            texte_contenu = texte_contenu + "\n\tdef test_toGet{0}(self):\n\t\tpprint ('test_toGet{0}')\n\t\tself.assertIsInstance(dao_{1}.toGet{0}(1),{2})\n\t\tp#print('SUCCES')\n".format(nomModeleEnUn,nom_modele,nomModele)
            texte_contenu = texte_contenu + "\n\tdef test_toGetList{0}(self):\n\t\tpprint ('test_toGetList{0}')\n\t\tself.assertIn(dao_{1}.toGet{0}(1),dao_{1}.toList{0}())\n\t\tp#print('SUCCES')".format(nomModeleEnUn,nom_modele,nomModele)
            texte_contenu = texte_contenu + "\n\tdef test_toDelete{0}(self):\n\t\tpprint ('test_toDelete{0}')\n\t\tself.assertTrue(dao_{1}.toDelete{0}(1))\n\t\tp#print('SUCCES')".format(nomModeleEnUn,nom_modele)

            fichier.write(texte_contenu)
            fichier.close()

        #print("End of Test Unit")

        #Fin des Opérations, on retourne Http
        messages.add_message(request, messages.SUCCESS, "Test Unitaire du {0} généré avec succès".format(nomModule))
        return HttpResponseRedirect(reverse('module_configuration_generate_test'))

    except Exception as e:
        #print("ERREUR !")
        #print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_generate_test'))


#TEST
def get_creer_selenium(request):
    permission_number = 563
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response
    try:

        context = {
            'title' : 'Generer un test selenium',
            'utilisateur': utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'sous_modules': sous_modules,
            'modules' : modules,
            "module" : vars_module,
            
            'menu' : 5
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/test/generate_selenium.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR MODEL ON GET CREER SELENIUM")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_modules'))


def post_creer_selenium(request):
    try:
        #Création du Module dans la table module

        nomModule = request.POST["nom_module"]
        list_fonction = request.POST.getlist("name_modele",None)
        list_properties = request.POST.getlist("property",None)

        try:
            path = os.path.abspath(os.path.curdir)
            path = path + "\\{0}\\tests\\webselenium".format(nomModule)
            #Creation du dossier tests et unit si ceux ci n'existent pas
            os.makedirs(path)
        except Exception as e:
            pass

        #Standardisation denomination module
        nomModulepart1 = nomModule[:6]
        nomModulepart2 = nomModule[6:]
        nomModuleCorrect = nomModulepart1.lower() + "_"+ nomModulepart2.lower()


        #Creation fichier init.py
        path2 = path + "\\__init__.py"
        fichier = codecs.open(utils.format_path(path2),"a", encoding='utf-8')
        fichier.close()

        #Ecrivons le fichier test.py
        path = path + "\\test.py"
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        selenium_driver = os.path.abspath(os.path.curdir) + "\\selenium_driver"
        selenium_driver = utils.format_path(selenium_driver)

        texte_write = "# -*- coding: utf-8 -*-\nfrom __future__ import unicode_literals\nfrom django.test import TestCase\n\nfrom selenium import webdriver\nfrom selenium.webdriver.common.keys import Keys\nfrom selenium.webdriver.firefox.firefox_binary import FirefoxBinary\nfrom selenium.webdriver.common.desired_capabilities import DesiredCapabilities\nfrom selenium.webdriver.support.ui import Select\nimport time\nfrom django.core.urlresolvers import reverse\nfrom django.test import Client\nfrom ErpBackOffice.models import Model_Personne\nfrom pprint import pprint\nfrom ErpBackOffice import views\n"
        texte_write = texte_write + "\n#login\n# TEST DE CONNEXION\n# A CONFIGURER EN FONCTION DU CHEMIN DE VOTRE NAVIGATEUR\nbinary = FirefoxBinary(r'C:\\Program Files\\Mozilla Firefox\\firefox.exe')\ncaps = DesiredCapabilities.FIREFOX.copy()\ncaps['marionette'] = True\nnavigator = webdriver.Firefox(firefox_binary=binary,capabilities=caps, executable_path='{0}\\geckodriver')\nnavigator.get('http://127.0.0.1:8000/utilisateur/connexion/')\nnavigator.maximize_window()\n# LOGIN\nloginBox = navigator.find_element_by_name('email')\nloginBox.clear()\nloginBox.send_keys('admin')\npasswordBox = navigator.find_element_by_name('password')\npasswordBox.clear()\npasswordBox.send_keys('Password01')\ntime.sleep(1)\nnavigator.find_element_by_id('btn_connecter').click()\nnavigator.find_element_by_id('navbar_link_{1}').click()\n#print('\\nDEBUT TEST DU {2} \\n')\n".format(selenium_driver,nomModuleCorrect,nomModule)

        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            try:
                nom_fonction = list_fonction[i]
            except Exception as e:
                pass
            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nom_modele_low = nom_modele.lower()
            nomModele = 'Model_{0}'.format(nom_modele)

            texte_write = texte_write + "\n#_{3} Appel du scénario pour le model {0}\nfrom {1}.tests.webselenium.model_{0} import {2}\n{2}(navigator)\n#print('\\n')\n".format(nom_modele_low,nomModule,nomModele,str(i+1))


        texte_write = texte_write + "#print('\\nFIN TEST DU {0} \\n')".format(nomModule.upper())
        fichier.write(texte_write)
        fichier.close()

        #CREATION DES SCENARIOS PROPRE A CHAQUE MODEL
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\tests\\webselenium".format(nomModule)
        for i in range(0,len(list_fonction)):
            nom_fonction = ""
            properties = ""
            try:
                nom_fonction = list_fonction[i]
                properties = list_properties[i]
            except Exception as e:
                pass

            nom_modele = nom_fonction.replace(" ","_").capitalize()
            nomModeleEnUn = ""
            nommodelesplit = nom_modele.split("_")
            i = 0
            while i < len(nommodelesplit):
                nomModeleEnUn = nomModeleEnUn + nommodelesplit[i].capitalize()
                i = i + 1
            nom_modele_low = nom_modele.lower()
            nomModele = 'Model_{0}'.format(nom_modele)



            path2 = path + "\\{0}.py".format(nomModele.lower())
            fichier = codecs.open(utils.format_path(path2),"a", encoding='utf-8')

            texte_contenu ="# -*- coding: utf-8 -*-\nfrom __future__ import unicode_literals\nfrom django.test import TestCase\nfrom selenium import webdriver\nfrom selenium.webdriver.common.keys import Keys\nfrom selenium.webdriver.firefox.firefox_binary import FirefoxBinary\nfrom selenium.webdriver.common.desired_capabilities import DesiredCapabilities\nfrom selenium.webdriver.support.ui import Select\nimport time\nfrom django.core.urlresolvers import reverse\nfrom django.test import Client\nfrom pprint import pprint\n\n#import dao relatif au model\nfrom ErpBackOffice.dao.dao_categorie_article import dao_{0}\n".format(nom_modele_low)
            texte_contenu = texte_contenu + "\nclass {1} (TestCase):\n\n\tdef __init__(self, navigator):\n\t\t#la liste des {0}\n\t\tnavigator.find_element_by_id('link_list_model_{0}').click()\n\t\t#test de création d une {0}\n\t\t{2} = {3}\n\t\t{0} = self.creation(navigator, {2})\n\n\t\t#retour à la liste des {0}\n\t\tnavigator.find_element_by_id('link_list_model_{0}').click()\n\n\t\t#test de modification d'un {0}\n\t\t{2} = {4}\n\t\tself.modification(navigator, {2}, {0})\n".format(nom_modele_low,nomModele,properties,"one example","another example")
            texte_contenu = texte_contenu + "\n\n\tdef creation(self, navigator, {1}):\n\t\t#print('test concernant la \"Création {0}\"')\n\t\tnavigator.find_element_by_id('btn_creer').click()\n\n\t\t#input_{1}\n\t\tinput_{1} = navigator.find_element_by_name('{1}')\n\t\tinput_{1}.clear()\n\t\tinput_{1}.send_keys({1})\n\n\t\treponse = self.verification({1})\n\t\tif reponse != None:\n\t\t\t#print('Test de création d une {0} est un échec car existe déjà')\n\t\t\treturn reponse\n\t\telse:\n\t\t\tnavigator.find_element_by_id('btn_valider').click()\n\t\t\treponse = self.verification({1})\n\t\t\tif reponse != None:\n\t\t\t\t#print('Test de création d une {0} est une réussite')\n\t\t\t\treturn reponse\n\t\t\telse:\n\t\t\t\t#print('Test de création d une {0} est un échec')\n\t\t\t\texit()\n".format(nom_modele_low,properties)
            texte_contenu = texte_contenu + "\n\tdef verification(self, {1}):\n\t\t#print('verification de l enregistrement')\n\t\treturn dao_{0}.toGet{2}By{3}({1})\n".format(nom_modele_low,properties,nomModeleEnUn,properties.capitalize())
            texte_contenu = texte_contenu + "\n\tdef modification(self, navigator, {1}, {0}):\n\t\t#print('test concernant la \"Modification {0}\"')\n\n\t\titem_id = 'link_item_' + str({0}.id)\n\t\tnavigator.find_element_by_id(item_id).click()\n\t\tnavigator.find_element_by_id('btn_modifier').click()\n\n\t\t#input_{1}\n\t\tinput_{1} = navigator.find_element_by_name('{1}')\n\t\tinput_{1}.clear()\n\t\tinput_{1}.send_keys({1})\n\t\treponse = self.verification({1})\n\t\tif reponse != None:\n\t\t\t#print('Test de modification d une {0} est un échec car existe déjà')\n\t\t\treturn reponse\n\t\telse:\n\t\t\tnavigator.find_element_by_id('btn_valider').click()\n\t\t\treponse = self.verification({1})\n\t\t\tif reponse != None:\n\t\t\t\t#print('Test de modification d une {0} est une réussite')\n\t\t\telse:\n\t\t\t\t#print('Test de modification d une {0} est un échec')\n\t\t\t\texit()".format(nom_modele_low,properties)

            fichier.write(texte_contenu)
            fichier.close()

        #print("End of Test Selenium")

        #Fin des Opérations, on retourne Http
        messages.add_message(request, messages.SUCCESS, "Test Selenium du {0} généré avec succès".format(nomModule))
        return HttpResponseRedirect(reverse('module_configuration_generate_selenium'))

    except Exception as e:
        #print("ERREUR !")
        #print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_generate_selenium'))


def get_creer_dao_template(request):
    permission_number = 555
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response


    try:
        context = {
            'title' : 'Nouveau dao & templates',
            'utilisateur' : utilisateur,
            
            'sous_modules': sous_modules,
            'modules':modules,
            'menu' : 4
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/module/dao_template/add.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR MODEL ON GET CREER DAO TEMPLATE")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_add_dao_template'))



def post_creer_dao_template(request):
    try:
        nomModule = request.POST['nomModule']
        nom_minim = nomModule[6:]
        nom_pattern = 'module_{0}'.format(unidecode.unidecode(nom_minim.lower().replace(" ","_")))
        nom_modele = request.POST["nom_modele"]
        list_nom_champ = request.POST.getlist("nom_champ",None)
        list_type_champ = request.POST.getlist("TypeId",None)


        #Standardisation denomination modele
        nom_modele = nom_modele.replace(" ","_").capitalize()
        nomModele = 'Model_{0}'.format(nom_modele)

        #Création Modele dans le fichier models.py
        #Creation dossier dao
        try:
            path = os.path.abspath(os.path.curdir)
            path = path + "\\{0}\\dao".format(nomModule)
            os.mkdir(utils.format_path(path))
        except Exception as e:
            pass

        nomdao="dao_{0}".format(nom_modele.lower())
        path = path + "\\{0}.py".format(nomdao)
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        texte_a_ajouter_dao_dossier_application = "from __future__ import unicode_literals\nfrom {0}.models import {1}\nfrom django.utils import timezone\n\nclass {2}(object):".format(nomModule,nomModele,nomdao)
        fichier.write(texte_a_ajouter_dao_dossier_application)
        fichier.close()
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_add_hors = "\n\tid = 0"
        #fichier.write(texte_add_hors)
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            try:
                texte_add = ""
                TypeId = int(list_type_champ[i])
                if TypeId == 1:
                    texte_add = "\n\t{0} = ''".format(nom_champ)
                elif TypeId == 2:
                    texte_add = "\n\t{0} = 0".format(nom_champ)
                elif TypeId == 3:
                    texte_add = "\n\t{0} = '2010-01-01'".format(nom_champ)
                elif TypeId == 4:
                    texte_add = "\n\t{0} = 0.0".format(nom_champ)
                elif TypeId == 5 :
                    texte_add = "\n\t{0} = False".format(nom_champ)
                elif TypeId == 6 :
                    texte_add = "\n\t{0} = None".format(nom_champ)
                    list_nom_champ[i] = '{0}_id'.format(nom_champ)
                else:
                    texte_add = "\n\t{0} = ''".format(nom_champ)
            except Exception as e:
                pass
            texte_add_hors = texte_add_hors + texte_add
            #fichier.write(texte_add)

        texte_add_hors = texte_add_hors + "\n\n\t@staticmethod\n\tdef toList():\n\t\treturn {1}.objects.all()\n\n\t@staticmethod\n\tdef toCreate(".format(nom_modele,nomModele)
        #fichier.write(texte_about_methode_dao)
        text_parenthese=""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            text_parenthese = text_parenthese + nom_champ + ","
        text_parenthese = text_parenthese[:len(text_parenthese)-1]
        text_parenthese = text_parenthese + "):"
        texte_add_hors = texte_add_hors + text_parenthese
        #fichier.write(texte_add_hors)

        texte_add_hors = texte_add_hors + "\n\t\ttry:\n\t\t\t{0} = {1}()".format(nom_modele.lower(),nomdao)
        #fichier.write(texte_about_methode_dao)
        text_parenthese = ""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            text_parenthese = text_parenthese + "\n\t\t\t{0}.{1} = {1}".format(nom_modele.lower(),nom_champ)
        texte_add_hors = texte_add_hors + text_parenthese
        #fichier.write(text_parenthese)
        texte_add_hors = texte_add_hors + "\n\t\t\treturn {0}\n\t\texcept Exception as e:\n\t\t\t#print('ERREUR LORS DE LA CREATION DE LA {1}')\n\t\t\t#print(e)\n\t\t\treturn None\n\n\t@staticmethod\n\tdef toSave(auteur, objet_dao_{2}):\n\t\ttry:\n\t\t\t{0}  = Model_{2}()".format(nom_modele.lower(),nom_modele.upper(),nom_modele.capitalize())
        #fichier.write(texte_about_methode_dao)

        text_parenthese = ""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            if nom_champ.upper() == "DATE_CREATION":
                text_parenthese = text_parenthese + "\n\t\t\t{0}.date_creation = timezone.now()"
            elif nom_champ.upper() == "CREATION_DATE":
                text_parenthese = text_parenthese + "\n\t\t\t{0}.creation_date = timezone.now()"
            else:
                text_parenthese = text_parenthese + "\n\t\t\t{0}.{1} = objet_dao_{2}.{1}".format(nom_modele.lower(),nom_champ.lower(),nom_modele.capitalize())

        texte_add_hors = texte_add_hors + text_parenthese

        #texte_add_hors = texte_add_hors + "\n\t\t\t{0}.created_at = timezone.now()\n\t\t\t{0}.updated_at = timezone.now()\n\t\t\t{0}.auteur_id = auteur.id\n".format(nom_modele.lower())
        #fichier.write(text_parenthese)

        texte_add_hors = texte_add_hors + "\n\t\t\t{0}.auteur_id = auteur.id\n\t\t\t{0}.save()\n\t\t\treturn {0}\n\t\texcept Exception as e:\n\t\t\t#print('ERREUR LORS DE L ENREGISTREMENT DE LA {1}')\n\t\t\t#print(e)\n\t\t\treturn None\n\n\t@staticmethod\n\tdef toUpdate(id, objet_dao_{2}):\n\t\ttry:\n\t\t\t{0} = Model_{2}.objects.get(pk = id)".format(nom_modele.lower(),nom_modele.upper(),nom_modele.capitalize())

        #fichier.write(texte_about_methode_dao)

        text_parenthese = ""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            text_parenthese = text_parenthese + "\n\t\t\t{0}.{1} =objet_dao_{2}.{1}".format(nom_modele.lower(),nom_champ,nom_modele.capitalize())
        #fichier.write(text_parenthese)
        texte_add_hors = texte_add_hors + text_parenthese
        #texte_add_hors = texte_add_hors + "\n\t\t\t{0}.updated_at = timezone.now()".format(nom_modele.lower())

        texte_add_hors = texte_add_hors + "\n\t\t\t{0}.save()\n\t\t\treturn {0}\n\t\texcept Exception as e:\n\t\t\t#print('ERREUR LORS DE LA MODIFICATION DE LA {1}')\n\t\t\t#print(e)\n\t\t\treturn None\n\n\t@staticmethod\n\tdef toGet(id):\n\t\ttry:\n\t\t\treturn Model_{2}.objects.get(pk = id)\n\t\texcept Exception as e:\n\t\t\treturn None\n\n\t@staticmethod\n\tdef toDelete(id):\n\t\ttry:\n\t\t\t{0} = Model_{2}.objects.get(pk = id)\n\t\t\t{0}.delete()\n\t\t\treturn True\n\t\texcept Exception as e:\n\t\t\treturn False".format(nom_modele.lower(),nom_modele.upper(),nom_modele.capitalize())
        fichier.write(texte_add_hors)
        fichier.close()

        #Creation des fonctions CRUD Dans views.py
        nameModuleUp = 'MODULE_{0}'.format(unidecode.unidecode(nom_minim.upper().replace(" ","_")))

        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\views.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        texte_addenda_views_py_appli = "\nfrom {4}.dao.{1} import {1}\n\ndef get_lister_{0}(request):\n\tis_connect = identite.est_connecte(request)\n\tif is_connect == False: return HttpResponseRedirect(reverse('backoffice_connexion'))\n\n\tmodel = {1}.toList{2}()\n\tcontext ={{'title' : 'Liste de {0}','model' : model,'utilisateur': utilisateur,'modules' : dao_module.toListModulesInstalles(),'module' : ErpModule.{3},'menu' : 1}}\n\ttemplate = loader.get_template('ErpProject/{4}/{0}/list.html')\n\treturn HttpResponse(template.render(context, request))".format(nom_modele.lower(),nomdao,nom_modele.capitalize(),nameModuleUp,nomModule)
        #fichier.write(texte_addenda_views_py_appli)
        #texte_addenda_views_py_appli = ""
        #I Stop Here, what to do is simple Continue with dao
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\ndef get_creer_{0}(request):\n\tis_connect=identite.est_connecte(request)\n\tif is_connect == False: return HttpResponseRedirect(reverse('backoffice_connexion'))\n\n\tcontext ={{'title' : 'Nouvelle {0}','utilisateur': utilisateur,'modules' : dao_module.toListModulesInstalles(),'module' : ErpModule.{1},'menu' : 2}}\n\ttemplate = loader.get_template('ErpProject/{2}/{0}/add.html')\n\treturn HttpResponse(template.render(context, request))".format(nom_modele.lower(),nameModuleUp,nomModule)
        #fichier.write(texte_addenda_views_py_appli)

        #texte_addenda_views_py_appli=""
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\ndef post_creer_{0}(request):\n\n\ttry:".format(nom_modele.lower())
        #fichier.write(texte_addenda_views_py_appli)
        texte_boucle=""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            texte_boucle = texte_boucle + "\n\t\t{0} = request.POST['{0}']".format(nom_champ)
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + texte_boucle
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\t\tauteur = identite.utilisateur(request)"
        #fichier.write(texte_boucle)

        #texte_addenda_views_py_appli=""
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\n\t\t{0}={1}.toCreate{2}(".format(nom_modele.lower(),nomdao,nom_modele.capitalize())
        text_parenthese = ""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            text_parenthese = text_parenthese + "{0},".format(nom_champ)
        text_parenthese = text_parenthese[:len(text_parenthese)-1]
        text_parenthese = text_parenthese + ")"
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + text_parenthese
        #fichier.write(text_parenthese)

        #texte_addenda_views_py_appli=""
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\t\t{0}={1}.toSave{2}(auteur, {0})".format(nom_modele.lower(),nomdao,nom_modele.capitalize())
        #fichier.write(texte_addenda_views_py_appli)
        #texte_addenda_views_py_appli=""
        texte_addenda_views_py_appli= texte_addenda_views_py_appli + "\n\t\treturn HttpResponseRedirect(reverse('{1}_list_{0}'))\n\texcept Exception as e:\n\t\t#print('Erreur lors de l enregistrement')\n\t\tauteur = identite.utilisateur(request)\n\t\tmonLog.error('{{}} :: {{}} :: \\nERREUR LORS DU POST CREER {2} \\n {{}}'.format(auteur.nom_complet, module,e))\n\t\t#print(e)\n\t\treturn HttpResponseRedirect(reverse('{1}_add_{0}'))".format(nom_modele.lower(),nom_pattern,nom_modele.upper())
        #texte_addenda_views_py_appli=""
        texte_addenda_views_py_appli= texte_addenda_views_py_appli + "\n\n\ndef get_details_{0}(request,ref):\n\tis_connect = identite.est_connecte(request)\n\tif is_connect == False: return HttpResponseRedirect(reverse('backoffice_connexion'))".format(nom_modele.lower())
        #fichier.write(texte_addenda_views_py_appli)

        texte_addenda_views_py_appli= texte_addenda_views_py_appli + "\n\ttry:\n\t\tref=int(ref)\n\t\t{0}={1}.toGet{2}(ref)\n\t\ttemplate = loader.get_template('ErpProject/{3}/{0}/item.html')\n\t\tcontext ={{'title' : 'Details d une {0}','{0}' : {0},'utilisateur': utilisateur,'modules' : dao_module.toListModulesInstalles(),'module' : ErpModule.{4},'menu' : 4}}\n\n\t\treturn HttpResponse(template.render(context, request))\n\texcept Exception as e:\n\t\t#print('Erreut Get Detail')\n\t\tauteur = identite.utilisateur(request)\n\t\tmonLog.error('{{}} :: {{}} :: \\nERREUR LORS DU GET DETAILS {6} \\n {{}}'.format(auteur.nom_complet, module,e))\n\t\t#print(e)\n\t\treturn HttpResponseRedirect(reverse('{5}_list_{0}'))".format(nom_modele.lower(),nomdao,nom_modele.capitalize(),nomModule,nameModuleUp,nom_pattern,nom_modele.upper())

        #Ajout de updaye
        texte_addenda_views_py_appli= texte_addenda_views_py_appli +"\ndef get_modifier_{0}(request,ref):\n\tis_connect=identite.est_connecte(request)\n\tif is_connect == False: return HttpResponseRedirect(reverse('backoffice_connexion'))\n\n\tref = int(ref)\n\tmodel = {1}.toGet{2}(ref)\n\tcontext ={{'title' : 'Modifier {2}','model':model, 'utilisateur': utilisateur,'modules' : dao_module.toListModulesInstalles(),'module' : ErpModule.{3},'menu' : 2}}\n\ttemplate = loader.get_template('ErpProject/{4}/{0}/update.html')\n\treturn HttpResponse(template.render(context, request))".format(nom_modele.lower(),nomdao,nom_modele.capitalize(),nameModuleUp,nomModule)

        #POST MODIFIER

        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\ndef post_modifier_{0}(request):\n\n\tid = int(request.POST['ref'])\n\ttry:".format(nom_modele.lower())
        texte_boucle=""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            texte_boucle = texte_boucle + "\n\t\t{0} = request.POST['{0}']".format(nom_champ)
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + texte_boucle
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\t\tauteur = identite.utilisateur(request)"
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\n\t\t{0}={1}.toCreate{2}(".format(nom_modele.lower(),nomdao,nom_modele.capitalize())
        text_parenthese = ""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            text_parenthese = text_parenthese + "{0},".format(nom_champ)
        text_parenthese = text_parenthese[:len(text_parenthese)-1]
        text_parenthese = text_parenthese + ")"
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + text_parenthese
        texte_addenda_views_py_appli = texte_addenda_views_py_appli + "\n\t\t{0}={1}.toUpdate{2}(id, {0})".format(nom_modele.lower(),nomdao,nom_modele.capitalize())
        #fichier.write(texte_addenda_views_py_appli)
        #texte_addenda_views_py_appli=""
        texte_addenda_views_py_appli= texte_addenda_views_py_appli + "\n\t\treturn HttpResponseRedirect(reverse('{1}_list_{0}'))\n\texcept Exception as e:\n\t\t#print('Erreur lors de l enregistrement')\n\t\tauteur = identite.utilisateur(request)\n\t\tmonLog.error('{{}} :: {{}} :: \\nERREUR LORS DU POST MODIFIER {2} \\n {{}}'.format(auteur.nom_complet, module,e))\n\t\t#print(e)\n\t\treturn HttpResponseRedirect(reverse('{1}_add_{0}'))".format(nom_modele.lower(),nom_pattern,nom_modele.upper())
        #texte_addenda_views_py_appli=""


        #END POST MODIFIER

        fichier.write(texte_addenda_views_py_appli)
        fichier.close()
        #Templates add list item
        try:
            path = os.path.abspath(os.path.curdir)
            path = path + "\\templates\\ErpProject\\{0}\\{1}".format(nomModule,nom_modele.lower())
            os.mkdir(utils.format_path(path))
        except Exception as e:
            pass
        path = path + "\\list.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        #print('in')

        texteTemplate = '''{{% extends "ErpProject/{0}/shared/layout.html" %}}
{{% block page %}}
<div class="row">
    <ul class="breadcrumb">
        <li><a><span class="mif-home"></span></a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_index' %}}">Accueil</a></li>
        <li>{{{{ title }}}}</li>
    </ul>
</div>


<div class="row">
    <div class="col-lg-12">
        <h2>Liste {1}</h2>
        <strong style="float: right;color: grey;opacity: 0.4;margin-top: -30px;">{{% now "jS F Y H:i" %}}</strong>

        <div class="separ" style="background-color: grey;opacity: 0.2"></div>

        <div class="panel panel-default" style="border: none; margin-top: 1rem;">
            <div class="panel panel-body" style="background-color:#f5f5f5;border: none;border-radius: none;">

                        <button onclick="javascript:window.location.assign('{{% url '{2}_add_{3}' %}}')" class="button small-button rounded primary chargement-au-click">Creer</button>

                <br>
                <br>


       <table id="example" class="display nowrap border bordered striped" cellspacing="0" style="width:100%">
            <thead>
                <tr>

        #fichier.write(texte)
                    <th style="width: 20px; background-color:#2e416a; white"></th><th>Id</th>'''.format(nomModule,nom_minim,nom_pattern,nom_modele.lower())
        textebcl=""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            textebcl=textebcl + "<th>{0}</th>".format(nom_champ)
        texteTemplate = texteTemplate + textebcl
        #fichier.write(textebcl)

        texteTemplate= texteTemplate + '''
                </tr>
            </thead>
            <tbody>
                {{% for item in model %}}
                    <tr>
                        <td>
                            <label class="small-check">
                                <input type="checkbox">
                                <span class="check"></span>
                            </label>
                        </td>
                        <td>
                            <a class="lien chargement-au-click" href="{{% url '{0}_detail_{1}' item.id %}}">{{{{ item.id }}}}</a>
                        </td>'''.format(nom_pattern,nom_modele.lower())
        #fichier.write(texte2)
        textebcl=""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            textebcl=textebcl + "\n<td>{{{{item.{0}}}}}</td>".format(nom_champ)
        texteTemplate = texteTemplate + textebcl
        #fichier.write(textebcl)
        texteTemplate = texteTemplate + '''
                </tr>
                    {% endfor %}

            </tbody>
        </table>


            </div>
        </div>
    </div>
    <!-- /.col-lg-12 -->
</div>
<script>
    $(document).ready(function() {
        $('#example').DataTable( {
            dom: 'Bfrtip',
            buttons: [
                {
                    extend: 'copyHtml5',
                    text: "<span class='mdi mdi-content-copy' style='font-size:1rem;color:#424892;'></span> Copy",
                    titleAttr: "Copier"
                },
                {
                    extend: 'excelHtml5',
                    text: "<span class='mdi mdi-file-excel' style='font-size:1rem;color:#236e43;'></span> Excel",
                    titleAttr: "Exporter en excel"
                },
                {
                    extend: 'csvHtml5',
                    text: "<span class='mdi mdi-file-outline' style='font-size:1rem;color:#9c27b0;'></span> CSV",
                    titleAttr: "Exporter en csv"
                },
                {
                    extend: 'pdfHtml5',
                    text: "<span class='mdi mdi-file-pdf' style='font-size:1rem;color:#f44336;'></span> PDF",
                    titleAttr: "Exporter en pdf"
                },
                //'copyHtml5',
                //'excelHtml5',
                //'csvHtml5',
                //'pdfHtml5'
            ]
        } );
    } );
</script>
{% endblock %}'''
        fichier.write(texteTemplate)
        fichier.close()
        #CC

        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}\\{1}".format(nomModule,nom_modele.lower())
        path = path + "\\add.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        #print('in1')
        texteTemplateLayout = '''{{% extends "ErpProject/{0}/shared/layout.html" %}} {{% block page %}}
<div class="row">
    <ul class="breadcrumb">
        <li><a><span class="mif-home"></span></a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_index' %}}">Accueil</a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_list_{3}' %}}">Liste</a></li>
        <li>{{{{ title }}}}</li>
    </ul>
</div>

<div class="row">
    <div class="col-lg-12">
        <h2>Ajouter un {1}</h2>
        <strong style="float: right;color: grey;opacity: 0.4;margin-top: -30px;">{{% now "jS F Y H:i" %}}</strong>

        <div class="separ" style="background-color: grey;opacity: 0.2"></div>

        <div class="panel panel-default" style="border: none; margin-top: 1rem;">
            <div class="panel panel-body" style="background-color:#f5f5f5;border: none;border-radius: none;">
                <div class="row only-on-small-screen">
                    <h2 class="text-light no-margin-left">{{{{ title }}}}</h2>
                </div>


                <div class="row only-on-large-screen">
                    <button onclick="javascript:document.getElementById('submit').click()" class="button small-button rounded primary_color_{{{{module.name|lower}}}}">
                        Valider
                    </button>
                    <button onclick="javascript:window.location.assign('{{% url '{2}_list_{3}' %}}')" class="button small-button rounded chargement-au-click" style="margin-left: 5px">Annuler</button>
                </div>
                <div class="row only-on-small-screen">
                    <button onclick="javascript:document.getElementById('submit').click()" class="button large-button rounded primary">Valider</button>
                    <button onclick="javascript:window.location.assign('{{% url '{2}_list_{3}' %}}')" class="button large-button rounded chargement-au-click" style="margin-left: 5px">Annuler</button>
                </div>

                <hr class="hr-ligne">

                {{% if messages %}}
                    {{% for message in messages %}}
                        {{% if message.level == DEFAULT_MESSAGE_LEVELS.ERROR %}}
                            <div class="row" style="margin:10px 0">
                                <div class="col-md-6 padding10 fg-white" style="background-color:#ff6a00;">
                                    <div class="row" style="width: 32px;">
                                        <span class="mif-info mif-2x"></span>
                                    </div>
                                    <div class="row" style="margin-left: 10px">
                                        <span class="notify-title"><b>Information :</b></span><br>
                                        <span class="notify-text">
                                            {{{{ message }}}}
                                        </span>
                                    </div>
                                </div>
                            </div>
                        {{% endif %}}
                    {{% endfor %}}
                {{% endif %}}

                <form id="form" method="POST" action="{{% url '{2}_post_add_{3}' %}}" data-role="validator" data-show-required-state="false" data-hint-mode="line" data-hint-background="bg-red" data-hint-color="fg-white" data-hide-error="5000" novalidate="novalidate"
                            data-on-error-input="notifyOnErrorInput" data-show-error-hint="false">
                            {{% csrf_token %}}
                            <input id="submit" type="submit" style="display: none">
                            <div class='row'>
                            '''.format(nomModule,nom_minim,nom_pattern,nom_modele.lower())

        #fichier.write(texte10)
        textebcl=""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            test = -1
            ask = nom_champ.find('_id')
            if ask == test :
                textebcl=""
                textebcl='''
                <div class="col-md-6">
                            <label>{0}</label>
                            <div class="input-control text full-size" data-role="input">
                                <input name="{1}" id="{1}" type="text" data-validate-func="required" data-validate-hint="Entrez {1} !">
                            </div>
                </div>
            '''.format(nom_champ.capitalize(),nom_champ)

            else:
                textebcl=""
                textebcl = '''
                <div class="col-md-6">
                        <label>{0}</label>
                        <div class="input-control text full-size">
                            <select name="{1}" id="{1}" onchange=""
                                data-validate-func="min"
                                data-validate-arg="1"
                                data-validate-hint="Selectionnez {0} svp.">
                                <option value="0">Selectionnez {0}</option>
                                {{% for item in model %}}
                                    <option value="{{{{ item.id }}}}"> {{{{ item.id }}}} </option>
                                {{% endfor %}}
                            </select>
                        </div>
                    </div>

                    '''.format(nom_champ.capitalize(),nom_champ)
            texteTemplateLayout = texteTemplateLayout + textebcl
        #texteTemplateLayout = texteTemplateLayout + textebcl
        #fichier.write(textebcl)
        texteTemplateLayout = texteTemplateLayout + '''
        </div>
        </form>
    </div>
        </div>
    </div>
    <!-- /.col-lg-12 -->
</div>


{% endblock %}'''
        fichier.write(texteTemplateLayout)
        fichier.close()
        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}\\{1}".format(nomModule,nom_modele.lower())
        path = path + "\\item.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        #print('six')
        texteTemplateItem='''{{% extends "ErpProject/{0}/shared/layout.html" %}}
{{% block page %}}
<div class="row">
    <ul class="breadcrumb">
        <li><a><span class="mif-home"></span></a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_index' %}}">Accueil</a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_list_{3}' %}}">Liste</a></li>
        <li>{{{{ title }}}}</li>
    </ul>
</div>

<div class="row">
    <div class="col-lg-12">
        <h2>Apercu {1}</h2>
        <strong style="float: right;color: grey;opacity: 0.4;margin-top: -30px;">{{% now "jS F Y H:i" %}}</strong>

        <div class="separ" style="background-color: grey;opacity: 0.2"></div>

        <div class="panel panel-default" style="border: none; margin-top: 1rem;">
            <div class="panel panel-body" style="background-color:#f5f5f5;border: none;border-radius: none;">

                <div class="row only-on-large-screen">
                    <button onclick="javascript:window.location.assign('{{% url '{2}_update_{3}' {3}.id %}}')" class="button small-button rounded primary_color_{{{{module.name|lower}}}}">
                        Modifier
                    </button>
                    <button onclick="javascript:window.location.assign('{{% url '{2}_list_{3}' %}}')" class="button small-button rounded chargement-au-click" style="margin-left: 5px">Annuler</button>
                </div>
                <div class="row only-on-small-screen">
                    <button onclick="javascript:window.location.assign('{{% url '{2}_update_{3}' {3}.id %}}')" class="button large-button rounded primary">Valider</button>
                    <button onclick="javascript:window.location.assign('{{% url '{2}_list_{3}' %}}')" class="button large-button rounded chargement-au-click" style="margin-left: 5px">Annuler</button>
                </div>

                <hr class="hr-ligne">


                {{% if messages %}}
                    {{% for message in messages %}}
                        {{% if message.level == DEFAULT_MESSAGE_LEVELS.ERROR %}}
                            <div class="row" style="margin:10px 0">
                                <div class="row cells12 padding10 fg-white" style="background-color:#ff6a00;">
                                    <div class="cell colspan3" style="width: 32px;">
                                        <span class="mif-info mif-2x"></span>
                                    </div>
                                    <div class="cell colspan9" style="margin-left: 10px">
                                        <span class="notify-title"><b>Information :</b></span><br>
                                        <span class="notify-text">
                                            {{{{ message }}}}
                                        </span>
                                    </div>
                                </div>
                            </div>
                        {{% endif %}}
                        {{% if message.level == DEFAULT_MESSAGE_LEVELS.SUCCESS %}}
                            <div class="row" style="margin:10px 0">
                                <div class="row cells12 padding10 fg-white" style="background-color:#36D900;">
                                    <div class="cell colspan3" style="width: 32px;">
                                        <span class="mif-info mif-2x"></span>
                                    </div>
                                    <div class="cell colspan9" style="margin-left: 10px">
                                        <span class="notify-title"><b>Information :</b></span><br>
                                        <span class="notify-text">
                                            {{{{ message }}}}
                                        </span>
                                    </div>
                                </div>
                            </div>
                        {{% endif %}}
                    {{% endfor %}}
                {{% endif %}}

                        <div class="row">
                            <div class="col-xs-12">
                                <div class="row padding20 no-padding-top no-padding-bottom" style="border-right: 1px solid #f7f5f5;">
                                    <p>
                    '''.format(nomModule,nom_minim,nom_pattern,nom_modele.lower())

        textebcl=""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            textebcl=textebcl + '''
            {0} :
                    <br>
                    <span class="sub-alt-header">{{{{ {1}.{2} }}}}</span>
                    <br>
                    <br>
            '''.format(nom_champ.capitalize(),nom_modele.lower(),nom_champ)
        texteTemplateItem = texteTemplateItem + textebcl
        #fichier.write(textebcl)

        texteTemplateItem = texteTemplateItem + '''
        </p>
                                </div>
                            </div>
                        </div>

           {% endblock %}}'''
        fichier.write(texteTemplateItem)
        fichier.close()

        #UPDATE HTML
        path = os.path.abspath(os.path.curdir)
        path = path + "\\templates\\ErpProject\\{0}\\{1}".format(nomModule,nom_modele.lower())
        path = path + "\\update.html"
        fichier = codecs.open(utils.format_path(path),"w", encoding='utf-8')
        #print('inNeuf')
        texteTemplateLayout = '''{{% extends "ErpProject/{0}/shared/layout.html" %}} {{% block page %}}
<div class="row">
    <ul class="breadcrumb">
        <li><a><span class="mif-home"></span></a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_index' %}}">Module {4}</a></li>
        <li><a class="chargement-au-click" href="{{% url '{2}_list_{3}' %}}">Liste</a></li>
        <li>{{{{ title }}}}</li>
    </ul>
</div>

<div class="row">
    <div class="col-lg-12">
        <h2>Modifier {1}</h2>
        <strong style="float: right;color: grey;opacity: 0.4;margin-top: -30px;">{{% now "jS F Y H:i" %}}</strong>

        <div class="separ" style="background-color: grey;opacity: 0.2"></div>

        <div class="panel panel-default" style="border: none; margin-top: 1rem;">
            <div class="panel panel-body" style="background-color:#f5f5f5;border: none;border-radius: none;">


                <div class="row only-on-small-screen">
                    <h2 class="text-light no-margin-left">{{{{ title }}}}</h2>
                </div>


                <div class="row only-on-large-screen">
                    <button onclick="javascript:document.getElementById('submit').click()" class="button small-button rounded primary_color_{{{{module.name|lower}}}}">
                        Valider
                    </button>
                    <button onclick="javascript:window.location.assign('{{% url '{2}_list_{3}' %}}')" class="button small-button rounded chargement-au-click" style="margin-left: 5px">Annuler</button>
                </div>
                <div class="row only-on-small-screen">
                    <button onclick="javascript:document.getElementById('submit').click()" class="button large-button rounded primary">Valider</button>
                    <button onclick="javascript:window.location.assign('{{% url '{2}_list_{3}' %}}')" class="button large-button rounded chargement-au-click" style="margin-left: 5px">Annuler</button>
                </div>

                <hr class="hr-ligne">

                {{% if messages %}}
                    {{% for message in messages %}}
                        {{% if message.level == DEFAULT_MESSAGE_LEVELS.ERROR %}}
                            <div class="row" style="margin:10px 0">
                                <div class="col-md-6 padding10 fg-white" style="background-color:#ff6a00;">
                                    <div class="row" style="width: 32px;">
                                        <span class="mif-info mif-2x"></span>
                                    </div>
                                    <div class="row" style="margin-left: 10px">
                                        <span class="notify-title"><b>Information :</b></span><br>
                                        <span class="notify-text">
                                            {{{{ message }}}}
                                        </span>
                                    </div>
                                </div>
                            </div>
                        {{% endif %}}
                    {{% endfor %}}
                {{% endif %}}
            <form id="form" method="POST" action="{{% url '{2}_post_update_{3}' %}}" data-role="validator" data-show-required-state="false" data-hint-mode="line" data-hint-background="bg-red" data-hint-color="fg-white" data-hide-error="5000" novalidate="novalidate"
                            data-on-error-input="notifyOnErrorInput" data-show-error-hint="false">
                            {{% csrf_token %}}
                            <input id="submit" type="submit" style="display: none">
                             <input name="ref" type="hidden" value="{{{{ model.id }}}}">
                            <div class="row">
                    '''.format(nomModule,nomModule.lower(),nom_pattern,nom_modele.lower(),nomModule.capitalize())

        #fichier.write(texte10)
        textebcl=""
        for i in range(0,len(list_nom_champ)):
            nom_champ = ""
            try:
                nom_champ = list_nom_champ[i]
                nom_champ = nom_champ.lower()
            except Exception as e:
                pass
            test = -1
            ask = nom_champ.find('_id')
            textebcl=""
            textebcl='''
                <div class="col-md-6">
                            <label>{0}</label>
                            <div class="input-control text full-size" data-role="input">
                                <input name="{1}" id="{1}" type="text" data-validate-func="required" data-validate-hint="Entrez {1} !" value = "{{{{model.{1}}}}}">
                            </div>
                </div>
            '''.format(nom_champ.capitalize(),nom_champ)


            texteTemplateLayout = texteTemplateLayout + textebcl
        #texteTemplateLayout = texteTemplateLayout + textebcl
        #fichier.write(textebcl)
        texteTemplateLayout = texteTemplateLayout + '''
        </div>
        </form>
    </div>
        </div>
    </div>
    <!-- /.col-lg-12 -->
</div>


{% endblock %}'''
        fichier.write(texteTemplateLayout)
        fichier.close()


        #END UPDATE

        #All End's up HEEERE
        path = os.path.abspath(os.path.curdir)
        path = path + "\\{0}\\urls.py".format(nomModule)
        fichier = codecs.open(utils.format_path(path),"a", encoding='utf-8')
        nomdao = "dao_{0}".format(nom_modele.lower())
        nom_pattern = 'module_{0}'.format(unidecode.unidecode(nom_minim.lower().replace(" ","_")))

        texte_a_ajouter_urls_py_dossier_ap="\nurlpatterns.append(url(r'^{0}/list', views.get_lister_{0}, name = '{1}_list_{0}'))".format(nom_modele.lower(),nom_pattern)
        #fichier.write(texte_a_ajouter_urls_py_dossier_ap)
        texte_a_ajouter_urls_py_dossier_ap= texte_a_ajouter_urls_py_dossier_ap + "\nurlpatterns.append(url(r'^{0}/add', views.get_creer_{0}, name = '{1}_add_{0}'))".format(nom_modele.lower(),nom_pattern)
        #fichier.write(texte_a_ajouter_urls_py_dossier_ap)
        texte_a_ajouter_urls_py_dossier_ap= texte_a_ajouter_urls_py_dossier_ap + "\nurlpatterns.append(url(r'^{0}/post_add', views.post_creer_{0}, name = '{1}_post_add_{0}'))".format(nom_modele.lower(),nom_pattern)
        #fichier.write(texte_a_ajouter_urls_py_dossier_ap)
        texte_a_ajouter_urls_py_dossier_ap= texte_a_ajouter_urls_py_dossier_ap + "\nurlpatterns.append(url(r'^{0}/item/(?P<ref>[0-9]+)/$', views.get_details_{0}, name = '{1}_detail_{0}'))".format(nom_modele.lower(),nom_pattern)
        texte_a_ajouter_urls_py_dossier_ap= texte_a_ajouter_urls_py_dossier_ap + "\nurlpatterns.append(url(r'^{0}/item/post_update/$', views.post_modifier_{0}, name = '{1}_post_update_{0}'))".format(nom_modele.lower(),nom_pattern)
        texte_a_ajouter_urls_py_dossier_ap= texte_a_ajouter_urls_py_dossier_ap + "\nurlpatterns.append(url(r'^{0}/item/(?P<ref>[0-9]+)/update$', views.get_modifier_{0}, name = '{1}_update_{0}'))".format(nom_modele.lower(),nom_pattern)
        fichier.write(texte_a_ajouter_urls_py_dossier_ap)
        fichier.close()

        #Fin des Opérations, on retourne Http
        messages.add_message(request, messages.SUCCESS, "Opération Réussie !")
        return HttpResponseRedirect(reverse('module_configuration_add_dao_template'))
    except Exception as e:
        #print("ERREUR !")
        #print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_add_dao_template'))


# SOUS MODULES CONTROLLERS
def get_lister_sous_modules_of_module(request, ref):
    permission_number = 554
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    ref = int(ref)
    model = dao_sous_module.toListSousModulesOf(ref)
    nom = dao_module.toGetModule(ref)
    context = {
        'title' : 'Liste des sous_modules',
        'model' : model,
        'nom_module': nom,
        "utilisateur" : utilisateur,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        "modules" : modules,
        "module" : vars_module,
        'sous_modules': sous_modules,   
        'menu' : 3
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/module/sous_module/list.html")
    return HttpResponse(template.render(context, request))

def get_creer_sous_module_of_module(request,ref):
    permission_number = 555
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    if response != None:
        return response

    ref = int(ref)
    nom_module = dao_module.toGetModule(ref)
    groupes = dao_groupe_menu.toListGroupeOfModule(ref)
    sousmodules = dao_sous_module.toListSousModulesOf(ref)
    modeles = ContentType.objects.all()
    context = {
        'title' : 'Nouveau sous-module',
        'groupes': groupes,
        'nom_module': nom_module,
        'sousmodules': sousmodules,
        'modeles': modeles,
        'utilisateur': utilisateur,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        'modules' : modules,
        "module" : vars_module,
        'sous_modules': sous_modules,    
        'menu' : 3
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/module/sous_module/add.html")
    return HttpResponse(template.render(context, request))

@transaction.atomic
def post_creer_sous_module_of_module(request,ref):
	sid = transaction.savepoint()
	try:   
		module_id = int(ref)
		moduleobject = dao_module.toGetModule(module_id)
		nom_sous_module = request.POST['nom_sous_module']
		description = request.POST['description']
		groupe = ""
		numero_ordre = request.POST['numero_ordre']	
		icon_menu = ""
		est_dashboard = False
		url_vers = request.POST['url_vers']
		model_principal_id = request.POST['model_principal_id'] if request.POST['model_principal_id'] else None
		groupe_menu_id = request.POST['groupe_menu_id'] if request.POST['groupe_menu_id'] else None 
		related_models   	= 	request.POST.getlist("related_models", [])
		auteur = identite.utilisateur(request)
		
		est_actif = False
		if "est_actif" in request.POST : est_actif = True
		
		est_model = False
		if "est_model" in request.POST : est_model = True
		
		generate_dao = False
		if "generate_dao" in request.POST : generate_dao = True
		
		generate_template = False
		if "generate_template" in request.POST : generate_template = True
		
		generate_api = False
		if "generate_api" in request.POST : generate_api = True
		
		generate_reporting = False
		if "generate_reporting" in request.POST : generate_reporting = True
		
		generate_bi = False
		if "generate_bi" in request.POST : generate_bi = True

		sousmodule = dao_sousmodule.toCreateSousmodule(module_id, nom_sous_module, description, groupe, icon_menu, url_vers, numero_ordre, est_model = est_model, est_dashboard = est_dashboard, est_actif = est_actif, model_principal_id = model_principal_id, groupe_menu_id = groupe_menu_id)
		sousmodule = dao_sousmodule.toSaveSousmodule(auteur, sousmodule)
		print("FIN CREATION SOUS MODULE")
				
		#GENERATION DAO
		if est_model and generate_dao :
			genDAOofModelContentType(model_principal_id, moduleobject.id)
			print("FIN GENERATION DAO")
		
		#GENERATION TEMPLATE
		if est_model and generate_template:
			nom_modele = genTemplateOfContentType(model_principal_id, moduleobject.id, related_models)
			print("FIN GENERATION TEMPLATE")
			permission = dao_permission.toCreatePermission(sousmodule.id, f"SUPPRIMER_{nom_modele.upper()}", dao_permission.toGetLatestNumeroOrdre() + 1)
			permission = dao_permission.toSavePermission(auteur, permission)
			print("FIN CREATION PERMISSION SUPPRIMER")
		
		#GENERATION REPORTING
		if est_model and generate_reporting:
			url_name_reporting = genReportingOfContentType(model_principal_id, moduleobject.id) 
			print("FIN GENERATION REPORTING")
			groupe_menu_reporting = dao_groupe_menu.toGetGroupeRapportOfModule(module_id) 
			if groupe_menu_reporting == None: 
				#On crée le groupe menu reporting
				designation = "Rapports"
				groupe_menu_reporting = dao_groupemenu.toCreateGroupemenu(designation,"file.svg", "Groupe Menu {}".format(designation), module.id, 8)
				groupe_menu_reporting = dao_groupemenu.toSaveGroupemenu(auteur, groupe_menu_reporting) 
				print("Groupe Menu 2 cree")
			numero_ordre = int(numero_ordre) + 30
			sousmodule_reporting = dao_sousmodule.toCreateSousmodule(module_id, "Rapport {}".format(nom_sous_module), "", "", "", url_name_reporting, numero_ordre, est_model = est_model, est_dashboard = False, est_actif = est_actif, model_principal_id = model_principal_id, groupe_menu_id = groupe_menu_reporting.id)
			sousmodule_reporting = dao_sousmodule.toSaveSousmodule(auteur, sousmodule_reporting)
			print("FIN CREATION SOUS MODULE")

		#GENERATION BI
		if est_model and generate_bi:
			url_name_bi = genBIOfContentType(model_principal_id, moduleobject.id) 
			print("FIN GENERATION BI")
			groupe_menu_bi = dao_groupe_menu.toGetGroupeAnalyseOfModule(module_id) 
			if groupe_menu_bi == None: 
				#On crée le groupe menu Analyses
				designation = "Analyses"
				groupe_menu_bi = dao_groupemenu.toCreateGroupemenu(designation,"line-chart-for-business.svg", "Groupe Menu {}".format(designation), module.id, 9)
				groupe_menu_bi = dao_groupemenu.toSaveGroupemenu(auteur, groupe_menu_bi) 
				print("Groupe Menu 3 cree")
			numero_ordre = int(numero_ordre) + 60
			sousmodule_bi = dao_sousmodule.toCreateSousmodule(module_id, "Analyse {}".format(nom_sous_module), "", "", "", url_name_bi, numero_ordre, est_model = est_model, est_dashboard = False, est_actif = est_actif, model_principal_id = model_principal_id, groupe_menu_id = groupe_menu_bi.id)
			sousmodule_bi = dao_sousmodule.toSaveSousmodule(auteur, sousmodule_bi)
			print("FIN CREATION SOUS MODULE")
					
		#GENERATION API
		if est_model and generate_api:
			genAPIOfContentType(model_principal_id, moduleobject.id)            
			print("FIN GENERATION API")
		

		#GESTION PERMISSION ET ACTION
		list_permission = request.POST.getlist('permission', None)
		list_action = request.POST.getlist('action', None)

		for i in range(0, len(list_permission)) :
			sous_module_id = sousmodule.id
			if est_model and generate_reporting and list_permission[i].startswith("ANALYSER_"):
				sous_module_id = sousmodule_reporting.id
			permission = dao_permission.toCreatePermission(sous_module_id, list_permission[i], dao_permission.toGetLatestNumeroOrdre() + 1)
			permission = dao_permission.toSavePermission(auteur, permission)
			action_string = list_action[i]
			list_action_string = action_string.split(",")
			for uneaction in list_action_string:
				action = dao_actionutilisateur.toCreateActionutilisateur(uneaction,"", "", permission.id)
				action = dao_actionutilisateur.toSaveActionutilisateur(auteur, action)
		print("FIN GESTION PERMISSION ET ACTION")
				
		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
		return HttpResponseRedirect(reverse('module_Configuration_detail_sousmodule', args=(sousmodule.id,)))
	except Exception as e:
		#print('Erreur lors de l enregistrement')
		#print(e)
		transaction.savepoint_rollback(sid)
		messages.add_message(request, messages.ERROR, e)
		return HttpResponseRedirect(reverse('module_configuration_add_sous_module', args=(ref,)))


def get_details_sous_module_of_module(request, ref, ref2):
    try:
        permission_number = 555
        modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

        if response != None:
            return response
        ref = int(ref)
        #print(ref)
        ref2 = int(ref2)
        model = dao_sous_module.toGetSousModule(ref2)

        #print(model.nom_sous_module)
        #print('template')
        context = {
            'title' : 'Details du sous-module '+ model.nom_sous_module,
            'model' : model,
            'ref': ref,
            'modules' : modules,
            "module" : vars_module,
            'sous_modules': sous_modules,           
            'utilisateur': utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'menu' : 3
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/module/sous_module/item.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_list_sous_modules" , args=(ref,)))

def get_modifier_sous_module_of_module(request, ref):
    permission_number = 556
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    if response != None:
        return response

    try:
        id = int(ref)
        module = dao_module.toGetModule(id)
        context = {
            'title' : 'Modifier %s' % module.nom_module,
            'model' : module,
            "utilisateur" : utilisateur,
            'sous_modules': sous_modules,
            "modules" : modules,
            "module" : vars_module,       
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'menu' : 3
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/module/sous_module/update.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR POUL")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_sous_modules' , args=(ref,)))

def post_modifier_sous_module_of_module(request):
    module_id = request.POST["module_id"]
    try:
        list_sous_module_id = request.POST.getlist('sous_module_id', None)
        list_numero_ordre = request.POST.getlist('numero_ordre', None)
        list_est_actif = request.POST.getlist('input_est_actif', None)

        for i in range(0, len(list_sous_module_id)):
            sous_module_id = int(list_sous_module_id[i])
            numero_ordre = int(list_numero_ordre[i])
            est_actif = False
            if int(list_est_actif[i]) == 1 : est_actif = True
            sous_module = dao_sous_module.toGetSousModule(sous_module_id)
            if sous_module != None:
                sous_module.numero_ordre = numero_ordre
                sous_module.est_actif = est_actif
                sous_module.save()
        messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
        return HttpResponseRedirect(reverse('module_configuration_details_module', args=(module_id,)))
    except Exception as e:
        #print("ERREUR !")
        #print(e)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_details_module', args=(module_id,)))

    # WORKFLOW CONTROLLERS
def get_lister_workflow(request):
    permission_number = 566
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    

    if response != None:
        return response

    workflows = dao_wkf_workflow.toListWorkflows()

    context = {
        'title' : 'Liste des workflows',
        "utilisateur" : utilisateur,
        "modules" : modules,
        'sous_modules': sous_modules,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        "model" : workflows,
        "module" : ErpModule.MODULE_CONFIGURATION,
        'menu' : 1
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/workflow/workflow/list.html")
    return HttpResponse(template.render(context, request))

def get_creer_workflow(request):
    permission_number = 567
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    contents = dao_wkf_workflow.toListObjectContenType()


    context = {
        'title' : 'Nouveau workflow',
        "utilisateur" : utilisateur,
        'contents':contents,
        'sous_modules': sous_modules,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        "modules" : modules,
        "module" : ErpModule.MODULE_CONFIGURATION,
        'menu' : 1
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/workflow/workflow/add.html")
    return HttpResponse(template.render(context, request))

@transaction.atomic
def post_creer_workflow(request):
    sid = transaction.savepoint()
    try:
        auteur =  identite.utilisateur(request)
        type_doc = request.POST["type_doc"]
        objet_id = request.POST["objet_id"]

        workflow = dao_wkf_workflow.toCreateWorkflow(type_doc, objet_id)
        workflow = dao_wkf_workflow.toSaveWorkflow(auteur,workflow)

        if workflow != None :
            messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
            transaction.savepoint_commit(sid)
            return HttpResponseRedirect(reverse('module_configuration_detail_workflow', args=(workflow.id,)))
        else :
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue lors de la création de l'utilisateur")
            return HttpResponseRedirect(reverse('module_configuration_add_workflow'))
    except Exception as e:
        #print("ERREUR !")
        #print(e)
        transaction.savepoint_rollback(sid)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_add_workflow'))

def get_details_workflow(request, ref):
    permission_number = 566
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    if response != None:
        return response

    try:
        ref = int(ref)
        workflow = dao_wkf_workflow.toGetWorkflow(ref)
        modules = dao_module.toListModulesInstalles()
        etapes = dao_wkf_etape.toListEtapeOfWorkflows(workflow.id)
        transitions = dao_wkf_transition.toListTransitionsOfWorkflow(workflow.id)

        #print("TRANSITIONS %s" % (transitions))
        context = {
            'title' : "Workflow : " + workflow.type_document,
            'model' : workflow,
            'utilisateur': utilisateur,
            'modules' : modules,
            'sous_modules': sous_modules,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            "etapes" : etapes,
            "transitions" : transitions,
            'roles_modules' : [],
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 1
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/workflow/workflow/item.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_utilisateurs'))

# ETAPE CONTROLLER
def get_creer_etape(request, ref):
    permission_number = 567
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    workflow = dao_wkf_workflow.toGetWorkflow(ref)
    exist = dao_wkf_etape.toGetEtapeInitialWorkflow(workflow.id)
    if exist == None:
        initiale_exist = False
    else:
        initiale_exist = True
    wfk_final = dao_wkf_etape.toGetEtapeFinalWorkflow(workflow.id)

    if wfk_final == None:
        final_exist = False
    else:
        final_exist = True

    context = {
        'title' : 'Nouvelle étape de workflow',
        "utilisateur" : utilisateur,
        "modules" : modules,
        'sous_modules': sous_modules,
        "module" : ErpModule.MODULE_CONFIGURATION,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        "workflow" : workflow,
        "initiale_exist" : initiale_exist,
        "final_exist": final_exist,
        'menu' : 1
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/workflow/etape/add.html")
    return HttpResponse(template.render(context, request))

@transaction.atomic
def post_creer_etape(request):
    sid = transaction.savepoint()
    try:
        auteur =  identite.utilisateur(request)
        workflow_id = request.POST["workflow_id"]
        designation = request.POST["designation"]
        label = request.POST["label"]
        est_initiale = request.POST["est_initiale"]
        est_final = request.POST["est_final"]
        est_decisive = request.POST["est_decisive"]
        initial = False
        final = False
        decisive = False

        if est_initiale == "2":
            initial = True

        if est_decisive == "2":
            decisive = True

        if est_final == "2":
            final = True

        #print(est_decisive, est_final, est_initiale)
        #print(flkdjlksjfkldj)


        etape = dao_wkf_etape.toCreateEtapeWorkflow(designation,label,workflow_id,initial, decisive, final)
        etape = dao_wkf_etape.toSaveEtapeWorkflow(auteur,etape)

        if etape != None :
            messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
            transaction.savepoint_commit(sid)
            return HttpResponseRedirect(reverse('module_configuration_detail_workflow', args=(workflow_id,)))
        else :
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue lors de la création de l'utilisateur")
            return HttpResponseRedirect(reverse('module_configuration_add_workflow'))

    except Exception as e:
        #print("ERREUR !")
        #print(e)
        transaction.savepoint_rollback(sid)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_add_workflow'))

def get_details_etape(request, ref):
    permission_number = 566
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response



    try:
        ref = int(ref)
        workflow = dao_wkf_workflow.toGetWorkflow(ref)
        modules = dao_module.toListModulesInstalles()

        context = {
            'title' : "Workflow : " + workflow.type_document,
            'model' : workflow,
            'utilisateur': utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'modules' : modules,
            'sous_modules': sous_modules,
            'roles_modules' : [],
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 1
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/workflow/workflow/item.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_utilisateurs'))

# TRANSITION CONTROLLER
def get_creer_transition(request, ref):
    permission_number = 567
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    if response != None:
        return response

    etapes = dao_wkf_etape.toListEtapeOfWorkflows(ref)
    conditions = dao_wkf_condition.toListConditions()

    context = {
        'title' : 'Nouvelle transition d\'un workflow',
        "utilisateur" : utilisateur,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        "modules" : modules,
        'sous_modules': sous_modules,
        "unite_fonctionnelles": [],
        "roles": dao_groupe_permission.toListGroupePermissions(),
        "module" : ErpModule.MODULE_CONFIGURATION,
        "etapes" : etapes,
        "conditions" : conditions,
        "workflow_id" : ref,
        'menu' : 1
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/workflow/transition/add.html")
    return HttpResponse(template.render(context, request))

@transaction.atomic
def post_creer_transition(request):
    sid = transaction.savepoint()
    try:
        auteur =  identite.utilisateur(request)
        et_source_id = request.POST["et_source_id"]
        et_destination_id = request.POST["et_destination_id"]
        role_id = request.POST["role_id"]
        condition_id = request.POST["condition_id"]
        url = request.POST["url"]
        workflow_id = request.POST["workflow_id"]
        num_ordre = request.POST["num_ordre"]
        if num_ordre == "":
            num_ordre = 0
        traitement = request.POST["traitement"]
        unite_fonctionnelle_id = request.POST["unite_fonctionnelle_id"]


        transition = dao_wkf_transition.toCreateTransition(et_source_id, et_destination_id,role_id,condition_id, url, num_ordre, traitement, unite_fonctionnelle_id)
        transition = dao_wkf_transition.toSaveTransition(auteur, transition)


        if transition != None :
            messages.add_message(request, messages.SUCCESS, "L'operation effectuée avec succès!")
            transaction.savepoint_commit(sid)
            return HttpResponseRedirect(reverse('module_configuration_detail_workflow', args=(workflow_id,)))
        else :
            transaction.savepoint_rollback(sid)
            messages.add_message(request, messages.ERROR, "Une erreur est survenue lors de la création de l'utilisateur")
            return HttpResponseRedirect(reverse('module_configuration_add_workflow'))

    except Exception as e:
        #print("ERREUR !")
        #print(e)
        transaction.savepoint_rollback(sid)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_add_workflow'))

def get_details_transition(request, ref):
    permission_number = 566
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    if response != None:
        return response

    try:
        ref = int(ref)
        workflow = dao_wkf_workflow.toGetWorkflow(ref)
        modules = dao_module.toListModulesInstalles()

        context = {
            'title' : "Workflow : " + workflow.type_document,
            'model' : workflow,
            'utilisateur': utilisateur,
            'actions':[],
            'sous_modules': sous_modules,
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'modules' : modules,
            'roles_modules' : [],
            "module" : ErpModule.MODULE_CONFIGURATION,
            'menu' : 1
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/workflow/workflow/item.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse('module_configuration_list_utilisateurs'))



# REGLE CONTROLLERS
def get_lister_regle(request):
    permission_number = 571
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)


    if response != None:
        return response

    model = dao_regle.toListRegles()

    context = {
        'title' : 'Liste des règles',
        'model' : model,
        "utilisateur" : utilisateur,
        'sous_modules': sous_modules,
        'actions':[],
        'organisation': dao_organisation.toGetMainOrganisation(),
        "modules" :modules,
        "module" : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/regle/list.html")
    return HttpResponse(template.render(context, request))

def get_details_regle(request, ref):
    permission_number = 571
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response
    try:
        ref = int(ref)
        model = dao_regle.toGetRegle(ref)
        sous_modules_list = []
        actions = []

        modules_du_role = dao_module.toListModulesByPermission(model.groupe_permission)
        modules_du_regle, sous_modules_list = dao_regle.toListModulesOfRegle(ref)

        for item in sous_modules_list:
            #actions.extend(dao_droit.toListDroitAutroses(item.id,model.designation))
            actions.extend(model.permissions.all())
        actions = list(set(actions))


        model = dao_regle.toGetRegle(ref)
        print("expression {}".format(model.expression))
        context = {
            'title' : model.designation,
            'model' : model,
            'modules' : modules,
            'modules_du_role' : modules_du_role,
            'modules_du_regle' : modules_du_regle,
            'sous_modules_list': sous_modules_list,
            'sous_modules': sous_modules,
            'actions_utilisateur': actions,
            "module" : ErpModule.MODULE_CONFIGURATION,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'utilisateur': utilisateur,
            'menu' : 2
        }
        template = loader.get_template("ErpProject/ModuleConfiguration/regle/item.html")
        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print("ERREUR")
        #print(e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_list_regle"))

def get_creer_regle(request):
    permission_number = 572
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    sous_modules_list = []
    actions_utilisateur = []
    groupe_permissions = dao_groupe_permission.toListGroupePermissions()

    for item in modules:
        sous_modules_list.extend(dao_sous_module.toListSousModulesOf(item.id))

    for item in sous_modules_list:
        #actions_utilisateur.extend(dao_droit.toListDroitNonAutroses(item.id,'role_nouveau'))
        actions_utilisateur.extend(dao_permission.toListPermissionsOfSousModule(item.id))

    type_condition_tests = dao_constante.toListTypeConditionTest()
    type_operation_tests = dao_constante.toListTypeOperationTest()

    context = {
        'title' : 'Nouvelle règle',
        'utilisateur' : utilisateur,
        'type_condition_tests': type_condition_tests,
        'type_operation_tests': type_operation_tests,
        'actions':[],
        'organisation': dao_organisation.toGetMainOrganisation(),
        'modules' : modules,
        'groupe_permissions': groupe_permissions,
        'sous_modules_list' : sous_modules_list,
        'sous_modules': sous_modules,
        'actions_utilisateur' : actions_utilisateur,
        "module" : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2
    }
    template = loader.get_template("ErpProject/ModuleConfiguration/regle/add.html")
    return HttpResponse(template.render(context, request))


@transaction.atomic
def post_creer_regle(request):
    sid = transaction.savepoint()
    try:
        #print(request.POST)
        erreur_survenue = False
        designation = request.POST["designation"]
        filtre = request.POST["filtre"]
        groupe_permission_id = request.POST["groupe_permission_id"]
        auteur = identite.utilisateur(request)

        list_module_id = request.POST.getlist('module_id', None)
        list_sous_module_id = request.POST.getlist('sous_module_id', None)
        list_permissions_id = request.POST.getlist('action_id', None)



        regle = dao_regle.toCreateRegle(designation, filtre, groupe_permission_id, list_permissions_id)
        regle = dao_regle.toSaveRegle(auteur, regle)
        for i in range(0, len(list_permissions_id)):
            permission_id = int(list_permissions_id[i])
            regle.permissions.add(permission_id)

        list_sequence = request.POST.getlist("sequence", None)
        list_condition_select = request.POST.getlist("condition_select", None)
        list_operation_select = request.POST.getlist("operation_select", None)
        list_code_input = request.POST.getlist("code_input", None)
        list_valeur_input = request.POST.getlist("valeur_input", None)

        for i in range(0, len(list_sequence)) :
            #print("Initialisation")
            sequence = int(list_sequence[i])
            operation_select = int(list_operation_select[i])
            condition_select = int(list_condition_select[i])
            code_input = list_code_input[i]
            valeur_input = list_valeur_input[i]

            #On s'assure que la première ligne sera toujours la condition du début pour la cohenrence
            if len(list_sequence) == 1: condition_select = 1
            elif int(list_condition_select[0]) not in (1, 4): condition_select = 1


            ligne = Model_LigneRegle(sequence = sequence, regle_id = regle.id, type_operation = operation_select, type_condition = condition_select, code = code_input, valeur = valeur_input)
            ligne.save()
            print("Ligne Test {} cree".format(ligne.sequence))

        return HttpResponseRedirect(reverse("module_configuration_details_regle", args=(regle.id,)))
    except Exception as e:
        print("ERREUR")
        print(e)
        transaction.savepoint_rollback(sid)
        messages.add_message(request, messages.ERROR, e)
        messages.error(request,e)
        return HttpResponseRedirect(reverse("module_configuration_add_regle"))

def get_lister_permission(request):

    permission_number = 575
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    model = dao_permission.toListPermission()
    context ={
                'title' : 'Liste des permissions',
                'model' : model,
                'sous_modules': sous_modules,
                'actions':[],
                'organisation': dao_organisation.toGetMainOrganisation(),
                'utilisateur': utilisateur,
                'modules' : modules,
                'module' : ErpModule.MODULE_CONFIGURATION,
                'degrade': 'module_configuration',
                'menu' : 1}
    template = loader.get_template('ErpProject/ModuleConfiguration/permission/list.html')
    return HttpResponse(template.render(context, request))

def get_creer_permission(request):
    permission_number = 576
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    context ={
            'title' : 'Nouvelle permission',
            'utilisateur': utilisateur,
            'modules' : modules,
            'sousmodule' : dao_sous_module.toListSousModule(),
            'sous_modules':sous_modules,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'module' : ErpModule.MODULE_CONFIGURATION,
            'menu' : 2}
    template = loader.get_template('ErpProject/ModuleConfiguration/permission/add.html')
    return HttpResponse(template.render(context, request))

def post_creer_permission(request):

    try:
        sous_module_id = request.POST['sous_module_id']
        designation = request.POST['designation']
        numero = request.POST['numero']
        auteur = identite.utilisateur(request)

        permission=dao_permission.toCreatePermission(sous_module_id,designation,numero)
        permission=dao_permission.toSavePermission(auteur, permission)
        return HttpResponseRedirect(reverse('module_Configuration_detail_permission', args=(permission.id,)))
    except Exception as e:
        #print('Erreur lors de l enregistrement')
        auteur = identite.utilisateur(request)
        monLog.error('{} :: {} :: \nERREUR LORS DU POST CREER PERMISSION \n {}'.format(auteur.nom_complet, module,e))
        #print(e)
        return HttpResponseRedirect(reverse('module_Configuration_add_permission'))


def get_details_permission(request,ref):
    permission_number = 575
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response
    try:
        ref=int(ref)
        permission=dao_permission.toGetPermission(ref)
        template = loader.get_template('ErpProject/ModuleConfiguration/permission/item.html')
        context ={
                'title' : 'Details d\'une permission',
                'permission' : permission,
                'utilisateur': utilisateur,
                'sous_modules': sous_modules,
                'modules' : modules,
                'actions':[],
                'organisation': dao_organisation.toGetMainOrganisation(),
                'module' : ErpModule.MODULE_CONFIGURATION,
                'menu' : 4
                }

        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print('Erreut Get Detail')
        auteur = identite.utilisateur(request)
        monLog.error('{} :: {} :: \nERREUR LORS DU GET DETAILS PERMISSION \n {}'.format(auteur.nom_complet, module,e))
        #print(e)
        return HttpResponseRedirect(reverse('module_Configuration_list_permission'))
def get_modifier_permission(request,ref):
    permission_number = 577
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    ref = int(ref)
    model = dao_permission.toGetPermission(ref)
    context ={
        'title' : 'Modifier Permission',
        'model':model,
        'sous_modules':sous_modules,
        'sousmodule' : dao_sous_module.toListSousModule(),
        'utilisateur': utilisateur,
        'modules' : modules,
        'actions':[],
        'organisation': dao_organisation.toGetMainOrganisation(),
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2}
    template = loader.get_template('ErpProject/ModuleConfiguration/permission/update.html')
    return HttpResponse(template.render(context, request))

def post_modifier_permission(request):

    id = int(request.POST['ref'])
    try:
        sous_module_id = request.POST['sous_module_id']
        designation = request.POST['designation']
        numero = request.POST['numero']
        auteur = identite.utilisateur(request)

        permission=dao_permission.toCreatePermission(sous_module_id,designation,numero)
        permission=dao_permission.toUpdatePermission(id, permission)
        return HttpResponseRedirect(reverse('module_Configuration_list_permission'))
    except Exception as e:
        #print('Erreur lors de l enregistrement')
        auteur = identite.utilisateur(request)
        monLog.error('{} :: {} :: \nERREUR LORS DU POST MODIFIER PERMISSION \n {}'.format(auteur.nom_complet, module,e))
        #print(e)
        return HttpResponseRedirect(reverse('module_Configuration_add_permission'))

def get_lister_actionutilisateur(request):
    permission_number = 579
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response
    model = dao_actionutilisateur.toListActionutilisateur()
    context ={
        'title' : 'Liste des actions',
        'model' : model,
        'sous_modules': sous_modules,
        'utilisateur': utilisateur,
        'actions':[],
        'organisation': dao_organisation.toGetMainOrganisation(),
        'modules' : modules,
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 1}
    template = loader.get_template('ErpProject/ModuleConfiguration/actionutilisateur/list.html')
    return HttpResponse(template.render(context, request))

def get_creer_actionutilisateur(request):
    permission_number = 580
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    context ={
        'title' : 'Nouvelle action',
        'utilisateur': utilisateur,
        'modules' : modules,
        "sous_modules": sous_modules,
        "permissions": dao_permission.toListPermissions(),
        'actions':[],
        'organisation': dao_organisation.toGetMainOrganisation(),
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2}
    template = loader.get_template('ErpProject/ModuleConfiguration/actionutilisateur/add.html')
    return HttpResponse(template.render(context, request))

def post_creer_actionutilisateur(request):

    try:
        nom_action = request.POST['nom_action']
        ref_action = request.POST['ref_action']
        description = request.POST['description']
        permission_id = request.POST['permission_id']
        auteur = identite.utilisateur(request)

        actionutilisateur=dao_actionutilisateur.toCreateActionutilisateur(nom_action,ref_action,description,permission_id)
        actionutilisateur=dao_actionutilisateur.toSaveActionutilisateur(auteur, actionutilisateur)
        return HttpResponseRedirect(reverse('module_Configuration_detail_actionutilisateur', args=(actionutilisateur.id,)))
    except Exception as e:
        #print('Erreur lors de l enregistrement')
        auteur = identite.utilisateur(request)
        monLog.error('{} :: {} :: \nERREUR LORS DU POST CREER ACTIONUTILISATEUR \n {}'.format(auteur.nom_complet, module,e))
        #print(e)
        return HttpResponseRedirect(reverse('module_Configuration_add_actionutilisateur'))


def get_details_actionutilisateur(request,ref):
    permission_number = 579
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    try:
        ref=int(ref)
        actionutilisateur=dao_actionutilisateur.toGetActionutilisateur(ref)
        template = loader.get_template('ErpProject/ModuleConfiguration/actionutilisateur/item.html')
        context ={
            'title' : 'Détails d\'une action',
            'actionutilisateur' : actionutilisateur,
            'utilisateur': utilisateur,
            'sous_modules': sous_modules,
            'modules' : modules,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'module' : ErpModule.MODULE_CONFIGURATION,
            'menu' : 4}

        return HttpResponse(template.render(context, request))
    except Exception as e:
        #print('Erreut Get Detail')
        auteur = identite.utilisateur(request)
        monLog.error('{} :: {} :: \nERREUR LORS DU GET DETAILS ACTIONUTILISATEUR \n {}'.format(auteur.nom_complet, module,e))
        #print(e)
        return HttpResponseRedirect(reverse('module_Configuration_list_actionutilisateur'))
def get_modifier_actionutilisateur(request,ref):
    permission_number = 581
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)

    if response != None:
        return response

    ref = int(ref)
    model = dao_actionutilisateur.toGetActionutilisateur(ref)
    context ={
        'title' : 'Modifier une action',
        'model':model,
        'utilisateur': utilisateur,
        'modules' : modules,
        'sous_modules':sous_modules,
        "permissions": dao_permission.toListPermissions(),
        'module' : ErpModule.MODULE_CONFIGURATION,
        'actions':[],
        'organisation': dao_organisation.toGetMainOrganisation(),
        'menu' : 2}
    template = loader.get_template('ErpProject/ModuleConfiguration/actionutilisateur/update.html')
    return HttpResponse(template.render(context, request))

def post_modifier_actionutilisateur(request):

    id = int(request.POST['ref'])
    try:
        nom_action = request.POST['nom_action']
        ref_action = request.POST['ref_action']
        description = request.POST['description']
        permission_id = request.POST['permission_id']
        auteur = identite.utilisateur(request)

        actionutilisateur=dao_actionutilisateur.toCreateActionutilisateur(nom_action,ref_action,description,permission_id)
        actionutilisateur=dao_actionutilisateur.toUpdateActionutilisateur(id, actionutilisateur)
        return HttpResponseRedirect(reverse('module_Configuration_list_actionutilisateur'))
    except Exception as e:
        #print('Erreur lors de l enregistrement')
        auteur = identite.utilisateur(request)
        monLog.error('{} :: {} :: \nERREUR LORS DU POST MODIFIER ACTIONUTILISATEUR \n {}'.format(auteur.nom_complet, module,e))
        #print(e)
        return HttpResponseRedirect(reverse('module_Configuration_add_actionutilisateur'))

def get_lister_sousmodule(request):
	permission_number = 587
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
	if response != None:
		return response

	model = dao_sousmodule.toListSousmodule()
	context = {
        'title' : 'Liste des menus',
        'model' : model,
        'utilisateur': utilisateur,
        'modules' : modules,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        'sous_modules': sous_modules,
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 1}
	template = loader.get_template('ErpProject/ModuleConfiguration/sousmodule/list.html')
	return HttpResponse(template.render(context, request))

def get_creer_sousmodule(request):
	permission_number = 588
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
	if response != None:
		return response

	context ={
        'title' : 'Nouveau menu',
        'utilisateur': utilisateur,
        'modules' : modules,
        'lesmodules': dao_module.toListModules(),
        'groupe_menus': dao_groupe_menu.toListGroupeMenus(),
        'contents': ContentType.objects.all(),
        'sous_modules': sous_modules,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2}
	template = loader.get_template('ErpProject/ModuleConfiguration/sousmodule/add.html')
	return HttpResponse(template.render(context, request))

def post_creer_sousmodule(request):

	try:
		module_id = request.POST['module_id']
		nom_sous_module = request.POST['nom_sous_module']
		description = request.POST['description']
		groupe = request.POST['groupe']
		icon_menu = request.POST['icon_menu']
		url_vers = request.POST['url_vers']
		numero_ordre = request.POST['numero_ordre']
		model_principal_id = request.POST['model_principal_id']
		groupe_menu_id = request.POST['groupe_menu_id']
		auteur = identite.utilisateur(request)

		est_model = False
		if "est_model" in request.POST : est_model = True

		est_dashboard = False
		if "est_dashboard" in request.POST : est_dashboard = True

		est_actif = False
		if "est_actif" in request.POST : est_actif = True

		sousmodule=dao_sousmodule.toCreateSousmodule(module_id,nom_sous_module,description,groupe,icon_menu,url_vers,numero_ordre,est_model,est_dashboard,est_actif,model_principal_id,groupe_menu_id)
		sousmodule=dao_sousmodule.toSaveSousmodule(auteur, sousmodule)
		return HttpResponseRedirect(reverse('module_Configuration_detail_sousmodule', args=(sousmodule.id,)))
	except Exception as e:
		print('Erreur lors de l enregistrement')
		auteur = identite.utilisateur(request)
		monLog.error('{} :: {} :: \nERREUR LORS DU POST CREER SOUSMODULE \n {}'.format(auteur.nom_complet, module,e))
		print(e)
		return HttpResponseRedirect(reverse('module_Configuration_add_sousmodule'))


def get_details_sousmodule(request,ref):
	permission_number = 587
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
	if response != None:
		return response
	try:
		ref=int(ref)
		sousmodule=dao_sousmodule.toGetSousmodule(ref)
		template = loader.get_template('ErpProject/ModuleConfiguration/sousmodule/item.html')
		context ={
            'title' : 'Details d\'un menu',
            'sousmodule' : sousmodule,
            'sous_modules': sous_modules,
            'utilisateur': utilisateur,
            'actions':[],
		    'organisation': dao_organisation.toGetMainOrganisation(),
            'modules' : modules,
            'module' : ErpModule.MODULE_CONFIGURATION,
            'menu' : 4}

		return HttpResponse(template.render(context, request))
	except Exception as e:
		#print('Erreut Get Detail')
		auteur = identite.utilisateur(request)
		monLog.error('{} :: {} :: \nERREUR LORS DU GET DETAILS SOUSMODULE \n {}'.format(auteur.nom_complet, module,e))
		#print(e)
		return HttpResponseRedirect(reverse('module_Configuration_list_sousmodule'))
def get_modifier_sousmodule(request,ref):
	permission_number = 589
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
	if response != None:
		return response

	ref = int(ref)
	model = dao_sousmodule.toGetSousmodule(ref)
	context ={
        'title' : 'Modifier un menu',
        'model':model,
        'utilisateur': utilisateur,
        'modules' : modules,
        'lesmodules': dao_module.toListModules(),
        'groupe_menus': dao_groupe_menu.toListGroupeMenus(),
        'contents': ContentType.objects.all(),
        'sous_modules': sous_modules,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2}
	template = loader.get_template('ErpProject/ModuleConfiguration/sousmodule/update.html')
	return HttpResponse(template.render(context, request))

def post_modifier_sousmodule(request):

	id = int(request.POST['ref'])
	try:
		module_id = request.POST['module_id']
		nom_sous_module = request.POST['nom_sous_module']
		description = request.POST['description']
		groupe = request.POST['groupe']
		icon_menu = request.POST['icon_menu']
		url_vers = request.POST['url_vers']
		numero_ordre = request.POST['numero_ordre']
		model_principal_id = request.POST['model_principal_id']
		groupe_menu_id = request.POST['groupe_menu_id']
		auteur = identite.utilisateur(request)

		est_model = False
		if "est_model" in request.POST : est_model = True

		est_dashboard = False
		if "est_dashboard" in request.POST : est_dashboard = True

		est_actif = False
		if "est_actif" in request.POST : est_actif = True

		sousmodule=dao_sousmodule.toCreateSousmodule(module_id,nom_sous_module,description,groupe,icon_menu,url_vers,numero_ordre,est_model,est_dashboard,est_actif,model_principal_id,groupe_menu_id)
		sousmodule=dao_sousmodule.toUpdateSousmodule(id, sousmodule)
		return HttpResponseRedirect(reverse('module_Configuration_detail_sousmodule', args=(sousmodule.id,)))
	except Exception as e:
		#print('Erreur lors de l enregistrement')
		auteur = identite.utilisateur(request)
		monLog.error('{} :: {} :: \nERREUR LORS DU POST MODIFIER SOUSMODULE \n {}'.format(auteur.nom_complet, module,e))
		#print(e)
		return HttpResponseRedirect(reverse('module_Configuration_list_sousmodule'))


def get_lister_groupemenu(request):
	permission_number = 583
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
	if response != None:
		return response

	model = dao_groupemenu.toListGroupemenu()
	context ={
        'title' : 'Liste des Groupe Menu',
        'model' : model,
        'utilisateur': utilisateur,
        'sous_modules': sous_modules,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        'modules' : modules,
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 1}
	template = loader.get_template('ErpProject/ModuleConfiguration/groupemenu/list.html')
	return HttpResponse(template.render(context, request))

def get_creer_groupemenu(request):
    permission_number = 584
    modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
    if response != None:
        return response

    try:
        module_id = int(request.GET.get("module_id",0))
    except Exception as e:
        module_id = 0

    context = {
        'title' : 'Nouveau Groupe Menu',
        'utilisateur': utilisateur,
        'modules' : modules,
        'module_id': module_id,
        'lesmodules': dao_module.toListModules(),
        'sous_modules': sous_modules,
        'actions':[],
        'organisation': dao_organisation.toGetMainOrganisation(),
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2}
    template = loader.get_template('ErpProject/ModuleConfiguration/groupemenu/add.html')
    return HttpResponse(template.render(context, request))

def post_creer_groupemenu(request):
    try:
        designation = request.POST['designation']
        icon_menu = request.POST['icon_menu']
        description = request.POST['description']
        module_id = request.POST['module_id']
        numero_ordre = request.POST['numero_ordre']
        module_page = int(request.POST['module_page'])
        auteur = identite.utilisateur(request)

        groupemenu=dao_groupemenu.toCreateGroupemenu(designation,icon_menu,description,module_id,numero_ordre)
        groupemenu=dao_groupemenu.toSaveGroupemenu(auteur, groupemenu)
        if module_page == 0: return HttpResponseRedirect(reverse('module_Configuration_detail_groupemenu', args=(groupemenu.id,)))
        else : return HttpResponseRedirect(reverse('module_configuration_details_module', args=(module_id,)))
    except Exception as e:
        print('Erreur lors de l enregistrement')
        auteur = identite.utilisateur(request)
        monLog.error('{} :: {} :: \nERREUR LORS DU POST CREER GROUPEMENU \n {}'.format(auteur.nom_complet, module,e))
        print(e)
        return HttpResponseRedirect(reverse('module_Configuration_add_groupemenu'))


def get_details_groupemenu(request,ref):
	permission_number = 583
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
	if response != None:
		return response
	try:
		ref=int(ref)
		groupemenu=dao_groupemenu.toGetGroupemenu(ref)
		template = loader.get_template('ErpProject/ModuleConfiguration/groupemenu/item.html')
		context ={
            'title' : 'Details d\'un Groupe Menu',
            'groupemenu' : groupemenu,
            'utilisateur': utilisateur,
            'modules' : modules,
            'sous_modules': sous_modules,
            'actions':[],
            'organisation': dao_organisation.toGetMainOrganisation(),
            'module' : ErpModule.MODULE_CONFIGURATION,
            'menu' : 4}

		return HttpResponse(template.render(context, request))
	except Exception as e:
		#print('Erreut Get Detail')
		auteur = identite.utilisateur(request)
		monLog.error('{} :: {} :: \nERREUR LORS DU GET DETAILS GROUPEMENU \n {}'.format(auteur.nom_complet, module,e))
		#print(e)
		return HttpResponseRedirect(reverse('module_Configuration_list_groupemenu'))
def get_modifier_groupemenu(request,ref):
	permission_number = 583
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
	if response != None:
		return response

	ref = int(ref)
	model = dao_groupemenu.toGetGroupemenu(ref)
	context ={
        'title' : 'Modifier un Groupe Menu',
        'model':model, 
        'utilisateur': utilisateur,
        'lesmodules': dao_module.toListModules(),
        'sous_modules': sous_modules,
        'actions':[],
        'organisation': dao_organisation.toGetMainOrganisation(),
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2
        }
	template = loader.get_template('ErpProject/ModuleConfiguration/groupemenu/update.html')
	return HttpResponse(template.render(context, request))

def post_modifier_groupemenu(request):

	id = int(request.POST['ref'])
	try:
		designation = request.POST['designation']
		icon_menu = request.POST['icon_menu']
		description = request.POST['description']
		module_id = request.POST['module_id']
		numero_ordre = request.POST['numero_ordre']
		auteur = identite.utilisateur(request)

		groupemenu=dao_groupemenu.toCreateGroupemenu(designation,icon_menu,description,module_id,numero_ordre)
		groupemenu=dao_groupemenu.toUpdateGroupemenu(id, groupemenu)
		return HttpResponseRedirect(reverse('module_Configuration_detail_groupemenu', args=(groupemenu.id,)))
	except Exception as e:
		#print('Erreur lors de l enregistrement')
		auteur = identite.utilisateur(request)
		monLog.error('{} :: {} :: \nERREUR LORS DU POST MODIFIER GROUPEMENU \n {}'.format(auteur.nom_complet, module,e))
		#print(e)
		return HttpResponseRedirect(reverse('module_Configuration_add_groupemenu'))




def get_creer_wizard_menu(request):
	permission_number = 588
	modules, sous_modules, utilisateur, groupe_permissions, response = auth.toGetAuthentification(permission_number, request, inspect.getframeinfo(inspect.currentframe()).function)
	if response != None:
		return response

	context ={
        'title' : 'Wizard Menu',
        'utilisateur': utilisateur,
        'modules' : modules,
        'lesmodules': dao_module.toListModules(),
        'groupe_menus': dao_groupe_menu.toListGroupeMenus(),
        'contents': ContentType.objects.all(),
        'sous_modules': sous_modules,
        'actions':[],
		'organisation': dao_organisation.toGetMainOrganisation(),
        'module' : ErpModule.MODULE_CONFIGURATION,
        'menu' : 2}
	template = loader.get_template('ErpProject/ModuleConfiguration/sousmodule/wizard.html')
	return HttpResponse(template.render(context, request))


@transaction.atomic
def post_creer_wizard_menu(request):
	sid = transaction.savepoint()
	try:
		module_id = request.POST['module_id']
		nom_sous_module = request.POST['nom_sous_module']
		description = request.POST['description']
		groupe = ""#request.POST['groupe']
		numero_ordre = request.POST['numero_ordre']	
		icon_menu = request.POST['icon_menu']
		url_vers = request.POST['url_vers']
		model_principal_id = request.POST['model_principal_id'] if request.POST['model_principal_id'] else None
		groupe_menu_id = request.POST['groupe_menu_id'] if request.POST['groupe_menu_id'] else None
		menu_id = request.POST['menu_id']
		if not groupe_menu_id:
			if menu_id and menu_id != 0:
				menu = dao_sous_module.toGetSousModule(menu_id)
				groupe_menu_id = menu.groupe_menu_id
				numero_ordre = menu.numero_ordre + 1
			else:
				groupe_menu_id = None
				
        
		auteur = identite.utilisateur(request)

		est_model = False
		if "est_model" in request.POST : est_model = True

		est_dashboard = False
		if "est_dashboard" in request.POST : est_dashboard = True

		est_actif = False
		if "est_actif" in request.POST : est_actif = True

		sousmodule=dao_sousmodule.toCreateSousmodule(module_id,nom_sous_module,description,groupe,icon_menu,url_vers,numero_ordre,est_model,est_dashboard,est_actif,model_principal_id,groupe_menu_id)
		sousmodule=dao_sousmodule.toSaveSousmodule(auteur, sousmodule)

		list_permission = request.POST.getlist('permission', None)
		list_action = request.POST.getlist('action', None)

		for i in range(0, len(list_permission)) :
			permission = dao_permission.toCreatePermission(sousmodule.id, list_permission[i],dao_permission.toGetLatestNumeroOrdre() + 1)
			permission = dao_permission.toSavePermission(auteur, permission)
			action_string = list_action[i]
			list_action_string = action_string.split(",")
			for uneaction in list_action_string:
				action = dao_actionutilisateur.toCreateActionutilisateur(uneaction,"", "", permission.id)
				action = dao_actionutilisateur.toSaveActionutilisateur(auteur, action)
        
		transaction.savepoint_commit(sid)	
		return HttpResponseRedirect(reverse('module_Configuration_detail_sousmodule', args=(sousmodule.id,)))
	except Exception as e:
		#print('Erreur lors de l enregistrement')
		transaction.savepoint_rollback(sid)
		auteur = identite.utilisateur(request)
		monLog.error('{} :: {} :: \nERREUR LORS DU POST CREER SOUSMODULE \n {}'.format(auteur.nom_complet, module,e))
		#print(e)
		return HttpResponseRedirect(reverse('module_Configuration_add_wizard_menu'))


def get_json_sous_modules(request):
	try:
		data = []
		ident = int(request.GET["ref"])
		sous_modules = dao_sous_module.toListSousModulesOf(ident)
		for sous_module in sous_modules:
			groupe_menu_designation = sous_module.groupe_menu.designation if sous_module.groupe_menu else None
			item = {
                "sous_module_id": sous_module.id,
                "nom_sous_module": sous_module.nom_sous_module,
                "est_dashboard": sous_module.est_dashboard,
                "numero_ordre": sous_module.numero_ordre,
                "est_actif": sous_module.est_actif,
                "groupe_menu_id": sous_module.groupe_menu_id,
                "groupe_menu_designation": groupe_menu_designation
                }
			data.append(item)
		#print(data)

		return JsonResponse(data, safe=False)
	except Exception as e:
		#print("error", e)
		return JsonResponse([], safe=False)


# AJAX
def ajax_get_related_models(request):
    try:
        data = []        
        model_id = request.GET["model_id"]        
        model = ContentType.objects.get(pk = model_id)
        nom_modele_class = model.model_class().__name__
        
        models = ContentType.objects.all()

        for item in models:
            model_class = item.model_class()
            if model_class != None:
                for field in model_class._meta.get_fields(include_parents=False, include_hidden=False):
                    if field.related_model != None and field.__class__.__name__ not in  ("ManyToManyField", "ManyToOneRel", "ManyToManyRel") and field.related_model.__name__ == nom_modele_class:
                        element = {
                            'content_id' : item.id,
                            'model_name' : item.model,
                            'model_class' : item.model_class().__name__,
                            'field_name' : field.name,
                            'field_type' : field.__class__.__name__,
                        }
                        data.append(element)

        #print("DATA %s" % data)
        return JsonResponse(data, safe=False)
    except Exception as e:
        return auth.toReturnApiFailed(request, e)
    
def ajax_get_urls_names(request):
    try:
        print("ajax_get_urls_names")
        data = {}        
        model_id = request.GET["model_id"] 
        module_id = request.GET["module_id"]       
        content_type = ContentType.objects.get(pk = model_id)
        model_class = content_type.model_class()
        nom_modele_verbose = model_class._meta.verbose_name
        nom_modele_verbose_plural = model_class._meta.verbose_name_plural
        
        module = dao_module.toGetModule(module_id)
        
        #Standardisation denomination modele
        nom_modele = content_type.model.replace("model_","").capitalize()
        nom_pattern = 'module_{0}'.format(unidecode.unidecode(module.nom_module.lower().replace(" ","_")))

        url_name_create = "{1}_add_{0}".format(nom_modele.lower(), nom_pattern)
        url_name_list = "{1}_list_{0}".format(nom_modele.lower(), nom_pattern)
        url_name_detail = "{1}_detail_{0}".format(nom_modele.lower(), nom_pattern)
        url_name_update = "{1}_update_{0}".format(nom_modele.lower(), nom_pattern)
        url_name_reporting = "{1}_get_generer_{0}".format(nom_modele.lower(), nom_pattern)
        
        data = {
            'lister' : f"{url_name_list},{url_name_detail}",
            'modifier' : url_name_update,
            'creer' : url_name_create,
            'rapport' : url_name_reporting,
            'nom_modele' : nom_modele.upper(),
            'nom_plural' : nom_modele_verbose_plural,
            'url_list' : url_name_list,
        }
        #print("DATA %s" % data)
        return JsonResponse(data, safe=False)
    except Exception as e:
        return auth.toReturnApiFailed(request, e)

# SOCIETE CONTROLLERS
from ModuleConfiguration.dao.dao_societe import dao_societe

def get_lister_societe(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_societe.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_societe.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des sociétés",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/societe/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_societe(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - Société",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_Societe(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'devises' : Model_Devise.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'type_periodes' : Model_Type_periode.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'communes' : Model_Commune.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'adresses' : Model_Adresse.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'contacts' : Model_Contact.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/societe/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_societe'))

@transaction.atomic
def post_creer_societe(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_societe'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		code = str(request.POST['code'])
		if code in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')

		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom de la société\' est obligatoire, Veuillez le renseigner SVP!')

		picture_icon = request.FILES['picture_icon'] if 'picture_icon' in request.FILES else None

		type = str(request.POST['type'])

		societe_id = makeIntId(request.POST['societe_id'])

		devise_id = makeIntId(request.POST['devise_id'])

		type_periode_id = makeIntId(request.POST['type_periode_id'])

		adress_email = str(request.POST['adress_email'])

		siteweb = str(request.POST['siteweb'])

		pays_id = makeIntId(request.POST['pays_id'])

		pays_adress = str(request.POST['pays_adress'])

		province_id = makeIntId(request.POST['province_id'])

		province_adress = str(request.POST['province_adress'])

		ville_id = makeIntId(request.POST['ville_id'])

		ville_adress = str(request.POST['ville_adress'])

		commune_id = makeIntId(request.POST['commune_id'])

		adresse_line1 = str(request.POST['adresse_line1'])

		adresse_line2 = str(request.POST['adresse_line2'])

		telephone_1 = str(request.POST['telephone_1'])

		telephone_2 = str(request.POST['telephone_2'])

		nbr_periode_gl = makeInt(request.POST['nbr_periode_gl'])

		nbr_periode_ar = makeInt(request.POST['nbr_periode_ar'])

		nbr_periode_ap = makeInt(request.POST['nbr_periode_ap'])

		nbr_periode_cm = makeInt(request.POST['nbr_periode_cm'])

		nbr_periode_fa = makeInt(request.POST['nbr_periode_fa'])

		nbr_periode_bgt = makeInt(request.POST['nbr_periode_bgt'])

		nbr_periode_py = makeInt(request.POST['nbr_periode_py'])

		period_begin_date = str(request.POST['period_begin_date'])
		is_formated, period_begin_date = checkDateTimeFormat(period_begin_date)
		if is_formated == False: return auth.toReturnApiFailed(request, 'Mauvais format Date et temps saisi', '', msg = 'La valeur saisi sur le champ \'Début exercice en cours\' ne correspond pas au format jj/mm/aaaa HH:MM:SS')

		period_end_date = str(request.POST['period_end_date'])
		is_formated, period_end_date = checkDateTimeFormat(period_end_date)
		if is_formated == False: return auth.toReturnApiFailed(request, 'Mauvais format Date et temps saisi', '', msg = 'La valeur saisi sur le champ \'Fin Exercice en cours\' ne correspond pas au format jj/mm/aaaa HH:MM:SS')

		description = str(request.POST['description'])

		autres_adresses = request.POST.getlist('autres_adresses', None)

		contacts = request.POST.getlist('contacts', None)

		auteur = identite.utilisateur(request)

		societe = dao_societe.toCreate(code = code, name = name, picture_icon = picture_icon, type = type, societe_id = societe_id, devise_id = devise_id, type_periode_id = type_periode_id, adress_email = adress_email, siteweb = siteweb, pays_id = pays_id, pays_adress = pays_adress, province_id = province_id, province_adress = province_adress, ville_id = ville_id, ville_adress = ville_adress, commune_id = commune_id, adresse_line1 = adresse_line1, adresse_line2 = adresse_line2, telephone_1 = telephone_1, telephone_2 = telephone_2, nbr_periode_gl = nbr_periode_gl, nbr_periode_ar = nbr_periode_ar, nbr_periode_ap = nbr_periode_ap, nbr_periode_cm = nbr_periode_cm, nbr_periode_fa = nbr_periode_fa, nbr_periode_bgt = nbr_periode_bgt, nbr_periode_py = nbr_periode_py, period_begin_date = period_begin_date, period_end_date = period_end_date, description = description, autres_adresses = autres_adresses, contacts = contacts)
		saved, societe, message = dao_societe.toSave(auteur, societe, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_societe.toListById(societe.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, societe)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : societe.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_societe(request,ref):
	try:
		same_perm_with = 'module_configuration_list_societe'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		societe = dao_societe.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(societe.id),'obj': str(societe)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_societe', args=(societe.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_societe(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		societe = auth.toGetWithRules(dao_societe.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if societe == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, societe) 

		context = {
			'title' : "Détails - Société : {}".format(societe),
			'model' : societe,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/societe/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_societe'))

def get_modifier_societe(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_societe.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - Société",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'devises' : Model_Devise.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'type_periodes' : Model_Type_periode.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'communes' : Model_Commune.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'adresses' : Model_Adresse.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'contacts' : Model_Contact.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/societe/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_societe(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_societe'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		code = str(request.POST['code'])
		if code in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')

		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom de la société\' est obligatoire, Veuillez le renseigner SVP!')

		picture_icon = request.FILES['picture_icon'] if 'picture_icon' in request.FILES else None

		type = str(request.POST['type'])

		societe_id = makeIntId(request.POST['societe_id'])

		devise_id = makeIntId(request.POST['devise_id'])

		type_periode_id = makeIntId(request.POST['type_periode_id'])

		adress_email = str(request.POST['adress_email'])

		siteweb = str(request.POST['siteweb'])

		pays_id = makeIntId(request.POST['pays_id'])

		pays_adress = str(request.POST['pays_adress'])

		province_id = makeIntId(request.POST['province_id'])

		province_adress = str(request.POST['province_adress'])

		ville_id = makeIntId(request.POST['ville_id'])

		ville_adress = str(request.POST['ville_adress'])

		commune_id = makeIntId(request.POST['commune_id'])

		adresse_line1 = str(request.POST['adresse_line1'])

		adresse_line2 = str(request.POST['adresse_line2'])

		telephone_1 = str(request.POST['telephone_1'])

		telephone_2 = str(request.POST['telephone_2'])

		nbr_periode_gl = makeInt(request.POST['nbr_periode_gl'])

		nbr_periode_ar = makeInt(request.POST['nbr_periode_ar'])

		nbr_periode_ap = makeInt(request.POST['nbr_periode_ap'])

		nbr_periode_cm = makeInt(request.POST['nbr_periode_cm'])

		nbr_periode_fa = makeInt(request.POST['nbr_periode_fa'])

		nbr_periode_bgt = makeInt(request.POST['nbr_periode_bgt'])

		nbr_periode_py = makeInt(request.POST['nbr_periode_py'])

		period_begin_date = str(request.POST['period_begin_date'])
		is_formated, period_begin_date = checkDateTimeFormat(period_begin_date)
		if is_formated == False: return auth.toReturnApiFailed(request, 'Mauvais format Date et temps saisi', '', msg = 'La valeur saisi sur le champ \'Début exercice en cours\' ne correspond pas au format jj/mm/aaaa HH:MM:SS')

		period_end_date = str(request.POST['period_end_date'])
		is_formated, period_end_date = checkDateTimeFormat(period_end_date)
		if is_formated == False: return auth.toReturnApiFailed(request, 'Mauvais format Date et temps saisi', '', msg = 'La valeur saisi sur le champ \'Fin Exercice en cours\' ne correspond pas au format jj/mm/aaaa HH:MM:SS')

		description = str(request.POST['description'])

		autres_adresses = request.POST.getlist('autres_adresses', None)

		contacts = request.POST.getlist('contacts', None)
		auteur = identite.utilisateur(request)

		societe = dao_societe.toCreate(code = code, name = name, picture_icon = picture_icon, type = type, societe_id = societe_id, devise_id = devise_id, type_periode_id = type_periode_id, adress_email = adress_email, siteweb = siteweb, pays_id = pays_id, pays_adress = pays_adress, province_id = province_id, province_adress = province_adress, ville_id = ville_id, ville_adress = ville_adress, commune_id = commune_id, adresse_line1 = adresse_line1, adresse_line2 = adresse_line2, telephone_1 = telephone_1, telephone_2 = telephone_2, nbr_periode_gl = nbr_periode_gl, nbr_periode_ar = nbr_periode_ar, nbr_periode_ap = nbr_periode_ap, nbr_periode_cm = nbr_periode_cm, nbr_periode_fa = nbr_periode_fa, nbr_periode_bgt = nbr_periode_bgt, nbr_periode_py = nbr_periode_py, period_begin_date = period_begin_date, period_end_date = period_end_date, description = description, autres_adresses = autres_adresses, contacts = contacts)
		saved, societe, message = dao_societe.toUpdate(id, societe, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_societe.toListById(societe.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : societe.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_societe(request,ref):
	try:
		same_perm_with = 'module_configuration_add_societe'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_societe.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'devises' : Model_Devise.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'type_periodes' : Model_Type_periode.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'communes' : Model_Commune.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'adresses' : Model_Adresse.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'contacts' : Model_Contact.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/societe/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_societe(request,ref):
	try:
		same_perm_with = 'module_configuration_list_societe'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		societe = auth.toGetWithRules(dao_societe.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if societe == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - Société : {}".format(societe),
			'model' : societe,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_societe.html', 'print_societe.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_societe(request):
	try:
		same_perm_with = 'module_configuration_add_societe'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_societe')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des sociétés",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/societe/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_societe(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_code = makeString(request.POST['code'])
		if header_code in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_code_id: {header_code_id}')

		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom de la société\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_type = makeString(request.POST['type'])
		#print(f'header_type_id: {header_type_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		header_devise_id = makeString(request.POST['devise_id'])
		#print(f'header_devise_id: {header_devise_id}')

		header_type_periode_id = makeString(request.POST['type_periode_id'])
		#print(f'header_type_periode_id: {header_type_periode_id}')

		header_adress_email = makeString(request.POST['adress_email'])
		#print(f'header_adress_email_id: {header_adress_email_id}')

		header_siteweb = makeString(request.POST['siteweb'])
		#print(f'header_siteweb_id: {header_siteweb_id}')

		header_pays_id = makeString(request.POST['pays_id'])
		#print(f'header_pays_id: {header_pays_id}')

		header_pays_adress = makeString(request.POST['pays_adress'])
		#print(f'header_pays_adress_id: {header_pays_adress_id}')

		header_province_id = makeString(request.POST['province_id'])
		#print(f'header_province_id: {header_province_id}')

		header_province_adress = makeString(request.POST['province_adress'])
		#print(f'header_province_adress_id: {header_province_adress_id}')

		header_ville_id = makeString(request.POST['ville_id'])
		#print(f'header_ville_id: {header_ville_id}')

		header_ville_adress = makeString(request.POST['ville_adress'])
		#print(f'header_ville_adress_id: {header_ville_adress_id}')

		header_commune_id = makeString(request.POST['commune_id'])
		#print(f'header_commune_id: {header_commune_id}')

		header_adresse_line1 = makeString(request.POST['adresse_line1'])
		#print(f'header_adresse_line1_id: {header_adresse_line1_id}')

		header_adresse_line2 = makeString(request.POST['adresse_line2'])
		#print(f'header_adresse_line2_id: {header_adresse_line2_id}')

		header_telephone_1 = makeString(request.POST['telephone_1'])
		#print(f'header_telephone_1_id: {header_telephone_1_id}')

		header_telephone_2 = makeString(request.POST['telephone_2'])
		#print(f'header_telephone_2_id: {header_telephone_2_id}')

		header_nbr_periode_gl = makeString(request.POST['nbr_periode_gl'])
		#print(f'header_nbr_periode_gl_id: {header_nbr_periode_gl_id}')

		header_nbr_periode_ar = makeString(request.POST['nbr_periode_ar'])
		#print(f'header_nbr_periode_ar_id: {header_nbr_periode_ar_id}')

		header_nbr_periode_ap = makeString(request.POST['nbr_periode_ap'])
		#print(f'header_nbr_periode_ap_id: {header_nbr_periode_ap_id}')

		header_nbr_periode_cm = makeString(request.POST['nbr_periode_cm'])
		#print(f'header_nbr_periode_cm_id: {header_nbr_periode_cm_id}')

		header_nbr_periode_fa = makeString(request.POST['nbr_periode_fa'])
		#print(f'header_nbr_periode_fa_id: {header_nbr_periode_fa_id}')

		header_nbr_periode_bgt = makeString(request.POST['nbr_periode_bgt'])
		#print(f'header_nbr_periode_bgt_id: {header_nbr_periode_bgt_id}')

		header_nbr_periode_py = makeString(request.POST['nbr_periode_py'])
		#print(f'header_nbr_periode_py_id: {header_nbr_periode_py_id}')

		header_period_begin_date = makeString(request.POST['period_begin_date'])
		#print(f'header_period_begin_date_id: {header_period_begin_date_id}')

		header_period_end_date = makeString(request.POST['period_end_date'])
		#print(f'header_period_end_date_id: {header_period_end_date_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		for i in df.index:

			code = ''
			if header_code != '': code = makeString(df[header_code][i])
			if code in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom de la société\' est obligatoire, Veuillez le renseigner SVP!')

			type = ''
			if header_type != '': type = makeString(df[header_type][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			devise_id = None
			if header_devise_id != '': devise_id = makeIntId(str(df[header_devise_id][i]))

			type_periode_id = None
			if header_type_periode_id != '': type_periode_id = makeIntId(str(df[header_type_periode_id][i]))

			adress_email = ''
			if header_adress_email != '': adress_email = makeString(df[header_adress_email][i])

			siteweb = ''
			if header_siteweb != '': siteweb = makeString(df[header_siteweb][i])

			pays_id = None
			if header_pays_id != '': pays_id = makeIntId(str(df[header_pays_id][i]))

			pays_adress = ''
			if header_pays_adress != '': pays_adress = makeString(df[header_pays_adress][i])

			province_id = None
			if header_province_id != '': province_id = makeIntId(str(df[header_province_id][i]))

			province_adress = ''
			if header_province_adress != '': province_adress = makeString(df[header_province_adress][i])

			ville_id = None
			if header_ville_id != '': ville_id = makeIntId(str(df[header_ville_id][i]))

			ville_adress = ''
			if header_ville_adress != '': ville_adress = makeString(df[header_ville_adress][i])

			commune_id = None
			if header_commune_id != '': commune_id = makeIntId(str(df[header_commune_id][i]))

			adresse_line1 = ''
			if header_adresse_line1 != '': adresse_line1 = makeString(df[header_adresse_line1][i])

			adresse_line2 = ''
			if header_adresse_line2 != '': adresse_line2 = makeString(df[header_adresse_line2][i])

			telephone_1 = ''
			if header_telephone_1 != '': telephone_1 = makeString(df[header_telephone_1][i])

			telephone_2 = ''
			if header_telephone_2 != '': telephone_2 = makeString(df[header_telephone_2][i])

			nbr_periode_gl = 0
			if header_nbr_periode_gl != '': nbr_periode_gl = makeInt(df[header_nbr_periode_gl][i])

			nbr_periode_ar = 0
			if header_nbr_periode_ar != '': nbr_periode_ar = makeInt(df[header_nbr_periode_ar][i])

			nbr_periode_ap = 0
			if header_nbr_periode_ap != '': nbr_periode_ap = makeInt(df[header_nbr_periode_ap][i])

			nbr_periode_cm = 0
			if header_nbr_periode_cm != '': nbr_periode_cm = makeInt(df[header_nbr_periode_cm][i])

			nbr_periode_fa = 0
			if header_nbr_periode_fa != '': nbr_periode_fa = makeInt(df[header_nbr_periode_fa][i])

			nbr_periode_bgt = 0
			if header_nbr_periode_bgt != '': nbr_periode_bgt = makeInt(df[header_nbr_periode_bgt][i])

			nbr_periode_py = 0
			if header_nbr_periode_py != '': nbr_periode_py = makeInt(df[header_nbr_periode_py][i])

			period_begin_date = None
			if header_period_begin_date != '': period_begin_date = df[header_period_begin_date][i]

			period_end_date = None
			if header_period_end_date != '': period_end_date = df[header_period_end_date][i]

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe = dao_societe.toCreate(code = code, name = name, type = type, societe_id = societe_id, devise_id = devise_id, type_periode_id = type_periode_id, adress_email = adress_email, siteweb = siteweb, pays_id = pays_id, pays_adress = pays_adress, province_id = province_id, province_adress = province_adress, ville_id = ville_id, ville_adress = ville_adress, commune_id = commune_id, adresse_line1 = adresse_line1, adresse_line2 = adresse_line2, telephone_1 = telephone_1, telephone_2 = telephone_2, nbr_periode_gl = nbr_periode_gl, nbr_periode_ar = nbr_periode_ar, nbr_periode_ap = nbr_periode_ap, nbr_periode_cm = nbr_periode_cm, nbr_periode_fa = nbr_periode_fa, nbr_periode_bgt = nbr_periode_bgt, nbr_periode_py = nbr_periode_py, period_begin_date = period_begin_date, period_end_date = period_end_date, description = description)
			saved, societe, message = dao_societe.toSave(auteur, societe)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_societe'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# SOCIETE API CONTROLLERS
def get_list_societe(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_societe.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'code' : str(item.code),
				'name' : str(item.name),
				'picture_icon' : item.picture_icon.url if item.picture_icon != None else None,
				'type' : str(item.type),
				'societe_id' : makeIntId(item.societe_id),
				'devise_id' : makeIntId(item.devise_id),
				'type_periode_id' : makeIntId(item.type_periode_id),
				'adress_email' : str(item.adress_email),
				'siteweb' : str(item.siteweb),
				'pays_id' : makeIntId(item.pays_id),
				'pays_adress' : str(item.pays_adress),
				'province_id' : makeIntId(item.province_id),
				'province_adress' : str(item.province_adress),
				'ville_id' : makeIntId(item.ville_id),
				'ville_adress' : str(item.ville_adress),
				'commune_id' : makeIntId(item.commune_id),
				'adresse_line1' : str(item.adresse_line1),
				'adresse_line2' : str(item.adresse_line2),
				'telephone_1' : str(item.telephone_1),
				'telephone_2' : str(item.telephone_2),
				'nbr_periode_gl' : makeInt(item.nbr_periode_gl),
				'nbr_periode_ar' : makeInt(item.nbr_periode_ar),
				'nbr_periode_ap' : makeInt(item.nbr_periode_ap),
				'nbr_periode_cm' : makeInt(item.nbr_periode_cm),
				'nbr_periode_fa' : makeInt(item.nbr_periode_fa),
				'nbr_periode_bgt' : makeInt(item.nbr_periode_bgt),
				'nbr_periode_py' : makeInt(item.nbr_periode_py),
				'period_begin_date' : item.period_begin_date,
				'period_end_date' : item.period_end_date,
				'description' : str(item.description),
				'statut_id' : makeIntId(item.statut_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_societe(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_societe.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'code' : str(model.code),
				'name' : str(model.name),
				'picture_icon' : model.picture_icon.url if model.picture_icon != None else None,
				'type' : str(model.type),
				'societe_id' : makeIntId(model.societe_id),
				'devise_id' : makeIntId(model.devise_id),
				'type_periode_id' : makeIntId(model.type_periode_id),
				'adress_email' : str(model.adress_email),
				'siteweb' : str(model.siteweb),
				'pays_id' : makeIntId(model.pays_id),
				'pays_adress' : str(model.pays_adress),
				'province_id' : makeIntId(model.province_id),
				'province_adress' : str(model.province_adress),
				'ville_id' : makeIntId(model.ville_id),
				'ville_adress' : str(model.ville_adress),
				'commune_id' : makeIntId(model.commune_id),
				'adresse_line1' : str(model.adresse_line1),
				'adresse_line2' : str(model.adresse_line2),
				'telephone_1' : str(model.telephone_1),
				'telephone_2' : str(model.telephone_2),
				'nbr_periode_gl' : makeInt(model.nbr_periode_gl),
				'nbr_periode_ar' : makeInt(model.nbr_periode_ar),
				'nbr_periode_ap' : makeInt(model.nbr_periode_ap),
				'nbr_periode_cm' : makeInt(model.nbr_periode_cm),
				'nbr_periode_fa' : makeInt(model.nbr_periode_fa),
				'nbr_periode_bgt' : makeInt(model.nbr_periode_bgt),
				'nbr_periode_py' : makeInt(model.nbr_periode_py),
				'period_begin_date' : model.period_begin_date,
				'period_end_date' : model.period_end_date,
				'description' : str(model.description),
				'statut_id' : makeIntId(model.statut_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_societe(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		code = ''
		if 'code' in request.POST : code = str(request.POST['code'])
		if code in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')

		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom de la société\' est obligatoire, Veuillez le renseigner SVP!')

		picture_icon = request.FILES['picture_icon'] if 'picture_icon' in request.FILES else None

		type = ''
		if 'type' in request.POST : type = str(request.POST['type'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		devise_id = None
		if 'devise' in request.POST : devise_id = makeIntId(request.POST['devise_id'])

		type_periode_id = None
		if 'type_periode' in request.POST : type_periode_id = makeIntId(request.POST['type_periode_id'])

		adress_email = ''
		if 'adress_email' in request.POST : adress_email = str(request.POST['adress_email'])

		siteweb = ''
		if 'siteweb' in request.POST : siteweb = str(request.POST['siteweb'])

		pays_id = None
		if 'pays' in request.POST : pays_id = makeIntId(request.POST['pays_id'])

		pays_adress = ''
		if 'pays_adress' in request.POST : pays_adress = str(request.POST['pays_adress'])

		province_id = None
		if 'province' in request.POST : province_id = makeIntId(request.POST['province_id'])

		province_adress = ''
		if 'province_adress' in request.POST : province_adress = str(request.POST['province_adress'])

		ville_id = None
		if 'ville' in request.POST : ville_id = makeIntId(request.POST['ville_id'])

		ville_adress = ''
		if 'ville_adress' in request.POST : ville_adress = str(request.POST['ville_adress'])

		commune_id = None
		if 'commune' in request.POST : commune_id = makeIntId(request.POST['commune_id'])

		adresse_line1 = ''
		if 'adresse_line1' in request.POST : adresse_line1 = str(request.POST['adresse_line1'])

		adresse_line2 = ''
		if 'adresse_line2' in request.POST : adresse_line2 = str(request.POST['adresse_line2'])

		telephone_1 = ''
		if 'telephone_1' in request.POST : telephone_1 = str(request.POST['telephone_1'])

		telephone_2 = ''
		if 'telephone_2' in request.POST : telephone_2 = str(request.POST['telephone_2'])

		nbr_periode_gl = 0
		if 'nbr_periode_gl' in request.POST : nbr_periode_gl = makeInt(request.POST['nbr_periode_gl'])

		nbr_periode_ar = 0
		if 'nbr_periode_ar' in request.POST : nbr_periode_ar = makeInt(request.POST['nbr_periode_ar'])

		nbr_periode_ap = 0
		if 'nbr_periode_ap' in request.POST : nbr_periode_ap = makeInt(request.POST['nbr_periode_ap'])

		nbr_periode_cm = 0
		if 'nbr_periode_cm' in request.POST : nbr_periode_cm = makeInt(request.POST['nbr_periode_cm'])

		nbr_periode_fa = 0
		if 'nbr_periode_fa' in request.POST : nbr_periode_fa = makeInt(request.POST['nbr_periode_fa'])

		nbr_periode_bgt = 0
		if 'nbr_periode_bgt' in request.POST : nbr_periode_bgt = makeInt(request.POST['nbr_periode_bgt'])

		nbr_periode_py = 0
		if 'nbr_periode_py' in request.POST : nbr_periode_py = makeInt(request.POST['nbr_periode_py'])

		period_begin_date = ''
		if 'period_begin_date' in request.POST : period_begin_date = str(request.POST['period_begin_date'])
		period_begin_date = timezone.datetime(int(period_begin_date[6:10]), int(period_begin_date[3:5]), int(period_begin_date[0:2]), int(period_begin_date[11:13]), int(period_begin_date[14:16]))

		period_end_date = ''
		if 'period_end_date' in request.POST : period_end_date = str(request.POST['period_end_date'])
		period_end_date = timezone.datetime(int(period_end_date[6:10]), int(period_end_date[3:5]), int(period_end_date[0:2]), int(period_end_date[11:13]), int(period_end_date[14:16]))

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		autres_adresses = []

		contacts = []

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		societe = dao_societe.toCreate(code = code, name = name, picture_icon = picture_icon, type = type, societe_id = societe_id, devise_id = devise_id, type_periode_id = type_periode_id, adress_email = adress_email, siteweb = siteweb, pays_id = pays_id, pays_adress = pays_adress, province_id = province_id, province_adress = province_adress, ville_id = ville_id, ville_adress = ville_adress, commune_id = commune_id, adresse_line1 = adresse_line1, adresse_line2 = adresse_line2, telephone_1 = telephone_1, telephone_2 = telephone_2, nbr_periode_gl = nbr_periode_gl, nbr_periode_ar = nbr_periode_ar, nbr_periode_ap = nbr_periode_ap, nbr_periode_cm = nbr_periode_cm, nbr_periode_fa = nbr_periode_fa, nbr_periode_bgt = nbr_periode_bgt, nbr_periode_py = nbr_periode_py, period_begin_date = period_begin_date, period_end_date = period_end_date, description = description, autres_adresses = autres_adresses, contacts = contacts)
		saved, societe, message = dao_societe.toSave(auteur, societe)

		if saved == False: raise Exception(message)

		objet = {
			'id' : societe.id,
			'code' : str(societe.code),
			'name' : str(societe.name),
			'picture_icon' : societe.picture_icon.url if societe.picture_icon != None else None,
			'type' : str(societe.type),
			'societe_id' : makeIntId(societe.societe_id),
			'devise_id' : makeIntId(societe.devise_id),
			'type_periode_id' : makeIntId(societe.type_periode_id),
			'adress_email' : str(societe.adress_email),
			'siteweb' : str(societe.siteweb),
			'pays_id' : makeIntId(societe.pays_id),
			'pays_adress' : str(societe.pays_adress),
			'province_id' : makeIntId(societe.province_id),
			'province_adress' : str(societe.province_adress),
			'ville_id' : makeIntId(societe.ville_id),
			'ville_adress' : str(societe.ville_adress),
			'commune_id' : makeIntId(societe.commune_id),
			'adresse_line1' : str(societe.adresse_line1),
			'adresse_line2' : str(societe.adresse_line2),
			'telephone_1' : str(societe.telephone_1),
			'telephone_2' : str(societe.telephone_2),
			'nbr_periode_gl' : makeInt(societe.nbr_periode_gl),
			'nbr_periode_ar' : makeInt(societe.nbr_periode_ar),
			'nbr_periode_ap' : makeInt(societe.nbr_periode_ap),
			'nbr_periode_cm' : makeInt(societe.nbr_periode_cm),
			'nbr_periode_fa' : makeInt(societe.nbr_periode_fa),
			'nbr_periode_bgt' : makeInt(societe.nbr_periode_bgt),
			'nbr_periode_py' : makeInt(societe.nbr_periode_py),
			'period_begin_date' : societe.period_begin_date,
			'period_end_date' : societe.period_end_date,
			'description' : str(societe.description),
			'statut_id' : makeIntId(societe.statut_id),
			'etat' : str(societe.etat),
			'creation_date' : societe.creation_date,
			'update_date' : societe.update_date,
			'update_by_id' : makeIntId(societe.update_by_id),
			'auteur_id' : makeIntId(societe.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

# CONTACT CONTROLLERS
from ModuleConfiguration.dao.dao_contact import dao_contact

def get_lister_contact(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_contact.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_contact.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des contacts",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/contact/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_contact(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - Contact",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_Contact(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'adresses' : Model_Adresse.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/contact/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_contact'))

@transaction.atomic
def post_creer_contact(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_contact'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Noms\' est obligatoire, Veuillez le renseigner SVP!')

		type = makeInt(request.POST['type'])

		nature = str(request.POST['nature'])

		email = str(request.POST['email'])

		siteweb = str(request.POST['siteweb'])

		function = str(request.POST['function'])

		country_id = makeIntId(request.POST['country_id'])

		adress_state_id = makeIntId(request.POST['adress_state_id'])

		adress_city_id = makeIntId(request.POST['adress_city_id'])

		adress_line1 = str(request.POST['adress_line1'])

		adress_line2 = str(request.POST['adress_line2'])

		phone_number = str(request.POST['phone_number'])

		phone_number_2 = str(request.POST['phone_number_2'])

		code_postal = str(request.POST['code_postal'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		autres_adresses = request.POST.getlist('autres_adresses', None)

		auteur = identite.utilisateur(request)

		contact = dao_contact.toCreate(name = name, type = type, nature = nature, email = email, siteweb = siteweb, function = function, country_id = country_id, adress_state_id = adress_state_id, adress_city_id = adress_city_id, adress_line1 = adress_line1, adress_line2 = adress_line2, phone_number = phone_number, phone_number_2 = phone_number_2, code_postal = code_postal, description = description, societe_id = societe_id, autres_adresses = autres_adresses)
		saved, contact, message = dao_contact.toSave(auteur, contact, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_contact.toListById(contact.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, contact)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : contact.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_contact(request,ref):
	try:
		same_perm_with = 'module_configuration_list_contact'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		contact = dao_contact.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(contact.id),'obj': str(contact)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_contact', args=(contact.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_contact(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		contact = auth.toGetWithRules(dao_contact.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if contact == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, contact) 

		context = {
			'title' : "Détails - Contact : {}".format(contact),
			'model' : contact,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/contact/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_contact'))

def get_modifier_contact(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_contact.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - Contact",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'adresses' : Model_Adresse.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/contact/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_contact(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_contact'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Noms\' est obligatoire, Veuillez le renseigner SVP!')

		type = makeInt(request.POST['type'])

		nature = str(request.POST['nature'])

		email = str(request.POST['email'])

		siteweb = str(request.POST['siteweb'])

		function = str(request.POST['function'])

		country_id = makeIntId(request.POST['country_id'])

		adress_state_id = makeIntId(request.POST['adress_state_id'])

		adress_city_id = makeIntId(request.POST['adress_city_id'])

		adress_line1 = str(request.POST['adress_line1'])

		adress_line2 = str(request.POST['adress_line2'])

		phone_number = str(request.POST['phone_number'])

		phone_number_2 = str(request.POST['phone_number_2'])

		code_postal = str(request.POST['code_postal'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		autres_adresses = request.POST.getlist('autres_adresses', None)
		auteur = identite.utilisateur(request)

		contact = dao_contact.toCreate(name = name, type = type, nature = nature, email = email, siteweb = siteweb, function = function, country_id = country_id, adress_state_id = adress_state_id, adress_city_id = adress_city_id, adress_line1 = adress_line1, adress_line2 = adress_line2, phone_number = phone_number, phone_number_2 = phone_number_2, code_postal = code_postal, description = description, societe_id = societe_id, autres_adresses = autres_adresses)
		saved, contact, message = dao_contact.toUpdate(id, contact, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_contact.toListById(contact.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : contact.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_contact(request,ref):
	try:
		same_perm_with = 'module_configuration_add_contact'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_contact.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'adresses' : Model_Adresse.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/contact/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_contact(request,ref):
	try:
		same_perm_with = 'module_configuration_list_contact'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		contact = auth.toGetWithRules(dao_contact.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if contact == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - Contact : {}".format(contact),
			'model' : contact,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_contact.html', 'print_contact.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_contact(request):
	try:
		same_perm_with = 'module_configuration_add_contact'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_contact')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des contacts",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/contact/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_contact(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Noms\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_type = makeString(request.POST['type'])
		#print(f'header_type_id: {header_type_id}')

		header_nature = makeString(request.POST['nature'])
		#print(f'header_nature_id: {header_nature_id}')

		header_email = makeString(request.POST['email'])
		#print(f'header_email_id: {header_email_id}')

		header_siteweb = makeString(request.POST['siteweb'])
		#print(f'header_siteweb_id: {header_siteweb_id}')

		header_function = makeString(request.POST['function'])
		#print(f'header_function_id: {header_function_id}')

		header_country_id = makeString(request.POST['country_id'])
		#print(f'header_country_id: {header_country_id}')

		header_adress_state_id = makeString(request.POST['adress_state_id'])
		#print(f'header_adress_state_id: {header_adress_state_id}')

		header_adress_city_id = makeString(request.POST['adress_city_id'])
		#print(f'header_adress_city_id: {header_adress_city_id}')

		header_adress_line1 = makeString(request.POST['adress_line1'])
		#print(f'header_adress_line1_id: {header_adress_line1_id}')

		header_adress_line2 = makeString(request.POST['adress_line2'])
		#print(f'header_adress_line2_id: {header_adress_line2_id}')

		header_phone_number = makeString(request.POST['phone_number'])
		#print(f'header_phone_number_id: {header_phone_number_id}')

		header_phone_number_2 = makeString(request.POST['phone_number_2'])
		#print(f'header_phone_number_2_id: {header_phone_number_2_id}')

		header_code_postal = makeString(request.POST['code_postal'])
		#print(f'header_code_postal_id: {header_code_postal_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		for i in df.index:

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Noms\' est obligatoire, Veuillez le renseigner SVP!')

			type = 0
			if header_type != '': type = makeInt(df[header_type][i])

			nature = ''
			if header_nature != '': nature = makeString(df[header_nature][i])

			email = ''
			if header_email != '': email = makeString(df[header_email][i])

			siteweb = ''
			if header_siteweb != '': siteweb = makeString(df[header_siteweb][i])

			function = ''
			if header_function != '': function = makeString(df[header_function][i])

			country_id = None
			if header_country_id != '': country_id = makeIntId(str(df[header_country_id][i]))

			adress_state_id = None
			if header_adress_state_id != '': adress_state_id = makeIntId(str(df[header_adress_state_id][i]))

			adress_city_id = None
			if header_adress_city_id != '': adress_city_id = makeIntId(str(df[header_adress_city_id][i]))

			adress_line1 = ''
			if header_adress_line1 != '': adress_line1 = makeString(df[header_adress_line1][i])

			adress_line2 = ''
			if header_adress_line2 != '': adress_line2 = makeString(df[header_adress_line2][i])

			phone_number = ''
			if header_phone_number != '': phone_number = makeString(df[header_phone_number][i])

			phone_number_2 = ''
			if header_phone_number_2 != '': phone_number_2 = makeString(df[header_phone_number_2][i])

			code_postal = ''
			if header_code_postal != '': code_postal = makeString(df[header_code_postal][i])

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			contact = dao_contact.toCreate(name = name, type = type, nature = nature, email = email, siteweb = siteweb, function = function, country_id = country_id, adress_state_id = adress_state_id, adress_city_id = adress_city_id, adress_line1 = adress_line1, adress_line2 = adress_line2, phone_number = phone_number, phone_number_2 = phone_number_2, code_postal = code_postal, description = description, societe_id = societe_id)
			saved, contact, message = dao_contact.toSave(auteur, contact)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_contact'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# CONTACT API CONTROLLERS
def get_list_contact(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_contact.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'name' : str(item.name),
				'type' : makeInt(item.type),
				'nature' : str(item.nature),
				'email' : str(item.email),
				'siteweb' : str(item.siteweb),
				'function' : str(item.function),
				'country_id' : makeIntId(item.country_id),
				'adress_state_id' : makeIntId(item.adress_state_id),
				'adress_city_id' : makeIntId(item.adress_city_id),
				'adress_line1' : str(item.adress_line1),
				'adress_line2' : str(item.adress_line2),
				'phone_number' : str(item.phone_number),
				'phone_number_2' : str(item.phone_number_2),
				'code_postal' : str(item.code_postal),
				'description' : str(item.description),
				'societe_id' : makeIntId(item.societe_id),
				'statut_id' : makeIntId(item.statut_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_contact(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_contact.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'name' : str(model.name),
				'type' : makeInt(model.type),
				'nature' : str(model.nature),
				'email' : str(model.email),
				'siteweb' : str(model.siteweb),
				'function' : str(model.function),
				'country_id' : makeIntId(model.country_id),
				'adress_state_id' : makeIntId(model.adress_state_id),
				'adress_city_id' : makeIntId(model.adress_city_id),
				'adress_line1' : str(model.adress_line1),
				'adress_line2' : str(model.adress_line2),
				'phone_number' : str(model.phone_number),
				'phone_number_2' : str(model.phone_number_2),
				'code_postal' : str(model.code_postal),
				'description' : str(model.description),
				'societe_id' : makeIntId(model.societe_id),
				'statut_id' : makeIntId(model.statut_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_contact(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Noms\' est obligatoire, Veuillez le renseigner SVP!')

		type = 0
		if 'type' in request.POST : type = makeInt(request.POST['type'])

		nature = ''
		if 'nature' in request.POST : nature = str(request.POST['nature'])

		email = ''
		if 'email' in request.POST : email = str(request.POST['email'])

		siteweb = ''
		if 'siteweb' in request.POST : siteweb = str(request.POST['siteweb'])

		function = ''
		if 'function' in request.POST : function = str(request.POST['function'])

		country_id = None
		if 'country' in request.POST : country_id = makeIntId(request.POST['country_id'])

		adress_state_id = None
		if 'adress_state' in request.POST : adress_state_id = makeIntId(request.POST['adress_state_id'])

		adress_city_id = None
		if 'adress_city' in request.POST : adress_city_id = makeIntId(request.POST['adress_city_id'])

		adress_line1 = ''
		if 'adress_line1' in request.POST : adress_line1 = str(request.POST['adress_line1'])

		adress_line2 = ''
		if 'adress_line2' in request.POST : adress_line2 = str(request.POST['adress_line2'])

		phone_number = ''
		if 'phone_number' in request.POST : phone_number = str(request.POST['phone_number'])

		phone_number_2 = ''
		if 'phone_number_2' in request.POST : phone_number_2 = str(request.POST['phone_number_2'])

		code_postal = ''
		if 'code_postal' in request.POST : code_postal = str(request.POST['code_postal'])

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		autres_adresses = []

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		contact = dao_contact.toCreate(name = name, type = type, nature = nature, email = email, siteweb = siteweb, function = function, country_id = country_id, adress_state_id = adress_state_id, adress_city_id = adress_city_id, adress_line1 = adress_line1, adress_line2 = adress_line2, phone_number = phone_number, phone_number_2 = phone_number_2, code_postal = code_postal, description = description, societe_id = societe_id, autres_adresses = autres_adresses)
		saved, contact, message = dao_contact.toSave(auteur, contact)

		if saved == False: raise Exception(message)

		objet = {
			'id' : contact.id,
			'name' : str(contact.name),
			'type' : makeInt(contact.type),
			'nature' : str(contact.nature),
			'email' : str(contact.email),
			'siteweb' : str(contact.siteweb),
			'function' : str(contact.function),
			'country_id' : makeIntId(contact.country_id),
			'adress_state_id' : makeIntId(contact.adress_state_id),
			'adress_city_id' : makeIntId(contact.adress_city_id),
			'adress_line1' : str(contact.adress_line1),
			'adress_line2' : str(contact.adress_line2),
			'phone_number' : str(contact.phone_number),
			'phone_number_2' : str(contact.phone_number_2),
			'code_postal' : str(contact.code_postal),
			'description' : str(contact.description),
			'societe_id' : makeIntId(contact.societe_id),
			'statut_id' : makeIntId(contact.statut_id),
			'etat' : str(contact.etat),
			'creation_date' : contact.creation_date,
			'update_date' : contact.update_date,
			'update_by_id' : makeIntId(contact.update_by_id),
			'auteur_id' : makeIntId(contact.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

# ADRESSE CONTROLLERS
from ModuleConfiguration.dao.dao_adresse import dao_adresse

def get_lister_adresse(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_adresse.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_adresse.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des adresses",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/adresse/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_adresse(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - Adresse",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_Adresse(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'communes' : Model_Commune.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/adresse/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_adresse'))

@transaction.atomic
def post_creer_adresse(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_adresse'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		type_adresse = makeInt(request.POST['type_adresse'])

		country_id = makeIntId(request.POST['country_id'])

		adress_state_id = makeIntId(request.POST['adress_state_id'])

		adress_city_id = makeIntId(request.POST['adress_city_id'])

		adress_township_id = makeIntId(request.POST['adress_township_id'])

		adress_line1 = str(request.POST['adress_line1'])

		adress_line2 = str(request.POST['adress_line2'])

		code_postal = str(request.POST['code_postal'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		auteur = identite.utilisateur(request)

		adresse = dao_adresse.toCreate(name = name, type_adresse = type_adresse, country_id = country_id, adress_state_id = adress_state_id, adress_city_id = adress_city_id, adress_township_id = adress_township_id, adress_line1 = adress_line1, adress_line2 = adress_line2, code_postal = code_postal, description = description, societe_id = societe_id)
		saved, adresse, message = dao_adresse.toSave(auteur, adresse, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_adresse.toListById(adresse.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, adresse)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : adresse.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_adresse(request,ref):
	try:
		same_perm_with = 'module_configuration_list_adresse'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		adresse = dao_adresse.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(adresse.id),'obj': str(adresse)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_adresse', args=(adresse.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_adresse(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		adresse = auth.toGetWithRules(dao_adresse.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if adresse == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, adresse) 

		context = {
			'title' : "Détails - Adresse : {}".format(adresse),
			'model' : adresse,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/adresse/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_adresse'))

def get_modifier_adresse(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_adresse.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - Adresse",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'communes' : Model_Commune.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/adresse/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_adresse(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_adresse'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		type_adresse = makeInt(request.POST['type_adresse'])

		country_id = makeIntId(request.POST['country_id'])

		adress_state_id = makeIntId(request.POST['adress_state_id'])

		adress_city_id = makeIntId(request.POST['adress_city_id'])

		adress_township_id = makeIntId(request.POST['adress_township_id'])

		adress_line1 = str(request.POST['adress_line1'])

		adress_line2 = str(request.POST['adress_line2'])

		code_postal = str(request.POST['code_postal'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])
		auteur = identite.utilisateur(request)

		adresse = dao_adresse.toCreate(name = name, type_adresse = type_adresse, country_id = country_id, adress_state_id = adress_state_id, adress_city_id = adress_city_id, adress_township_id = adress_township_id, adress_line1 = adress_line1, adress_line2 = adress_line2, code_postal = code_postal, description = description, societe_id = societe_id)
		saved, adresse, message = dao_adresse.toUpdate(id, adresse, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_adresse.toListById(adresse.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : adresse.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_adresse(request,ref):
	try:
		same_perm_with = 'module_configuration_add_adresse'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_adresse.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'communes' : Model_Commune.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/adresse/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_adresse(request,ref):
	try:
		same_perm_with = 'module_configuration_list_adresse'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		adresse = auth.toGetWithRules(dao_adresse.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if adresse == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - Adresse : {}".format(adresse),
			'model' : adresse,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_adresse.html', 'print_adresse.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_adresse(request):
	try:
		same_perm_with = 'module_configuration_add_adresse'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_adresse')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des adresses",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/adresse/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_adresse(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_type_adresse = makeString(request.POST['type_adresse'])
		#print(f'header_type_adresse_id: {header_type_adresse_id}')

		header_country_id = makeString(request.POST['country_id'])
		#print(f'header_country_id: {header_country_id}')

		header_adress_state_id = makeString(request.POST['adress_state_id'])
		#print(f'header_adress_state_id: {header_adress_state_id}')

		header_adress_city_id = makeString(request.POST['adress_city_id'])
		#print(f'header_adress_city_id: {header_adress_city_id}')

		header_adress_township_id = makeString(request.POST['adress_township_id'])
		#print(f'header_adress_township_id: {header_adress_township_id}')

		header_adress_line1 = makeString(request.POST['adress_line1'])
		#print(f'header_adress_line1_id: {header_adress_line1_id}')

		header_adress_line2 = makeString(request.POST['adress_line2'])
		#print(f'header_adress_line2_id: {header_adress_line2_id}')

		header_code_postal = makeString(request.POST['code_postal'])
		#print(f'header_code_postal_id: {header_code_postal_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		for i in df.index:

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

			type_adresse = 0
			if header_type_adresse != '': type_adresse = makeInt(df[header_type_adresse][i])

			country_id = None
			if header_country_id != '': country_id = makeIntId(str(df[header_country_id][i]))

			adress_state_id = None
			if header_adress_state_id != '': adress_state_id = makeIntId(str(df[header_adress_state_id][i]))

			adress_city_id = None
			if header_adress_city_id != '': adress_city_id = makeIntId(str(df[header_adress_city_id][i]))

			adress_township_id = None
			if header_adress_township_id != '': adress_township_id = makeIntId(str(df[header_adress_township_id][i]))

			adress_line1 = ''
			if header_adress_line1 != '': adress_line1 = makeString(df[header_adress_line1][i])

			adress_line2 = ''
			if header_adress_line2 != '': adress_line2 = makeString(df[header_adress_line2][i])

			code_postal = ''
			if header_code_postal != '': code_postal = makeString(df[header_code_postal][i])

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			adresse = dao_adresse.toCreate(name = name, type_adresse = type_adresse, country_id = country_id, adress_state_id = adress_state_id, adress_city_id = adress_city_id, adress_township_id = adress_township_id, adress_line1 = adress_line1, adress_line2 = adress_line2, code_postal = code_postal, description = description, societe_id = societe_id)
			saved, adresse, message = dao_adresse.toSave(auteur, adresse)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_adresse'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# ADRESSE API CONTROLLERS
def get_list_adresse(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_adresse.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'name' : str(item.name),
				'type_adresse' : makeInt(item.type_adresse),
				'country_id' : makeIntId(item.country_id),
				'adress_state_id' : makeIntId(item.adress_state_id),
				'adress_city_id' : makeIntId(item.adress_city_id),
				'adress_township_id' : makeIntId(item.adress_township_id),
				'adress_line1' : str(item.adress_line1),
				'adress_line2' : str(item.adress_line2),
				'code_postal' : str(item.code_postal),
				'description' : str(item.description),
				'societe_id' : makeIntId(item.societe_id),
				'statut_id' : makeIntId(item.statut_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_adresse(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_adresse.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'name' : str(model.name),
				'type_adresse' : makeInt(model.type_adresse),
				'country_id' : makeIntId(model.country_id),
				'adress_state_id' : makeIntId(model.adress_state_id),
				'adress_city_id' : makeIntId(model.adress_city_id),
				'adress_township_id' : makeIntId(model.adress_township_id),
				'adress_line1' : str(model.adress_line1),
				'adress_line2' : str(model.adress_line2),
				'code_postal' : str(model.code_postal),
				'description' : str(model.description),
				'societe_id' : makeIntId(model.societe_id),
				'statut_id' : makeIntId(model.statut_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_adresse(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		type_adresse = 0
		if 'type_adresse' in request.POST : type_adresse = makeInt(request.POST['type_adresse'])

		country_id = None
		if 'country' in request.POST : country_id = makeIntId(request.POST['country_id'])

		adress_state_id = None
		if 'adress_state' in request.POST : adress_state_id = makeIntId(request.POST['adress_state_id'])

		adress_city_id = None
		if 'adress_city' in request.POST : adress_city_id = makeIntId(request.POST['adress_city_id'])

		adress_township_id = None
		if 'adress_township' in request.POST : adress_township_id = makeIntId(request.POST['adress_township_id'])

		adress_line1 = ''
		if 'adress_line1' in request.POST : adress_line1 = str(request.POST['adress_line1'])

		adress_line2 = ''
		if 'adress_line2' in request.POST : adress_line2 = str(request.POST['adress_line2'])

		code_postal = ''
		if 'code_postal' in request.POST : code_postal = str(request.POST['code_postal'])

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		adresse = dao_adresse.toCreate(name = name, type_adresse = type_adresse, country_id = country_id, adress_state_id = adress_state_id, adress_city_id = adress_city_id, adress_township_id = adress_township_id, adress_line1 = adress_line1, adress_line2 = adress_line2, code_postal = code_postal, description = description, societe_id = societe_id)
		saved, adresse, message = dao_adresse.toSave(auteur, adresse)

		if saved == False: raise Exception(message)

		objet = {
			'id' : adresse.id,
			'name' : str(adresse.name),
			'type_adresse' : makeInt(adresse.type_adresse),
			'country_id' : makeIntId(adresse.country_id),
			'adress_state_id' : makeIntId(adresse.adress_state_id),
			'adress_city_id' : makeIntId(adresse.adress_city_id),
			'adress_township_id' : makeIntId(adresse.adress_township_id),
			'adress_line1' : str(adresse.adress_line1),
			'adress_line2' : str(adresse.adress_line2),
			'code_postal' : str(adresse.code_postal),
			'description' : str(adresse.description),
			'societe_id' : makeIntId(adresse.societe_id),
			'statut_id' : makeIntId(adresse.statut_id),
			'etat' : str(adresse.etat),
			'creation_date' : adresse.creation_date,
			'update_date' : adresse.update_date,
			'update_by_id' : makeIntId(adresse.update_by_id),
			'auteur_id' : makeIntId(adresse.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

# PAYS CONTROLLERS
from ModuleConfiguration.dao.dao_pays import dao_pays

def get_lister_pays(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_pays.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_pays.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des pays",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/pays/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_pays(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - Pays",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_Pays(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/pays/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_pays'))

@transaction.atomic
def post_creer_pays(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_pays'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		code = str(request.POST['code'])
		if code in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		auteur = identite.utilisateur(request)

		pays = dao_pays.toCreate(name = name, code = code, description = description, societe_id = societe_id)
		saved, pays, message = dao_pays.toSave(auteur, pays, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_pays.toListById(pays.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, pays)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : pays.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_pays(request,ref):
	try:
		same_perm_with = 'module_configuration_list_pays'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		pays = dao_pays.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(pays.id),'obj': str(pays)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_pays', args=(pays.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_pays(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		pays = auth.toGetWithRules(dao_pays.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if pays == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, pays) 

		context = {
			'title' : "Détails - Pays : {}".format(pays),
			'model' : pays,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/pays/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_pays'))

def get_modifier_pays(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_pays.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - Pays",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/pays/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_pays(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_pays'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		code = str(request.POST['code'])
		if code in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])
		auteur = identite.utilisateur(request)

		pays = dao_pays.toCreate(name = name, code = code, description = description, societe_id = societe_id)
		saved, pays, message = dao_pays.toUpdate(id, pays, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_pays.toListById(pays.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : pays.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_pays(request,ref):
	try:
		same_perm_with = 'module_configuration_add_pays'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_pays.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/pays/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_pays(request,ref):
	try:
		same_perm_with = 'module_configuration_list_pays'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		pays = auth.toGetWithRules(dao_pays.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if pays == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - Pays : {}".format(pays),
			'model' : pays,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_pays.html', 'print_pays.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_pays(request):
	try:
		same_perm_with = 'module_configuration_add_pays'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_pays')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des pays",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/pays/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_pays(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_code = makeString(request.POST['code'])
		if header_code in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_code_id: {header_code_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		for i in df.index:

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

			code = ''
			if header_code != '': code = makeString(df[header_code][i])
			if code in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			pays = dao_pays.toCreate(name = name, code = code, description = description, societe_id = societe_id)
			saved, pays, message = dao_pays.toSave(auteur, pays)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_pays'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# PAYS API CONTROLLERS
def get_list_pays(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_pays.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'name' : str(item.name),
				'code' : str(item.code),
				'description' : str(item.description),
				'societe_id' : makeIntId(item.societe_id),
				'statut_id' : makeIntId(item.statut_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_pays(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_pays.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'name' : str(model.name),
				'code' : str(model.code),
				'description' : str(model.description),
				'societe_id' : makeIntId(model.societe_id),
				'statut_id' : makeIntId(model.statut_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_pays(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		code = ''
		if 'code' in request.POST : code = str(request.POST['code'])
		if code in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Code\' est obligatoire, Veuillez le renseigner SVP!')

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		pays = dao_pays.toCreate(name = name, code = code, description = description, societe_id = societe_id)
		saved, pays, message = dao_pays.toSave(auteur, pays)

		if saved == False: raise Exception(message)

		objet = {
			'id' : pays.id,
			'name' : str(pays.name),
			'code' : str(pays.code),
			'description' : str(pays.description),
			'societe_id' : makeIntId(pays.societe_id),
			'statut_id' : makeIntId(pays.statut_id),
			'etat' : str(pays.etat),
			'creation_date' : pays.creation_date,
			'update_date' : pays.update_date,
			'update_by_id' : makeIntId(pays.update_by_id),
			'auteur_id' : makeIntId(pays.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

# PROVINCE CONTROLLERS
from ModuleConfiguration.dao.dao_province import dao_province

def get_lister_province(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_province.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_province.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des provinces",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/province/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_province(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - Province",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_Province(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/province/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_province'))

@transaction.atomic
def post_creer_province(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_province'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		country_id = makeIntId(request.POST['country_id'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		auteur = identite.utilisateur(request)

		province = dao_province.toCreate(name = name, country_id = country_id, description = description, societe_id = societe_id)
		saved, province, message = dao_province.toSave(auteur, province, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_province.toListById(province.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, province)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : province.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_province(request,ref):
	try:
		same_perm_with = 'module_configuration_list_province'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		province = dao_province.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(province.id),'obj': str(province)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_province', args=(province.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_province(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		province = auth.toGetWithRules(dao_province.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if province == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, province) 

		context = {
			'title' : "Détails - Province : {}".format(province),
			'model' : province,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/province/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_province'))

def get_modifier_province(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_province.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - Province",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/province/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_province(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_province'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		country_id = makeIntId(request.POST['country_id'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])
		auteur = identite.utilisateur(request)

		province = dao_province.toCreate(name = name, country_id = country_id, description = description, societe_id = societe_id)
		saved, province, message = dao_province.toUpdate(id, province, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_province.toListById(province.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : province.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_province(request,ref):
	try:
		same_perm_with = 'module_configuration_add_province'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_province.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'payss' : Model_Pays.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/province/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_province(request,ref):
	try:
		same_perm_with = 'module_configuration_list_province'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		province = auth.toGetWithRules(dao_province.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if province == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - Province : {}".format(province),
			'model' : province,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_province.html', 'print_province.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_province(request):
	try:
		same_perm_with = 'module_configuration_add_province'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_province')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des provinces",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/province/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_province(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_country_id = makeString(request.POST['country_id'])
		#print(f'header_country_id: {header_country_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		for i in df.index:

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

			country_id = None
			if header_country_id != '': country_id = makeIntId(str(df[header_country_id][i]))

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			province = dao_province.toCreate(name = name, country_id = country_id, description = description, societe_id = societe_id)
			saved, province, message = dao_province.toSave(auteur, province)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_province'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# PROVINCE API CONTROLLERS
def get_list_province(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_province.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'name' : str(item.name),
				'country_id' : makeIntId(item.country_id),
				'description' : str(item.description),
				'societe_id' : makeIntId(item.societe_id),
				'statut_id' : makeIntId(item.statut_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_province(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_province.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'name' : str(model.name),
				'country_id' : makeIntId(model.country_id),
				'description' : str(model.description),
				'societe_id' : makeIntId(model.societe_id),
				'statut_id' : makeIntId(model.statut_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_province(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		country_id = None
		if 'country' in request.POST : country_id = makeIntId(request.POST['country_id'])

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		province = dao_province.toCreate(name = name, country_id = country_id, description = description, societe_id = societe_id)
		saved, province, message = dao_province.toSave(auteur, province)

		if saved == False: raise Exception(message)

		objet = {
			'id' : province.id,
			'name' : str(province.name),
			'country_id' : makeIntId(province.country_id),
			'description' : str(province.description),
			'societe_id' : makeIntId(province.societe_id),
			'statut_id' : makeIntId(province.statut_id),
			'etat' : str(province.etat),
			'creation_date' : province.creation_date,
			'update_date' : province.update_date,
			'update_by_id' : makeIntId(province.update_by_id),
			'auteur_id' : makeIntId(province.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

# VILLE CONTROLLERS
from ModuleConfiguration.dao.dao_ville import dao_ville

def get_lister_ville(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_ville.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_ville.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des villes",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/ville/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_ville(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - Ville",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_Ville(),
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/ville/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_ville'))

@transaction.atomic
def post_creer_ville(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_ville'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		province_id = makeIntId(request.POST['province_id'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		auteur = identite.utilisateur(request)

		ville = dao_ville.toCreate(name = name, province_id = province_id, description = description, societe_id = societe_id)
		saved, ville, message = dao_ville.toSave(auteur, ville, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_ville.toListById(ville.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, ville)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : ville.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_ville(request,ref):
	try:
		same_perm_with = 'module_configuration_list_ville'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ville = dao_ville.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(ville.id),'obj': str(ville)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_ville', args=(ville.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_ville(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		ville = auth.toGetWithRules(dao_ville.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if ville == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, ville) 

		context = {
			'title' : "Détails - Ville : {}".format(ville),
			'model' : ville,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/ville/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_ville'))

def get_modifier_ville(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_ville.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - Ville",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/ville/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_ville(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_ville'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		province_id = makeIntId(request.POST['province_id'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])
		auteur = identite.utilisateur(request)

		ville = dao_ville.toCreate(name = name, province_id = province_id, description = description, societe_id = societe_id)
		saved, ville, message = dao_ville.toUpdate(id, ville, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_ville.toListById(ville.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : ville.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_ville(request,ref):
	try:
		same_perm_with = 'module_configuration_add_ville'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_ville.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/ville/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_ville(request,ref):
	try:
		same_perm_with = 'module_configuration_list_ville'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		ville = auth.toGetWithRules(dao_ville.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if ville == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - Ville : {}".format(ville),
			'model' : ville,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_ville.html', 'print_ville.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_ville(request):
	try:
		same_perm_with = 'module_configuration_add_ville'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_ville')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des villes",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/ville/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_ville(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_province_id = makeString(request.POST['province_id'])
		#print(f'header_province_id: {header_province_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		for i in df.index:

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

			province_id = None
			if header_province_id != '': province_id = makeIntId(str(df[header_province_id][i]))

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			ville = dao_ville.toCreate(name = name, province_id = province_id, description = description, societe_id = societe_id)
			saved, ville, message = dao_ville.toSave(auteur, ville)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_ville'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# VILLE API CONTROLLERS
def get_list_ville(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_ville.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'name' : str(item.name),
				'province_id' : makeIntId(item.province_id),
				'description' : str(item.description),
				'societe_id' : makeIntId(item.societe_id),
				'statut_id' : makeIntId(item.statut_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_ville(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_ville.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'name' : str(model.name),
				'province_id' : makeIntId(model.province_id),
				'description' : str(model.description),
				'societe_id' : makeIntId(model.societe_id),
				'statut_id' : makeIntId(model.statut_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_ville(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		province_id = None
		if 'province' in request.POST : province_id = makeIntId(request.POST['province_id'])

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		ville = dao_ville.toCreate(name = name, province_id = province_id, description = description, societe_id = societe_id)
		saved, ville, message = dao_ville.toSave(auteur, ville)

		if saved == False: raise Exception(message)

		objet = {
			'id' : ville.id,
			'name' : str(ville.name),
			'province_id' : makeIntId(ville.province_id),
			'description' : str(ville.description),
			'societe_id' : makeIntId(ville.societe_id),
			'statut_id' : makeIntId(ville.statut_id),
			'etat' : str(ville.etat),
			'creation_date' : ville.creation_date,
			'update_date' : ville.update_date,
			'update_by_id' : makeIntId(ville.update_by_id),
			'auteur_id' : makeIntId(ville.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

# DISTRICT CONTROLLERS
from ModuleConfiguration.dao.dao_district import dao_district

def get_lister_district(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_district.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_district.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des districts",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/district/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_district(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - District",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_District(),
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/district/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_district'))

@transaction.atomic
def post_creer_district(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_district'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		province_id = makeIntId(request.POST['province_id'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		auteur = identite.utilisateur(request)

		district = dao_district.toCreate(name = name, province_id = province_id, description = description, societe_id = societe_id)
		saved, district, message = dao_district.toSave(auteur, district, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_district.toListById(district.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, district)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : district.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_district(request,ref):
	try:
		same_perm_with = 'module_configuration_list_district'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		district = dao_district.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(district.id),'obj': str(district)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_district', args=(district.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_district(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		district = auth.toGetWithRules(dao_district.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if district == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, district) 

		context = {
			'title' : "Détails - District : {}".format(district),
			'model' : district,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/district/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_district'))

def get_modifier_district(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_district.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - District",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/district/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_district(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_district'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		province_id = makeIntId(request.POST['province_id'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])
		auteur = identite.utilisateur(request)

		district = dao_district.toCreate(name = name, province_id = province_id, description = description, societe_id = societe_id)
		saved, district, message = dao_district.toUpdate(id, district, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_district.toListById(district.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : district.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_district(request,ref):
	try:
		same_perm_with = 'module_configuration_add_district'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_district.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'provinces' : Model_Province.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/district/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_district(request,ref):
	try:
		same_perm_with = 'module_configuration_list_district'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		district = auth.toGetWithRules(dao_district.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if district == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - District : {}".format(district),
			'model' : district,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_district.html', 'print_district.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_district(request):
	try:
		same_perm_with = 'module_configuration_add_district'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_district')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des districts",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/district/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_district(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_province_id = makeString(request.POST['province_id'])
		#print(f'header_province_id: {header_province_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		for i in df.index:

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

			province_id = None
			if header_province_id != '': province_id = makeIntId(str(df[header_province_id][i]))

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			district = dao_district.toCreate(name = name, province_id = province_id, description = description, societe_id = societe_id)
			saved, district, message = dao_district.toSave(auteur, district)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_district'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# DISTRICT API CONTROLLERS
def get_list_district(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_district.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'name' : str(item.name),
				'province_id' : makeIntId(item.province_id),
				'description' : str(item.description),
				'societe_id' : makeIntId(item.societe_id),
				'statut_id' : makeIntId(item.statut_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_district(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_district.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'name' : str(model.name),
				'province_id' : makeIntId(model.province_id),
				'description' : str(model.description),
				'societe_id' : makeIntId(model.societe_id),
				'statut_id' : makeIntId(model.statut_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_district(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		province_id = None
		if 'province' in request.POST : province_id = makeIntId(request.POST['province_id'])

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		district = dao_district.toCreate(name = name, province_id = province_id, description = description, societe_id = societe_id)
		saved, district, message = dao_district.toSave(auteur, district)

		if saved == False: raise Exception(message)

		objet = {
			'id' : district.id,
			'name' : str(district.name),
			'province_id' : makeIntId(district.province_id),
			'description' : str(district.description),
			'societe_id' : makeIntId(district.societe_id),
			'statut_id' : makeIntId(district.statut_id),
			'etat' : str(district.etat),
			'creation_date' : district.creation_date,
			'update_date' : district.update_date,
			'update_by_id' : makeIntId(district.update_by_id),
			'auteur_id' : makeIntId(district.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

# COMMUNE CONTROLLERS
from ModuleConfiguration.dao.dao_commune import dao_commune

def get_lister_commune(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_commune.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_commune.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des communes",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/commune/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_commune(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - Commune",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_Commune(),
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/commune/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_commune'))

@transaction.atomic
def post_creer_commune(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_commune'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		ville_id = makeIntId(request.POST['ville_id'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		auteur = identite.utilisateur(request)

		commune = dao_commune.toCreate(name = name, ville_id = ville_id, description = description, societe_id = societe_id)
		saved, commune, message = dao_commune.toSave(auteur, commune, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_commune.toListById(commune.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, commune)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : commune.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_commune(request,ref):
	try:
		same_perm_with = 'module_configuration_list_commune'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		commune = dao_commune.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(commune.id),'obj': str(commune)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_commune', args=(commune.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_commune(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		commune = auth.toGetWithRules(dao_commune.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if commune == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, commune) 

		context = {
			'title' : "Détails - Commune : {}".format(commune),
			'model' : commune,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/commune/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_commune'))

def get_modifier_commune(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_commune.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - Commune",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/commune/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_commune(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_commune'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		ville_id = makeIntId(request.POST['ville_id'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])
		auteur = identite.utilisateur(request)

		commune = dao_commune.toCreate(name = name, ville_id = ville_id, description = description, societe_id = societe_id)
		saved, commune, message = dao_commune.toUpdate(id, commune, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_commune.toListById(commune.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : commune.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_commune(request,ref):
	try:
		same_perm_with = 'module_configuration_add_commune'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_commune.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'villes' : Model_Ville.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/commune/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_commune(request,ref):
	try:
		same_perm_with = 'module_configuration_list_commune'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		commune = auth.toGetWithRules(dao_commune.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if commune == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - Commune : {}".format(commune),
			'model' : commune,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_commune.html', 'print_commune.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_commune(request):
	try:
		same_perm_with = 'module_configuration_add_commune'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_commune')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des communes",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/commune/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_commune(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_ville_id = makeString(request.POST['ville_id'])
		#print(f'header_ville_id: {header_ville_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		for i in df.index:

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

			ville_id = None
			if header_ville_id != '': ville_id = makeIntId(str(df[header_ville_id][i]))

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			commune = dao_commune.toCreate(name = name, ville_id = ville_id, description = description, societe_id = societe_id)
			saved, commune, message = dao_commune.toSave(auteur, commune)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_commune'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# COMMUNE API CONTROLLERS
def get_list_commune(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_commune.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'name' : str(item.name),
				'ville_id' : makeIntId(item.ville_id),
				'description' : str(item.description),
				'societe_id' : makeIntId(item.societe_id),
				'statut_id' : makeIntId(item.statut_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_commune(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_commune.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'name' : str(model.name),
				'ville_id' : makeIntId(model.ville_id),
				'description' : str(model.description),
				'societe_id' : makeIntId(model.societe_id),
				'statut_id' : makeIntId(model.statut_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_commune(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Nom\' est obligatoire, Veuillez le renseigner SVP!')

		ville_id = None
		if 'ville' in request.POST : ville_id = makeIntId(request.POST['ville_id'])

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		commune = dao_commune.toCreate(name = name, ville_id = ville_id, description = description, societe_id = societe_id)
		saved, commune, message = dao_commune.toSave(auteur, commune)

		if saved == False: raise Exception(message)

		objet = {
			'id' : commune.id,
			'name' : str(commune.name),
			'ville_id' : makeIntId(commune.ville_id),
			'description' : str(commune.description),
			'societe_id' : makeIntId(commune.societe_id),
			'statut_id' : makeIntId(commune.statut_id),
			'etat' : str(commune.etat),
			'creation_date' : commune.creation_date,
			'update_date' : commune.update_date,
			'update_by_id' : makeIntId(commune.update_by_id),
			'auteur_id' : makeIntId(commune.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())


# TYPE_PERIODE CONTROLLERS
from ModuleConfiguration.dao.dao_type_periode import dao_type_periode

def get_lister_type_periode(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		view, query, page, count = utils.get_list_request(request)
		#print(f'view {view} query {query} page {page} count {count}')

		#*******Filtre sur les règles **********#
		model = auth.toListWithRules(dao_type_periode.toList(query,utilisateur), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		model = pagination.toGetData(model, page, count)

		if request.method == 'POST':
			context = {
				'error' : False,
				'message' : 'Recupération effectuée avec succès',
				'model' : dao_type_periode.toListJson(model.object_list),
				'view' : view,
				'query' : query,
				'page' : page,
				'count' : count,
			}
			context = pagination.toAddVarsToContext(model, context)
			return JsonResponse(context, safe=False)

		isPopup = False
		if 'isPopup' in request.GET:
			isPopup = True
			view = 'list'

		context = {
			'title' : "Liste des types période",
			'model' : model,
			'view' : view,
			'query' : query,
			'page' : page,
			'count' : count,
			'isPopup' : isPopup,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation()
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/type_periode/list.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		if request.method == 'POST': return auth.toReturnApiFailed(request, e, traceback.format_exc())
		else: return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_index'))

def get_creer_type_periode(request):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		context = {
			'title' : "Formulaire d'enregistrement - Type Période",
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
			'model' : Model_Type_periode(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/type_periode/add.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_type_periode'))

@transaction.atomic
def post_creer_type_periode(request):
	sid = transaction.savepoint()
	try:
		same_perm_with = 'module_configuration_add_type_periode'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Désignation\' est obligatoire, Veuillez le renseigner SVP!')

		periodicite = str(request.POST['periodicite'])

		nombre_par_exercice = makeInt(request.POST['nombre_par_exercice'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])

		auteur = identite.utilisateur(request)

		type_periode = dao_type_periode.toCreate(name = name, periodicite = periodicite, nombre_par_exercice = nombre_par_exercice, description = description, societe_id = societe_id)
		saved, type_periode, message = dao_type_periode.toSave(auteur, type_periode, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_type_periode.toListById(type_periode.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la création', '', msg = 'Vous n\'êtes pas habilité(e) de créer cet objet avec certaines informations que vous avez saisies !')

		#Initialisation du workflow
		wkf_task.initializeWorkflow(auteur, type_periode)

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Enregistrement effectué avec succès',
			'isPopup': isPopup,
			'id' : type_periode.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_select_type_periode(request,ref):
	try:
		same_perm_with = 'module_configuration_list_type_periode'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		type_periode = dao_type_periode.toGet(ref)

		if 'isPopup' in request.GET:
			popup_response_data = json.dumps({'value': str(type_periode.id),'obj': str(type_periode)})
			return TemplateResponse(request, 'ErpProject/ErpBackOffice/popup_response.html', { 'popup_response_data': popup_response_data })

		return HttpResponseRedirect(reverse('module_configuration_detail_type_periode', args=(type_periode.id,)))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_details_type_periode(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		type_periode = auth.toGetWithRules(dao_type_periode.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if type_periode == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		historique, transitions_etapes_suivantes, content_type_id, documents = wkf_task.get_details(utilisateur, type_periode) 

		context = {
			'title' : "Détails - Type Période : {}".format(type_periode),
			'model' : type_periode,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'historique': historique,
			'etapes_suivantes' : transitions_etapes_suivantes,
			'content_type_id': content_type_id,
			'documents': documents,
			'roles': groupe_permissions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/type_periode/item.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc(), reverse('module_configuration_list_type_periode'))

def get_modifier_type_periode(request,ref):
	try:
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request)
		if response != None: return response

		ref = int(ref)
		model = dao_type_periode.toGet(ref)
		context = {
			'title' : "Formulaire de mise à jour - Type Période",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/type_periode/update.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_modifier_type_periode(request):
	sid = transaction.savepoint()
	id = int(request.POST['ref'])
	try:
		same_perm_with = 'module_configuration_update_type_periode'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response


		name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Désignation\' est obligatoire, Veuillez le renseigner SVP!')

		periodicite = str(request.POST['periodicite'])

		nombre_par_exercice = makeInt(request.POST['nombre_par_exercice'])

		description = str(request.POST['description'])

		societe_id = makeIntId(request.POST['societe_id'])
		auteur = identite.utilisateur(request)

		type_periode = dao_type_periode.toCreate(name = name, periodicite = periodicite, nombre_par_exercice = nombre_par_exercice, description = description, societe_id = societe_id)
		saved, type_periode, message = dao_type_periode.toUpdate(id, type_periode, auteur, request.POST)

		if saved == False: raise Exception(message)

		#*******Filtre sur les règles **********#
		model = auth.toGetWithRules(dao_type_periode.toListById(type_periode.id), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if model == None: 
			transaction.savepoint_rollback(sid)
			return auth.toReturnApiFailed(request, 'Erreur: Violation de règle sur la modification', '', msg = 'Vous n\'êtes pas habilité(e) de modifier cet objet avec certaines informations que vous avez saisies !')

		isPopup = 0
		if 'isPopup' in request.POST: isPopup = 1

		transaction.savepoint_commit(sid)
		context = {
			'error' : False,
			'message' : 'Mise à jour effectuée avec succès',
			'isPopup': isPopup,
			'id' : type_periode.id,
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_dupliquer_type_periode(request,ref):
	try:
		same_perm_with = 'module_configuration_add_type_periode'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)
		model = dao_type_periode.toGet(ref)
		context = {
			'title' : "Formulaire d'enregistrement",
			'model':model,
			'utilisateur': utilisateur,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
			'societes' : Model_Societe.objects.filter(Q(societe_id = utilisateur.societe_id) | Q(societe__code = 'MD')).order_by('-id')[:10],
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/type_periode/duplicate.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_imprimer_type_periode(request,ref):
	try:
		same_perm_with = 'module_configuration_list_type_periode'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		ref = int(ref)

		#*******Filtre sur les règles **********#
		type_periode = auth.toGetWithRules(dao_type_periode.toListById(ref), permission, groupe_permissions, utilisateur)
		#******* End Regle *******************#

		if type_periode == None:  return HttpResponseRedirect(reverse('backoffice_erreur_autorisation'))

		context = {
			'title' : "Détails - Type Période : {}".format(type_periode),
			'model' : type_periode,
			'utilisateur' : utilisateur,
			'user_actions': actions,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation': dao_organisation.toGetMainOrganisation(),
		}

		return weasy_print('ErpProject/ModuleConfiguration/reporting/print_type_periode.html', 'print_type_periode.pdf', context, request)
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

def get_upload_type_periode(request):
	try:
		same_perm_with = 'module_configuration_add_type_periode'
		modules, sous_modules, utilisateur, groupe_permissions, permission, actions, response = auth.toGetAuthPerm(request, same_perm_with)
		if response != None: return response

		model_content_type = dao_query_builder.toGetContentTypeByName('model_type_periode')
		champs = dao_query_builder.toListFieldOfModel(model_content_type.id)

		context = {
			'title' : "Import de la liste des types période",
			'utilisateur' : utilisateur,
			'champs': champs,
			'user_actions': actions,
			'isPopup': True if 'isPopup' in request.GET else False,
			'modules' : modules,
			'sous_modules': sous_modules,
			'module' : vars_module,
			'organisation' : dao_organisation.toGetMainOrganisation(),
		}
		template = loader.get_template('ErpProject/ModuleConfiguration/type_periode/upload.html')
		return HttpResponse(template.render(context, request))
	except Exception as e:
		return auth.toReturnFailed(request, e, traceback.format_exc())

@transaction.atomic
def post_upload_type_periode(request):
	sid = transaction.savepoint()
	try:
		media_dir = settings.MEDIA_ROOT + '/excel/'
		file_name = ''
		randomId = randint(111, 999)
		if 'file_upload' in request.FILES:
			file = request.FILES['file_upload']
			save_path = os.path.join(media_dir, str(randomId) + '.xlsx')
			if default_storage.exists(save_path):
				default_storage.delete(save_path)
			file_name = default_storage.save(save_path, file)
		else: file_name = ''
		sheet = str(request.POST['sheet'])

		df = pd.read_excel(io=save_path, sheet_name=sheet, engine='openpyxl')
		df = df.fillna('') #Replace all nan value

		auteur = identite.utilisateur(request)


		header_name = makeString(request.POST['name'])
		if header_name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Désignation\' est obligatoire, Veuillez le renseigner SVP!')
		#print(f'header_name_id: {header_name_id}')

		header_periodicite = makeString(request.POST['periodicite'])
		#print(f'header_periodicite_id: {header_periodicite_id}')

		header_nombre_par_exercice = makeString(request.POST['nombre_par_exercice'])
		#print(f'header_nombre_par_exercice_id: {header_nombre_par_exercice_id}')

		header_description = makeString(request.POST['description'])
		#print(f'header_description_id: {header_description_id}')

		header_societe_id = makeString(request.POST['societe_id'])
		#print(f'header_societe_id: {header_societe_id}')

		for i in df.index:

			name = ''
			if header_name != '': name = makeString(df[header_name][i])
			if name in (None, '') : return auth.toReturnFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Désignation\' est obligatoire, Veuillez le renseigner SVP!')

			periodicite = ''
			if header_periodicite != '': periodicite = makeString(df[header_periodicite][i])

			nombre_par_exercice = 0
			if header_nombre_par_exercice != '': nombre_par_exercice = makeInt(df[header_nombre_par_exercice][i])

			description = ''
			if header_description != '': description = makeString(df[header_description][i])

			societe_id = None
			if header_societe_id != '': societe_id = makeIntId(str(df[header_societe_id][i]))

			type_periode = dao_type_periode.toCreate(name = name, periodicite = periodicite, nombre_par_exercice = nombre_par_exercice, description = description, societe_id = societe_id)
			saved, type_periode, message = dao_type_periode.toSave(auteur, type_periode)

			if saved == False: raise Exception(message)

		transaction.savepoint_commit(sid)
		messages.add_message(request, messages.SUCCESS, 'Les enregistrements se sont effectué avec succès!')
		return HttpResponseRedirect(reverse('module_configuration_list_type_periode'))
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnFailed(request, e, traceback.format_exc())

# TYPE_PERIODE API CONTROLLERS
def get_list_type_periode(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		filtered = False
		if 'filtered' in request.GET : filtered = str(request.GET['filtered'])
		date_from = None
		if 'date_from' in request.GET : date_from = request.GET['date_from']
		date_to = None
		if 'date_to' in request.GET : date_to = request.GET['date_to']
		query = ''
		if 'query' in request.GET : query = str(request.GET['query'])

		listes = []
		model = dao_type_periode.toList()
		#model = pagination.toGet(request, model)

		for item in model:
			element = {
				'id' : item.id,
				'name' : str(item.name),
				'periodicite' : str(item.periodicite),
				'nombre_par_exercice' : makeInt(item.nombre_par_exercice),
				'description' : str(item.description),
				'statut_id' : makeIntId(item.statut_id),
				'societe_id' : makeIntId(item.societe_id),
				'etat' : str(item.etat),
				'creation_date' : item.creation_date,
				'update_date' : item.update_date,
				'update_by_id' : makeIntId(item.update_by_id),
				'auteur_id' : makeIntId(item.auteur_id),
			}
			listes.append(element)

		context = {
			'error' : False,
			'message' : 'Liste récupérée',
			'datas' : listes
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

def get_item_type_periode(request):
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')

		id = 0
		if 'id' in request.GET : id = int(request.GET['id'])

		item = {}
		model = dao_type_periode.toGet(id)
		if model != None :
			item = {
				'id' : model.id,
				'name' : str(model.name),
				'periodicite' : str(model.periodicite),
				'nombre_par_exercice' : makeInt(model.nombre_par_exercice),
				'description' : str(model.description),
				'statut_id' : makeIntId(model.statut_id),
				'societe_id' : makeIntId(model.societe_id),
				'etat' : str(model.etat),
				'creation_date' : model.creation_date,
				'update_date' : model.update_date,
				'update_by_id' : makeIntId(model.update_by_id),
				'auteur_id' : makeIntId(model.auteur_id),
			}

		context = {
			'error' : False,
			'message' : 'Objet récupéré',
			'item' : item
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		return auth.toReturnApiFailed(request, e, traceback.format_exc())

@api_view(['POST'])
@transaction.atomic
def post_create_type_periode(request):
	sid = transaction.savepoint()
	try:
		context = {}
		#token = request.META.get('HTTP_TOKEN')
		#if not token: raise Exception('Erreur, Token manquant')


		name = ''
		if 'name' in request.POST : name = str(request.POST['name'])
		if name in (None, '') : return auth.toReturnApiFailed(request, 'Champ obligatoire non saisi', '', msg = 'Le Champ \'Désignation\' est obligatoire, Veuillez le renseigner SVP!')

		periodicite = ''
		if 'periodicite' in request.POST : periodicite = str(request.POST['periodicite'])

		nombre_par_exercice = 0
		if 'nombre_par_exercice' in request.POST : nombre_par_exercice = makeInt(request.POST['nombre_par_exercice'])

		description = ''
		if 'description' in request.POST : description = str(request.POST['description'])

		societe_id = None
		if 'societe' in request.POST : societe_id = makeIntId(request.POST['societe_id'])

		auteur_id = None
		if 'auteur' in request.POST : auteur_id = makeIntId(request.POST['auteur_id'])

		auteur = dao_utilisateur.toGetUtilisateur(auteur_id)

		type_periode = dao_type_periode.toCreate(name = name, periodicite = periodicite, nombre_par_exercice = nombre_par_exercice, description = description, societe_id = societe_id)
		saved, type_periode, message = dao_type_periode.toSave(auteur, type_periode)

		if saved == False: raise Exception(message)

		objet = {
			'id' : type_periode.id,
			'name' : str(type_periode.name),
			'periodicite' : str(type_periode.periodicite),
			'nombre_par_exercice' : makeInt(type_periode.nombre_par_exercice),
			'description' : str(type_periode.description),
			'statut_id' : makeIntId(type_periode.statut_id),
			'societe_id' : makeIntId(type_periode.societe_id),
			'etat' : str(type_periode.etat),
			'creation_date' : type_periode.creation_date,
			'update_date' : type_periode.update_date,
			'update_by_id' : makeIntId(type_periode.update_by_id),
			'auteur_id' : makeIntId(type_periode.auteur_id),
		}
		transaction.savepoint_commit(sid)

		context = {
			'error' : False,
			'message' : 'Enregistrement éffectué avec succès',
			'item' : objet
		}
		return JsonResponse(context, safe=False)
	except Exception as e:
		transaction.savepoint_rollback(sid)
		return auth.toReturnApiFailed(request, e, traceback.format_exc())